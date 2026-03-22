from __future__ import annotations

import io
import tempfile
import unittest
import zipfile
from dataclasses import replace
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from services.tax_package import (
    PackageSnapshot,
    PackageStats,
    _classify_reference_material_type,
    _collect_withholding_summary_rows,
    _extend_review_items,
    _reference_transaction_comparison,
    _resolve_reference_material_comparison,
    _source_labels,
    describe_tax_package_profile,
    build_tax_package_zip_from_snapshot,
)


class TaxPackageServiceTest(unittest.TestCase):
    def setUp(self) -> None:
        self.tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmpdir.cleanup)

        self.evidence_path = Path(self.tmpdir.name) / "receipt.pdf"
        self.evidence_path.write_bytes(b"%PDF-1.4\nsample evidence\n")
        self.official_path = Path(self.tmpdir.name) / "hometax_payment.csv"
        self.official_path.write_text("조회일,2026-03-10\n최근 납부일,납부금액 합계,세목\n2026-03-09,150000,종합소득세\n", encoding="utf-8")

        stats = PackageStats(
            month_key="2026-03",
            period_start_kst="2026-03-01",
            period_end_kst="2026-03-31",
            generated_at_kst="2026-03-17 09:30",
            tx_total=2,
            tx_in_count=1,
            tx_out_count=1,
            sum_in_total=3200000,
            sum_out_total=180000,
            income_included_total=3200000,
            income_excluded_non_income_total=0,
            income_unknown_count=0,
            expense_business_total=180000,
            expense_personal_total=0,
            expense_mixed_total=0,
            expense_unknown_total=0,
            evidence_missing_required_count=0,
            evidence_missing_required_amount=0,
            evidence_missing_maybe_count=1,
            evidence_missing_maybe_amount=22000,
            evidence_attached_count=1,
            review_needed_count=1,
            tax_rate=0.15,
            tax_buffer_total=120000,
            tax_buffer_target=480000,
            tax_buffer_shortage=360000,
        )

        self.snapshot = PackageSnapshot(
            root_name="세무사전달패키지_2026-03_테스터",
            download_name="세무사전달패키지_2026-03_테스터.zip",
            display_name="테스터",
            stats=stats,
            transactions=[
                {
                    "tx_id": 101,
                    "occurred_at_kst": "2026-03-03 10:00",
                    "direction_label": "출금",
                    "amount_krw": 180000,
                    "counterparty": "카페 샘플",
                    "memo": "회의비",
                    "source_label": "자동연동",
                    "provider_label": "은행연동",
                    "classification_result_label": "업무지출",
                    "business_related_label": "예",
                    "evidence_status_label": "첨부됨",
                    "representative_evidence_type": "PDF 증빙",
                    "evidence_count": 1,
                    "evidence_zip_path": "attachments/evidence/101_receipt.pdf",
                    "trust_label": "반영됨",
                    "calculation_included_label": "예",
                    "recheck_required_label": "아니오",
                    "recheck_reason": "",
                    "evidence_note": "카드 전표 첨부",
                },
                {
                    "tx_id": 102,
                    "occurred_at_kst": "2026-03-05 16:20",
                    "direction_label": "출금",
                    "amount_krw": 22000,
                    "counterparty": "문구점",
                    "memo": "사무용품",
                    "source_label": "수동입력",
                    "provider_label": "없음",
                    "classification_result_label": "미확정",
                    "business_related_label": "미확정",
                    "evidence_status_label": "확인 필요",
                    "representative_evidence_type": "",
                    "evidence_count": 0,
                    "evidence_zip_path": "",
                    "trust_label": "재확인필요",
                    "calculation_included_label": "보류",
                    "recheck_required_label": "예",
                    "recheck_reason": "지출 분류가 아직 확정되지 않았습니다",
                    "evidence_note": "",
                },
            ],
            evidences=[
                {
                    "증빙번호": 9001,
                    "연결거래번호": 101,
                    "거래일시": "2026-03-03 10:00",
                    "거래처": "카페 샘플",
                    "금액": 180000,
                    "증빙종류": "PDF 증빙",
                    "파일명": "receipt.pdf",
                    "파일열기": ("열기", "attachments/evidence/101_receipt.pdf"),
                    "저장위치": "attachments/evidence/101_receipt.pdf",
                    "업로드일시": "2026-03-03 10:10",
                    "신뢰구분": "반영됨",
                    "계산반영여부": "예",
                    "재확인필요여부": "아니오",
                    "메모": "카드 전표 첨부",
                    "_zip_path": "attachments/evidence/101_receipt.pdf",
                    "_abs_path": self.evidence_path,
                }
            ],
            review_items=[
                {
                    "항목번호": 1,
                    "항목유형": "거래검토",
                    "관련자료구분": "거래내역",
                    "관련번호": 102,
                    "요약설명": "지출 분류가 아직 확정되지 않았습니다",
                    "현재상태": "미확정",
                    "필요한확인내용": "업무/개인/혼합 중 하나로 확정해 주세요",
                    "우선순위": "보통",
                    "메모": "사무용품",
                }
            ],
            evidence_missing_items=[
                {
                    "거래번호": 102,
                    "거래일시": "2026-03-05 16:20",
                    "거래처": "문구점",
                    "금액": 22000,
                    "증빙상태": "확인 필요",
                    "필요한확인내용": "업무 관련이면 증빙을 첨부하고, 아니면 불필요로 표시해 주세요",
                    "우선순위": "보통",
                }
            ],
            review_trade_items=[
                {
                    "거래번호": 102,
                    "거래일시": "2026-03-05 16:20",
                    "자료출처": "수동입력",
                    "거래처": "문구점",
                    "금액": 22000,
                    "현재상태": "미확정",
                    "재확인사유": "업무/개인 판단이 확정되지 않았습니다",
                    "필요한확인내용": "업무/개인/혼합 중 하나로 확정해 주세요",
                }
            ],
            included_source_labels=["수동입력", "자동연동"],
            business_status_rows=[
                {
                    "user_type": "프리랜서(3.3)",
                    "health_insurance_status": "지역가입자",
                    "vat_status": "미확인",
                    "business_registration_status": "미확인",
                    "business_account_usage_status": "미연결",
                    "business_card_usage_status": "미확인",
                    "onboarding_basis": "온보딩 입력값",
                    "note": "건강보험/과세 상태 중 일부 미확인",
                }
            ],
            withholding_summary_rows=[
                {
                    "document_scope": "미확인",
                    "withholding_period_basis": "미확인",
                    "paid_tax_period_basis": "미확인",
                    "has_withholding_data": "아니오",
                    "gross_pay_total_krw": "",
                    "withholding_tax_total_krw": "",
                    "has_paid_tax_data": "아니오",
                    "paid_tax_total_krw": "",
                    "payer_reference": "미확인",
                    "other_income_flag": "예(입력값 기준)",
                    "source_basis": "온보딩 입력값",
                    "needs_review": "예",
                    "note": "대상 월에 포함된 원천징수/기납부세액 공식자료가 없습니다",
                }
            ],
            vat_summary_rows=[
                {
                    "vat_status": "미확인",
                    "recent_vat_filing_status": "자료 없음",
                    "tax_invoice_sales_total_krw": "",
                    "tax_invoice_purchase_total_krw": "",
                    "card_purchase_total_krw": "",
                    "cash_receipt_purchase_total_krw": "",
                    "source_basis": "거래 집계 참고",
                    "needs_review": "예",
                    "note": "과세 상태 또는 부가세 관련 공식자료 요약이 미확인입니다",
                }
            ],
            nhis_pension_summary_rows=[
                {
                    "health_insurance_status": "지역가입자",
                    "period_basis": "2026-03",
                    "nhis_total_krw": "",
                    "has_nhis_data": "아니오",
                    "has_pension_data": "아니오",
                    "pension_total_krw": "",
                    "source_basis": "사용자 입력",
                    "needs_review": "예",
                    "note": "건강보험/국민연금 공식자료 요약이 없습니다",
                }
            ],
        )

        official_stats = replace(
            stats,
            review_needed_count=6,
            official_data_total=2,
            official_data_parsed_count=2,
            official_data_review_count=0,
            official_data_unsupported_count=0,
            official_data_failed_count=0,
        )
        self.snapshot_with_official = PackageSnapshot(
            root_name="세무사전달패키지_2026-03_테스터",
            download_name="세무사전달패키지_2026-03_테스터.zip",
            display_name="테스터",
            stats=official_stats,
            transactions=self.snapshot.transactions,
            evidences=self.snapshot.evidences,
            evidence_missing_items=self.snapshot.evidence_missing_items,
            review_trade_items=self.snapshot.review_trade_items,
            included_source_labels=self.snapshot.included_source_labels,
            review_items=[
                *self.snapshot.review_items,
                {
                    "항목번호": 2,
                    "항목유형": "공식자료재확인",
                    "관련자료구분": "공식자료",
                    "관련번호": 7001,
                    "요약설명": "홈택스 납부내역 재확인 필요",
                    "현재상태": "반영 가능 / 검증 미실시",
                    "필요한확인내용": "공식자료 요약값과 원본 기준이 맞는지 다시 확인해 주세요",
                    "우선순위": "보통",
                    "메모": "검증 미실시",
                },
                {
                    "항목번호": 3,
                    "항목유형": "공식자료교차검증재확인",
                    "관련자료구분": "공식자료",
                    "관련번호": 7001,
                    "요약설명": "공식자료 교차검증 차이 확인 필요",
                    "현재상태": "비교 가능한 거래와 차이 있음",
                    "필요한확인내용": "공식자료 요약과 비교 가능한 거래 사이에 차이가 있어 세무사 재확인이 필요합니다.",
                    "우선순위": "보통",
                    "메모": "홈택스 납부내역 / 같은 성격의 거래 1건이 있지만 금액 또는 날짜가 다릅니다.",
                },
                {
                    "항목번호": 4,
                    "항목유형": "참고자료검토",
                    "관련자료구분": "참고자료",
                    "관련번호": 8101,
                    "요약설명": "참고자료 보조 설명 연결 확인 필요",
                    "현재상태": "참고용",
                    "필요한확인내용": "공식자료와 연결되는 설명인지, 참고용 메모인지 확인해 주세요",
                    "우선순위": "보통",
                    "메모": "수입 구조 메모",
                },
                {
                    "항목번호": 5,
                    "항목유형": "부가세자료누락",
                    "관련자료구분": "공식자료",
                    "관련번호": "",
                    "요약설명": "부가세 신고 자료 추가 확인 필요",
                    "현재상태": "자료 없음",
                    "필요한확인내용": "최근 부가세 신고 자료와 부가세 관련 공식자료를 다시 확인해 주세요",
                    "우선순위": "높음",
                    "메모": "공식자료 요약값 / 사용자 입력",
                },
                {
                    "항목번호": 6,
                    "항목유형": "건보자료누락",
                    "관련자료구분": "공식자료",
                    "관련번호": "",
                    "요약설명": "건강보험 자료 추가 확인 필요",
                    "현재상태": "자료 미첨부",
                    "필요한확인내용": "건강보험 납부 자료 또는 안전한 요약 자료가 있으면 추가 확인해 주세요",
                    "우선순위": "보통",
                    "메모": "공식자료 요약값 / 사용자 입력",
                },
            ],
            official_documents=[
                {
                    "자료번호": 7001,
                    "기관명": "국세청(홈택스)",
                    "문서종류": "홈택스 납부내역",
                    "기준일": "2026-03-10",
                    "원본파일명": "hometax_payment.csv",
                    "읽기상태": "반영 가능",
                    "검증상태": "검증 미실시",
                    "구조확인": "구조 확인됨",
                    "신뢰등급": "구조 확인됨 (B)",
                    "핵심값요약": "기준일: 2026-03-10 / 납부세액 합계: 150,000원",
                    "교차검증상태": "불일치",
                    "교차검증사유": "같은 성격의 거래 1건이 있지만 금액 또는 날짜가 다릅니다.",
                    "교차검증재확인필요여부": "예",
                    "목록반영여부": "예",
                    "원본첨부여부": "아니오",
                    "재확인필요여부": "예",
                    "메모": "검증 미실시 / 원본 파일은 기본 패키지에 포함하지 않습니다",
                    "_cross_validation_status_key": "mismatch",
                    "_attachment_index_key": "official-7001",
                    "_period_basis": "2026-03-10",
                    "_summary_items": [
                        {"label": "기준일", "value": "2026-03-10"},
                        {"label": "납부세액 합계", "value": "150,000원"},
                    ],
                },
                {
                    "자료번호": 7002,
                    "기관명": "국민건강보험공단",
                    "문서종류": "건강보험 자격 관련 문서",
                    "기준일": "2026-03-11",
                    "원본파일명": "nhis_eligibility.pdf",
                    "읽기상태": "반영 가능",
                    "검증상태": "검증 미실시",
                    "구조확인": "구조 확인됨",
                    "신뢰등급": "구조 확인됨 (B)",
                    "핵심값요약": "기준일: 2026-03-11",
                    "교차검증상태": "비교 불가",
                    "교차검증사유": "비교 가능한 공식자료 범위가 아니어서 교차검증 v1 비교를 하지 않았습니다.",
                    "교차검증재확인필요여부": "아니오",
                    "목록반영여부": "예",
                    "원본첨부여부": "아니오",
                    "재확인필요여부": "아니오",
                    "메모": "원본 파일은 기본 패키지에 포함하지 않습니다",
                    "_cross_validation_status_key": "reference_only",
                    "_attachment_index_key": "official-7002",
                    "_period_basis": "2026-03-11",
                    "_summary_items": [
                        {"label": "기준일", "value": "2026-03-11"},
                    ],
                }
            ],
            business_status_rows=self.snapshot.business_status_rows,
            withholding_summary_rows=[
                {
                    "document_scope": "지급명세서 계열 / 홈택스 납부내역",
                    "withholding_period_basis": "2026-03-01 ~ 2026-03-31",
                    "paid_tax_period_basis": "2026-03",
                    "has_withholding_data": "예",
                    "gross_pay_total_krw": 3200000,
                    "withholding_tax_total_krw": 330000,
                    "has_paid_tax_data": "예",
                    "paid_tax_total_krw": 150000,
                    "payer_reference": "샘플지급처",
                    "other_income_flag": "예(입력값 기준)",
                    "source_basis": "공식자료 요약값 / 온보딩 입력값",
                    "needs_review": "아니오",
                    "note": "",
                }
            ],
            vat_summary_rows=[
                {
                    "vat_status": "과세사업자/부가세 대상이에요",
                    "recent_vat_filing_status": "예",
                    "tax_invoice_sales_total_krw": 2100000,
                    "tax_invoice_purchase_total_krw": 560000,
                    "card_purchase_total_krw": 180000,
                    "cash_receipt_purchase_total_krw": 90000,
                    "source_basis": "공식자료 요약값 / 사용자 입력",
                    "needs_review": "아니오",
                    "note": "",
                }
            ],
            nhis_pension_summary_rows=[
                {
                    "health_insurance_status": "지역가입자",
                    "period_basis": "2026-03",
                    "nhis_total_krw": 119000,
                    "has_nhis_data": "예",
                    "has_pension_data": "예",
                    "pension_total_krw": 152000,
                    "source_basis": "공식자료 요약값 / 사용자 입력",
                    "needs_review": "아니오",
                    "note": "",
                }
            ],
            reference_material_rows=[
                {
                    "reference_material_id": 8101,
                    "title": "기납부세액 메모",
                    "reference_type": "참고자료",
                    "reported_period": "2026-03",
                    "reported_amount_krw": 330000,
                    "linked_official_doc_type": "홈택스 납부내역",
                    "link_status": "공식자료 요약과 차이 있음",
                    "link_status_key": "official_difference",
                    "comparison_basis": "공식자료 요약 대비",
                    "comparison_target": "홈택스 납부내역",
                    "difference_krw": 180000,
                    "difference_description": "기재 금액과 연결된 공식자료 요약값 차이를 확인해 주세요",
                    "needs_review": "예",
                    "note": "연결된 공식자료와 금액 차이가 있어 세무사 확인이 필요합니다",
                    "_original_filename": "income_note.pdf",
                    "_attachment_index_key": "reference-8101",
                    "_period_basis": "2026-03",
                },
                {
                    "reference_material_id": 8102,
                    "title": "월 매출 정리 메모",
                    "reference_type": "추가설명",
                    "reported_period": "2026-03",
                    "reported_amount_krw": 3200000,
                    "linked_official_doc_type": "",
                    "link_status": "거래 합계와 대체로 일치",
                    "link_status_key": "transaction_match",
                    "comparison_basis": "거래 합계 대비",
                    "comparison_target": "월간 수입 합계",
                    "difference_krw": 0,
                    "difference_description": "기재 금액과 대상 월 거래 합계 차이가 없습니다",
                    "needs_review": "아니오",
                    "note": "공식자료 대체가 아니라 보조 설명 자료로 전달합니다",
                    "_original_filename": "sales_note.pdf",
                    "_attachment_index_key": "reference-8102",
                    "_period_basis": "2026-03",
                },
            ],
        )

    def _build_zip(
        self,
        snapshot: PackageSnapshot | None = None,
        profile_code: str | None = None,
    ) -> tuple[bytes, zipfile.ZipFile, str]:
        zip_io, filename = build_tax_package_zip_from_snapshot(snapshot or self.snapshot, profile_code=profile_code)
        payload = zip_io.getvalue()
        archive = zipfile.ZipFile(io.BytesIO(payload))
        self.addCleanup(archive.close)
        return payload, archive, filename

    def test_zip_contains_expected_files_and_keeps_summary_outputs_when_no_docs(self) -> None:
        _, archive, filename = self._build_zip()
        names = set(archive.namelist())
        root = "세무사전달패키지_2026-03_테스터"
        self.assertEqual(filename, "세무사전달패키지_2026-03_테스터.zip")

        self.assertIn(f"{root}/00_패키지요약.xlsx", names)
        self.assertIn(f"{root}/01_사업_상태_요약.xlsx", names)
        self.assertIn(f"{root}/03_거래원장.xlsx", names)
        self.assertIn(f"{root}/04_증빙상태표.xlsx", names)
        self.assertIn(f"{root}/05_원천징수_기납부세액_요약.xlsx", names)
        self.assertIn(f"{root}/06_세무사_확인필요목록.xlsx", names)
        self.assertIn(f"{root}/07_첨부인덱스.xlsx", names)
        self.assertIn(f"{root}/08_부가세_자료_요약.xlsx", names)
        self.assertIn(f"{root}/09_건보_연금_요약.xlsx", names)
        self.assertIn(f"{root}/10_참고자료_요약.xlsx", names)
        self.assertIn(f"{root}/attachments/", names)
        self.assertIn(f"{root}/attachments/evidence/", names)
        self.assertIn(f"{root}/attachments/evidence/101_receipt.pdf", names)

        self.assertFalse(any(name.startswith(f"{root}/공식자료/") for name in names))
        self.assertFalse(any(name.startswith(f"{root}/attachments/official/") for name in names))

        summary_wb = load_workbook(io.BytesIO(archive.read(f"{root}/00_패키지요약.xlsx")))
        official_wb = summary_wb
        official_ws = official_wb["공식자료목록"]
        self.assertEqual(official_ws["C2"].value, "현재 포함된 공식자료 없음")

    def test_package_guide_sheet_explains_scope_and_limitations(self) -> None:
        _, archive, _ = self._build_zip()
        wb = load_workbook(io.BytesIO(archive.read("세무사전달패키지_2026-03_테스터/00_패키지요약.xlsx")))
        ws = wb["패키지안내"]
        guide = "\n".join(str(cell.value or "") for cell in ws["A"][1:])

        self.assertIn("01_사업_상태_요약.xlsx", guide)
        self.assertIn("05_원천징수_기납부세액_요약.xlsx", guide)
        self.assertIn("08_부가세_자료_요약.xlsx", guide)
        self.assertIn("09_건보_연금_요약.xlsx", guide)
        self.assertIn("07_첨부인덱스.xlsx", guide)
        self.assertIn("10_참고자료_요약.xlsx", guide)
        self.assertIn("공식자료 원본 파일", guide)
        self.assertIn("참고자료 원본", guide)
        self.assertIn("attachments/evidence/", guide)
        self.assertIn("거래당 대표 증빙 1개 기준", guide)
        self.assertIn("검증상태가 '검증 미실시'", guide)
        self.assertIn("교차검증 v1", guide)
        self.assertIn("비교 가능한 공식자료만", guide)

    def test_summary_workbook_includes_cross_validation_counts_and_note(self) -> None:
        _, archive, _ = self._build_zip(self.snapshot_with_official)
        root = "세무사전달패키지_2026-03_테스터"
        summary_wb = load_workbook(io.BytesIO(archive.read(f"{root}/00_패키지요약.xlsx")))
        summary_ws = summary_wb["패키지요약"]
        top_labels = [summary_ws.cell(row_idx, 1).value for row_idx in range(2, 12)]
        rows = {
            summary_ws.cell(row_idx, 1).value: summary_ws.cell(row_idx, 2).value
            for row_idx in range(2, summary_ws.max_row + 1)
        }

        self.assertEqual(
            top_labels[:8],
            [
                "사용자명",
                "대상 기간",
                "생성일시",
                "검토 시작",
                "1차 검토 흐름",
                "즉시 재확인 항목 수",
                "증빙 확인 필요",
                "원천징수·기납부세액 상태",
            ],
        )
        self.assertEqual(rows["검토 시작"], "00_패키지요약.xlsx → 06_세무사_확인필요목록.xlsx")
        self.assertIn("05_원천징수_기납부세액_요약.xlsx", str(rows["1차 검토 흐름"]))
        self.assertEqual(rows["즉시 재확인 항목 수"], 6)
        self.assertEqual(rows["증빙 확인 필요"], "필수 0건 / 확인 1건")
        self.assertEqual(rows["원천징수·기납부세액 상태"], "원천징수·기납부세액 자료 있음")
        self.assertEqual(rows["부가세 상태"], "부가세 신고 자료 확인됨")
        self.assertEqual(rows["건보·연금 상태"], "건보·연금 자료 있음")
        self.assertEqual(rows["참고자료 검토 상태"], "재확인 필요 1건 / 총 2건")
        self.assertEqual(rows["교차검증 일치 문서 수"], 0)
        self.assertEqual(rows["교차검증 부분 일치 문서 수"], 0)
        self.assertEqual(rows["교차검증 재확인 필요 문서 수"], 0)
        self.assertEqual(rows["교차검증 불일치 문서 수"], 1)
        self.assertEqual(rows["교차검증 비교 불가 문서 수"], 1)
        self.assertIn("교차검증 v1 기준", str(rows["교차검증 안내"]))
        self.assertIn("비교 가능한 공식자료만", str(rows["교차검증 안내"]))

    def test_business_status_and_withholding_workbooks_are_added(self) -> None:
        _, archive, _ = self._build_zip()
        root = "세무사전달패키지_2026-03_테스터"

        business_wb = load_workbook(io.BytesIO(archive.read(f"{root}/01_사업_상태_요약.xlsx")))
        business_ws = business_wb["사업 상태 요약"]
        business_headers = {cell.value: idx + 1 for idx, cell in enumerate(business_ws[1])}
        self.assertEqual(business_ws.cell(2, business_headers["항목명"]).value, "사용자 유형")
        self.assertEqual(business_ws.cell(2, business_headers["현재 값"]).value, "프리랜서(3.3)")
        self.assertEqual(business_ws.cell(2, business_headers["값 출처"]).value, "사용자 입력")
        self.assertEqual(business_ws.cell(2, business_headers["확인 수준"]).value, "참고용")

        withholding_wb = load_workbook(io.BytesIO(archive.read(f"{root}/05_원천징수_기납부세액_요약.xlsx")))
        withholding_ws = withholding_wb["원천징수·기납부세액 요약"]
        withholding_headers = {cell.value: idx + 1 for idx, cell in enumerate(withholding_ws[1])}
        self.assertEqual(withholding_ws.cell(2, withholding_headers["자료 종류"]).value, "미확인")
        self.assertEqual(withholding_ws.cell(2, withholding_headers["원천징수 기준 기간"]).value, "미확인")
        self.assertEqual(withholding_ws.cell(2, withholding_headers["원천징수 자료 있음"]).value, "아니오")
        self.assertEqual(withholding_ws.cell(2, withholding_headers["다른 소득 있음"]).value, "예(입력값 기준)")
        self.assertEqual(withholding_ws.cell(2, withholding_headers["기준 자료"]).value, "온보딩 입력값")
        self.assertEqual(withholding_ws.cell(2, withholding_headers["재확인 필요"]).value, "예")

        vat_wb = load_workbook(io.BytesIO(archive.read(f"{root}/08_부가세_자료_요약.xlsx")))
        vat_ws = vat_wb["부가세 자료 요약"]
        vat_headers = {cell.value: idx + 1 for idx, cell in enumerate(vat_ws[1])}
        self.assertEqual(vat_ws.cell(2, vat_headers["과세 상태"]).value, "미확인")
        self.assertEqual(vat_ws.cell(2, vat_headers["재확인 필요"]).value, "예")

        nhis_wb = load_workbook(io.BytesIO(archive.read(f"{root}/09_건보_연금_요약.xlsx")))
        nhis_ws = nhis_wb["건보·연금 요약"]
        nhis_headers = {cell.value: idx + 1 for idx, cell in enumerate(nhis_ws[1])}
        self.assertEqual(nhis_ws.cell(2, nhis_headers["건강보험 상태"]).value, "지역가입자")
        self.assertEqual(nhis_ws.cell(2, nhis_headers["건강보험 자료 있음"]).value, "아니오")

    def test_reference_material_workbook_and_review_expansion_are_added(self) -> None:
        _, archive, _ = self._build_zip(self.snapshot_with_official)
        root = "세무사전달패키지_2026-03_테스터"

        reference_wb = load_workbook(io.BytesIO(archive.read(f"{root}/10_참고자료_요약.xlsx")))
        reference_ws = reference_wb["참고자료 요약"]
        reference_headers = {cell.value: idx + 1 for idx, cell in enumerate(reference_ws[1])}
        self.assertEqual(reference_ws.cell(2, reference_headers["제목"]).value, "기납부세액 메모")
        self.assertEqual(reference_ws.cell(2, reference_headers["연결 상태"]).value, "공식자료 요약과 차이 있음")
        self.assertEqual(reference_ws.cell(2, reference_headers["비교 기준"]).value, "공식자료 요약 대비")
        self.assertEqual(reference_ws.cell(2, reference_headers["차이 설명"]).value, "기재 금액과 연결된 공식자료 요약값 차이를 확인해 주세요")
        self.assertEqual(reference_ws.cell(2, reference_headers["재확인 필요"]).value, "예")
        self.assertEqual(reference_ws.cell(3, reference_headers["연결 상태"]).value, "거래 합계와 대체로 일치")
        self.assertEqual(reference_ws.cell(3, reference_headers["비교 기준"]).value, "거래 합계 대비")
        self.assertEqual(reference_ws.cell(3, reference_headers["비교 대상"]).value, "월간 수입 합계")

        review_wb = load_workbook(io.BytesIO(archive.read(f"{root}/06_세무사_확인필요목록.xlsx")))
        review_ws = review_wb["세무사_확인필요목록"]
        review_headers = {cell.value: idx + 1 for idx, cell in enumerate(review_ws[1])}
        self.assertEqual(
            [cell.value for cell in review_ws[1][:7]],
            ["항목번호", "우선확인순서", "우선순위", "요약설명", "현재상태", "필요한확인내용", "항목유형"],
        )
        review_types = [review_ws.cell(row_idx, review_headers["항목유형"]).value for row_idx in range(2, review_ws.max_row + 1)]
        self.assertIn("공식자료재확인", review_types)
        self.assertIn("공식자료교차검증재확인", review_types)
        self.assertIn("참고자료검토", review_types)
        self.assertIn("부가세자료누락", review_types)
        self.assertIn("건보자료누락", review_types)
        self.assertEqual(review_ws.cell(2, review_headers["우선확인순서"]).value, 1)
        self.assertEqual(review_ws.cell(2, review_headers["우선순위"]).value, "높음")
        self.assertEqual(review_ws.cell(2, review_headers["항목유형"]).value, "거래검토")
        self.assertEqual(review_ws.cell(3, review_headers["항목유형"]).value, "부가세자료누락")
        self.assertEqual(review_ws.cell(4, review_headers["항목유형"]).value, "공식자료교차검증재확인")
        self.assertIn("세액 영향", str(review_ws.cell(2, review_headers["우선순위 기준"]).value))
        cross_validation_rows = [
            row_idx
            for row_idx in range(2, review_ws.max_row + 1)
            if review_ws.cell(row_idx, review_headers["항목유형"]).value == "공식자료교차검증재확인"
        ]
        self.assertEqual(len(cross_validation_rows), 1)
        cross_validation_row = cross_validation_rows[0]
        self.assertEqual(review_ws.cell(cross_validation_row, review_headers["관련번호"]).value, 7001)
        self.assertEqual(review_ws.cell(cross_validation_row, review_headers["현재상태"]).value, "비교 가능한 거래와 차이 있음")

    def test_workbooks_include_relative_evidence_hyperlinks(self) -> None:
        _, archive, _ = self._build_zip()

        tx_wb = load_workbook(io.BytesIO(archive.read("세무사전달패키지_2026-03_테스터/03_거래원장.xlsx")))
        tx_ws = tx_wb["거래원장"]
        tx_headers = {cell.value: idx + 1 for idx, cell in enumerate(tx_ws[1])}
        tx_link_cell = tx_ws.cell(2, tx_headers["대표첨부열기"])
        self.assertEqual(tx_link_cell.value, "열기")
        self.assertIsNotNone(tx_link_cell.hyperlink)
        self.assertEqual(tx_link_cell.hyperlink.target, "attachments/evidence/101_receipt.pdf")

        evidence_wb = load_workbook(io.BytesIO(archive.read("세무사전달패키지_2026-03_테스터/04_증빙상태표.xlsx")))
        evidence_ws = evidence_wb["증빙상태표"]
        evidence_headers = {cell.value: idx + 1 for idx, cell in enumerate(evidence_ws[1])}
        evidence_link_cell = evidence_ws.cell(2, evidence_headers["첨부열기"])
        self.assertEqual(evidence_link_cell.value, "열기")
        self.assertIsNotNone(evidence_link_cell.hyperlink)
        self.assertEqual(evidence_link_cell.hyperlink.target, "attachments/evidence/101_receipt.pdf")

        attachment_wb = load_workbook(io.BytesIO(archive.read("세무사전달패키지_2026-03_테스터/07_첨부인덱스.xlsx")))
        attachment_ws = attachment_wb["첨부인덱스"]
        attachment_headers = {cell.value: idx + 1 for idx, cell in enumerate(attachment_ws[1])}
        attachment_link_cell = attachment_ws.cell(2, attachment_headers["파일 열기"])
        self.assertEqual(attachment_link_cell.value, "열기")
        self.assertIsNotNone(attachment_link_cell.hyperlink)
        self.assertEqual(attachment_link_cell.hyperlink.target, "attachments/evidence/101_receipt.pdf")

    def test_official_data_is_summarized_without_raw_attachment_when_docs_exist(self) -> None:
        _, archive, _ = self._build_zip(self.snapshot_with_official)
        names = set(archive.namelist())
        root = "세무사전달패키지_2026-03_테스터"

        self.assertNotIn(f"{root}/공식자료/홈택스_납부내역_2026-03-10.csv", names)
        self.assertIn(f"{root}/00_패키지요약.xlsx", names)
        self.assertIn(f"{root}/07_첨부인덱스.xlsx", names)

        official_wb = load_workbook(io.BytesIO(archive.read(f"{root}/00_패키지요약.xlsx")))
        official_ws = official_wb["공식자료목록"]
        headers = {cell.value: idx + 1 for idx, cell in enumerate(official_ws[1])}
        self.assertEqual(official_ws.cell(2, headers["원본첨부여부"]).value, "아니오")
        self.assertEqual(official_ws.cell(2, headers["목록반영여부"]).value, "예")

        withholding_ws = load_workbook(io.BytesIO(archive.read(f"{root}/05_원천징수_기납부세액_요약.xlsx")))["원천징수·기납부세액 요약"]
        withholding_headers = {cell.value: idx + 1 for idx, cell in enumerate(withholding_ws[1])}
        self.assertEqual(withholding_ws.cell(2, withholding_headers["자료 종류"]).value, "지급명세서 계열 / 홈택스 납부내역")
        self.assertEqual(withholding_ws.cell(2, withholding_headers["원천징수 기준 기간"]).value, "2026-03-01 ~ 2026-03-31")
        self.assertEqual(withholding_ws.cell(2, withholding_headers["기납부세액 기준 기간"]).value, "2026-03")
        self.assertEqual(withholding_ws.cell(2, withholding_headers["총지급액 합계"]).value, 3200000)
        self.assertEqual(withholding_ws.cell(2, withholding_headers["지급처 참조"]).value, "샘플지급처")
        self.assertEqual(withholding_ws.cell(2, withholding_headers["재확인 필요"]).value, "아니오")

        vat_ws = load_workbook(io.BytesIO(archive.read(f"{root}/08_부가세_자료_요약.xlsx")))["부가세 자료 요약"]
        vat_headers = {cell.value: idx + 1 for idx, cell in enumerate(vat_ws[1])}
        self.assertEqual(vat_ws.cell(2, vat_headers["과세 상태"]).value, "과세사업자/부가세 대상이에요")
        self.assertEqual(vat_ws.cell(2, vat_headers["최근 부가세 신고 여부"]).value, "예")
        self.assertEqual(vat_ws.cell(2, vat_headers["세금계산서 매출 합계"]).value, 2100000)

        nhis_ws = load_workbook(io.BytesIO(archive.read(f"{root}/09_건보_연금_요약.xlsx")))["건보·연금 요약"]
        nhis_headers = {cell.value: idx + 1 for idx, cell in enumerate(nhis_ws[1])}
        self.assertEqual(nhis_ws.cell(2, nhis_headers["건강보험 자료 있음"]).value, "예")
        self.assertEqual(nhis_ws.cell(2, nhis_headers["국민연금 납부 자료 있음"]).value, "예")

        summary_ws = official_wb["공식자료상태요약"]
        summary_headers = {cell.value: idx + 1 for idx, cell in enumerate(summary_ws[1])}
        summary_rows = {
            summary_ws.cell(row_idx, summary_headers["문서종류"]).value: {
                header: summary_ws.cell(row_idx, col_idx).value
                for header, col_idx in summary_headers.items()
            }
            for row_idx in range(2, summary_ws.max_row + 1)
        }
        self.assertEqual(summary_rows["홈택스 납부내역"]["읽기 가능 건수"], 1)
        self.assertEqual(summary_rows["홈택스 납부내역"]["교차검증 불일치 건수"], 1)
        self.assertEqual(summary_rows["건강보험 자격 관련 문서"]["교차검증 비교 불가 건수"], 1)

        key_ws = official_wb["공식자료핵심값"]
        self.assertEqual(key_ws["A2"].value, 7001)
        self.assertEqual(key_ws["D2"].value, "기준일")

        attachment_ws = load_workbook(io.BytesIO(archive.read(f"{root}/07_첨부인덱스.xlsx")))["첨부인덱스"]
        attachment_headers = {cell.value: idx + 1 for idx, cell in enumerate(attachment_ws[1])}
        attachment_rows = [
            {
                header: attachment_ws.cell(row_idx, col_idx).value
                for header, col_idx in attachment_headers.items()
            }
            for row_idx in range(2, attachment_ws.max_row + 1)
        ]
        official_attachment = next(row for row in attachment_rows if row["자료 유형"] == "홈택스 납부내역 원본")
        reference_attachment = next(row for row in attachment_rows if row["자료 유형"] == "참고자료 원본")
        self.assertEqual(official_attachment["패키지 포함 상태"], "기본 제외")
        self.assertEqual(official_attachment["상대경로"] or "", "")
        self.assertEqual(reference_attachment["패키지 포함 상태"], "기본 제외")

    def test_official_data_sheet_includes_cross_validation_status_reason_and_non_comparable_handling(self) -> None:
        _, archive, _ = self._build_zip(self.snapshot_with_official)
        root = "세무사전달패키지_2026-03_테스터"

        official_wb = load_workbook(io.BytesIO(archive.read(f"{root}/00_패키지요약.xlsx")))
        official_ws = official_wb["공식자료목록"]
        headers = {cell.value: idx + 1 for idx, cell in enumerate(official_ws[1])}

        self.assertEqual(official_ws.cell(2, headers["교차검증 상태"]).value, "불일치")
        self.assertIn("같은 성격의 거래", str(official_ws.cell(2, headers["교차검증 사유"]).value))
        self.assertEqual(official_ws.cell(2, headers["교차검증 재확인 필요"]).value, "예")

        self.assertEqual(official_ws.cell(3, headers["교차검증 상태"]).value, "비교 불가")
        self.assertIn("비교 가능한 공식자료 범위가 아니어서", str(official_ws.cell(3, headers["교차검증 사유"]).value))
        self.assertEqual(official_ws.cell(3, headers["교차검증 재확인 필요"]).value, "아니오")

    def test_reference_and_attachment_sheets_follow_review_friendly_order(self) -> None:
        reordered_snapshot = replace(
            self.snapshot_with_official,
            reference_material_rows=list(reversed(self.snapshot_with_official.reference_material_rows)),
        )
        _, archive, _ = self._build_zip(reordered_snapshot)
        root = "세무사전달패키지_2026-03_테스터"

        reference_wb = load_workbook(io.BytesIO(archive.read(f"{root}/10_참고자료_요약.xlsx")))
        reference_ws = reference_wb["참고자료 요약"]
        reference_headers = {cell.value: idx + 1 for idx, cell in enumerate(reference_ws[1])}
        self.assertEqual(reference_ws.cell(2, reference_headers["연결 상태"]).value, "공식자료 요약과 차이 있음")
        self.assertEqual(reference_ws.cell(3, reference_headers["연결 상태"]).value, "거래 합계와 대체로 일치")

        attachment_wb = load_workbook(io.BytesIO(archive.read(f"{root}/07_첨부인덱스.xlsx")))
        attachment_ws = attachment_wb["첨부인덱스"]
        attachment_headers = {cell.value: idx + 1 for idx, cell in enumerate(attachment_ws[1])}
        self.assertEqual(attachment_ws.cell(2, attachment_headers["패키지 포함 상태"]).value, "포함")
        self.assertEqual(attachment_ws.cell(3, attachment_headers["패키지 포함 상태"]).value, "기본 제외")

    def test_collect_withholding_summary_rows_extracts_period_scope_and_payer_reference(self) -> None:
        documents = [
            {
                "문서종류": "홈택스 원천징수 관련 문서",
                "기준일": "2026-03-31",
                "_period_basis": "2026-03-31",
                "_summary_items": [
                    {"label": "원천징수세액 합계", "value": "330,000원"},
                    {"label": "총지급액 합계", "value": "3,200,000원"},
                ],
                "_summary_values": {
                    "withholding_material_kind": "지급명세서 계열",
                    "period_start": "2026-03-01",
                    "period_end": "2026-03-31",
                    "payer_reference": "샘플지급처",
                    "gross_pay_total_krw": 3200000,
                    "withheld_tax_total_krw": 330000,
                },
            },
            {
                "문서종류": "홈택스 납부내역",
                "기준일": "2026-03-10",
                "_period_basis": "2026-03",
                "_summary_items": [
                    {"label": "납부세액 합계", "value": "150,000원"},
                ],
                "_summary_values": {
                    "paid_tax_total_krw": 150000,
                    "period_summary": "2026-03",
                },
            },
        ]

        with patch(
            "services.tax_package.build_onboarding_reflection",
            return_value={
                "is_freelancer": True,
                "is_employee_sidejob": False,
                "has_specific_user_type": True,
            },
        ):
            rows = _collect_withholding_summary_rows(7, documents)

        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["document_scope"], "지급명세서 계열 / 홈택스 납부내역")
        self.assertEqual(row["withholding_period_basis"], "2026-03-01 ~ 2026-03-31")
        self.assertEqual(row["paid_tax_period_basis"], "2026-03")
        self.assertEqual(row["gross_pay_total_krw"], 3200000)
        self.assertEqual(row["withholding_tax_total_krw"], 330000)
        self.assertEqual(row["paid_tax_total_krw"], 150000)
        self.assertEqual(row["payer_reference"], "샘플지급처")
        self.assertEqual(row["needs_review"], "아니오")

    def test_extend_review_items_adds_granular_withholding_reasons(self) -> None:
        rows = _extend_review_items(
            review_items=[],
            official_documents=[],
            business_status_rows=[
                {
                    "user_type": "프리랜서(3.3)",
                    "health_insurance_status": "지역가입자",
                    "vat_status": "과세사업자/부가세 대상이에요",
                    "note": "",
                }
            ],
            withholding_summary_rows=[
                {
                    "has_withholding_data": "예",
                    "has_paid_tax_data": "예",
                    "other_income_flag": "예(입력값 기준)",
                    "document_scope": "지급명세서 계열 / 홈택스 납부내역",
                    "withholding_period_basis": "미확인",
                    "paid_tax_period_basis": "2026-03 외 1건",
                    "gross_pay_total_krw": "",
                    "withholding_tax_total_krw": "",
                    "paid_tax_total_krw": "",
                    "payer_reference": "미확인",
                    "source_basis": "공식자료 요약값 / 온보딩 입력값",
                    "note": "총지급액 합계가 현재 구조화되지 않아 합계를 확정하지 못했습니다",
                }
            ],
            vat_summary_rows=[],
            nhis_pension_summary_rows=[],
            reference_material_rows=[],
        )
        rows_by_type = {row["항목유형"]: row for row in rows}

        self.assertIn("원천징수기준기간확인", rows_by_type)
        self.assertIn("원천징수지급처참조미확인", rows_by_type)
        self.assertIn("총지급액추출불가", rows_by_type)
        self.assertIn("원천징수세액추출불가", rows_by_type)
        self.assertIn("기납부세액추출불가", rows_by_type)
        self.assertEqual(rows_by_type["원천징수기준기간확인"]["현재상태"], "원천징수: 미확인 / 기납부세액: 2026-03 외 1건")
        self.assertEqual(rows_by_type["원천징수지급처참조미확인"]["현재상태"], "미확인")
        self.assertEqual(rows_by_type["총지급액추출불가"]["현재상태"], "총지급액 추출 불가")
        self.assertEqual(rows_by_type["원천징수세액추출불가"]["현재상태"], "원천징수세액 추출 불가")
        self.assertEqual(rows_by_type["기납부세액추출불가"]["현재상태"], "기납부세액 추출 불가")

    def test_official_and_evidence_sheets_apply_review_friendly_sorting(self) -> None:
        reordered_snapshot = replace(
            self.snapshot_with_official,
            official_documents=list(reversed(self.snapshot_with_official.official_documents)),
            transactions=list(reversed(self.snapshot_with_official.transactions)),
        )
        _, archive, _ = self._build_zip(reordered_snapshot)
        root = "세무사전달패키지_2026-03_테스터"

        summary_wb = load_workbook(io.BytesIO(archive.read(f"{root}/00_패키지요약.xlsx")))
        official_ws = summary_wb["공식자료목록"]
        official_headers = {cell.value: idx + 1 for idx, cell in enumerate(official_ws[1])}
        self.assertEqual(official_ws.cell(2, official_headers["교차검증 상태"]).value, "불일치")
        self.assertEqual(official_ws.cell(3, official_headers["교차검증 상태"]).value, "비교 불가")

        evidence_wb = load_workbook(io.BytesIO(archive.read(f"{root}/04_증빙상태표.xlsx")))
        linked_ws = evidence_wb["거래별대표첨부"]
        linked_headers = {cell.value: idx + 1 for idx, cell in enumerate(linked_ws[1])}
        self.assertEqual(linked_ws.cell(2, linked_headers["거래번호"]).value, 102)
        self.assertEqual(linked_ws.cell(2, linked_headers["증빙상태"]).value, "확인 필요")

    def test_profile_description_exposes_common_income_vat_and_nhis_variants(self) -> None:
        common = describe_tax_package_profile("common")
        income = describe_tax_package_profile("comprehensive_income")
        vat = describe_tax_package_profile("vat_review")
        nhis = describe_tax_package_profile("nhis_pension_check")

        self.assertEqual(common["display_name"], "공통형")
        self.assertEqual(income["display_name"], "종합소득세용")
        self.assertEqual(vat["display_name"], "부가세용")
        self.assertEqual(nhis["display_name"], "건보·연금 점검용")
        self.assertEqual(common["selection_description"], "전체 자료를 종합적으로 점검하는 기본 패키지")
        self.assertEqual(income["selection_description"], "원천징수·기납부세액·거래/증빙 검토를 우선 보는 패키지")
        self.assertEqual(vat["selection_description"], "부가세 관련 자료와 재확인 항목을 우선 보는 패키지")
        self.assertEqual(nhis["selection_description"], "건강보험·국민연금 관련 자료와 재확인 포인트를 먼저 보는 패키지")
        self.assertEqual(common["included_workbooks"][0]["filename"], "00_패키지요약.xlsx")
        self.assertEqual(income["included_workbooks"][3]["filename"], "05_원천징수_기납부세액_요약.xlsx")
        self.assertEqual(vat["included_workbooks"][2]["filename"], "08_부가세_자료_요약.xlsx")
        self.assertEqual(nhis["included_workbooks"][2]["filename"], "09_건보_연금_요약.xlsx")
        self.assertNotIn("09_건보_연금_요약.xlsx", [item["filename"] for item in vat["included_workbooks"]])
        self.assertNotIn("05_원천징수_기납부세액_요약.xlsx", [item["filename"] for item in nhis["included_workbooks"]])
        self.assertNotIn("08_부가세_자료_요약.xlsx", [item["filename"] for item in nhis["included_workbooks"]])

    def test_comprehensive_income_profile_changes_filename_and_summary_focus(self) -> None:
        _, archive, filename = self._build_zip(self.snapshot_with_official, profile_code="comprehensive_income")
        root = "세무사전달패키지_종합소득세용_2026-03_테스터"

        self.assertEqual(filename, f"{root}.zip")
        summary_wb = load_workbook(io.BytesIO(archive.read(f"{root}/00_패키지요약.xlsx")))
        summary_ws = summary_wb["패키지요약"]
        top_labels = [summary_ws.cell(row_idx, 1).value for row_idx in range(2, 11)]
        rows = {
            summary_ws.cell(row_idx, 1).value: summary_ws.cell(row_idx, 2).value
            for row_idx in range(2, summary_ws.max_row + 1)
        }

        self.assertEqual(
            top_labels[:9],
            [
                "사용자명",
                "대상 기간",
                "생성일시",
                "검토 시작",
                "1차 검토 흐름",
                "즉시 재확인 항목 수",
                "원천징수·기납부세액 상태",
                "공식자료 교차검증 상태",
                "증빙 확인 필요",
            ],
        )
        self.assertIn("05_원천징수_기납부세액_요약.xlsx", str(rows["1차 검토 흐름"]))
        self.assertIn("03_거래원장.xlsx", str(rows["1차 검토 흐름"]))

        review_ws = load_workbook(io.BytesIO(archive.read(f"{root}/06_세무사_확인필요목록.xlsx")))["세무사_확인필요목록"]
        review_headers = {cell.value: idx + 1 for idx, cell in enumerate(review_ws[1])}
        ordered_types = [review_ws.cell(row_idx, review_headers["항목유형"]).value for row_idx in range(2, 6)]
        self.assertEqual(ordered_types[:4], ["거래검토", "부가세자료누락", "공식자료교차검증재확인", "공식자료재확인"])

    def test_vat_review_profile_changes_filename_summary_focus_and_review_order(self) -> None:
        _, archive, filename = self._build_zip(self.snapshot_with_official, profile_code="vat_review")
        root = "세무사전달패키지_부가세용_2026-03_테스터"

        self.assertEqual(filename, f"{root}.zip")
        self.assertNotIn(f"{root}/09_건보_연금_요약.xlsx", set(archive.namelist()))
        summary_wb = load_workbook(io.BytesIO(archive.read(f"{root}/00_패키지요약.xlsx")))
        summary_ws = summary_wb["패키지요약"]
        top_labels = [summary_ws.cell(row_idx, 1).value for row_idx in range(2, 11)]
        rows = {
            summary_ws.cell(row_idx, 1).value: summary_ws.cell(row_idx, 2).value
            for row_idx in range(2, summary_ws.max_row + 1)
        }

        self.assertEqual(
            top_labels[:9],
            [
                "사용자명",
                "대상 기간",
                "생성일시",
                "검토 시작",
                "1차 검토 흐름",
                "부가세 상태",
                "부가세 재확인 항목 수",
                "세금계산서/매입자료 요약",
                "즉시 재확인 항목 수",
            ],
        )
        self.assertIn("08_부가세_자료_요약.xlsx", str(rows["1차 검토 흐름"]))
        self.assertIn("03_거래원장.xlsx", str(rows["1차 검토 흐름"]))

        review_ws = load_workbook(io.BytesIO(archive.read(f"{root}/06_세무사_확인필요목록.xlsx")))["세무사_확인필요목록"]
        review_headers = {cell.value: idx + 1 for idx, cell in enumerate(review_ws[1])}
        ordered_types = [review_ws.cell(row_idx, review_headers["항목유형"]).value for row_idx in range(2, 6)]
        self.assertEqual(ordered_types[:4], ["부가세자료누락", "거래검토", "공식자료교차검증재확인", "공식자료재확인"])

    def test_nhis_pension_check_profile_changes_filename_summary_focus_and_review_order(self) -> None:
        nhis_snapshot = replace(
            self.snapshot_with_official,
            review_items=[
                *self.snapshot_with_official.review_items,
                {
                    "항목번호": 7,
                    "항목유형": "연금자료누락",
                    "관련자료구분": "공식자료",
                    "관련번호": "",
                    "요약설명": "국민연금 자료 추가 확인 필요",
                    "현재상태": "자료 미첨부",
                    "필요한확인내용": "국민연금 납부 자료 또는 안전한 요약 자료가 있으면 추가 확인해 주세요",
                    "우선순위": "보통",
                    "메모": "공식자료 요약값 / 사용자 입력",
                },
                {
                    "항목번호": 8,
                    "항목유형": "건보연금상태확인",
                    "관련자료구분": "사용자 상태",
                    "관련번호": "",
                    "요약설명": "건강보험·국민연금 상태 추가 확인 필요",
                    "현재상태": "미확인",
                    "필요한확인내용": "현재 건강보험 상태와 국민연금 납부 여부를 다시 확인해 주세요",
                    "우선순위": "보통",
                    "메모": "사용자 입력 / 공식자료 요약",
                },
            ],
        )
        _, archive, filename = self._build_zip(nhis_snapshot, profile_code="nhis_pension_check")
        root = "세무사전달패키지_건보연금점검용_2026-03_테스터"

        self.assertEqual(filename, f"{root}.zip")
        names = set(archive.namelist())
        self.assertNotIn(f"{root}/05_원천징수_기납부세액_요약.xlsx", names)
        self.assertNotIn(f"{root}/08_부가세_자료_요약.xlsx", names)
        summary_wb = load_workbook(io.BytesIO(archive.read(f"{root}/00_패키지요약.xlsx")))
        summary_ws = summary_wb["패키지요약"]
        top_labels = [summary_ws.cell(row_idx, 1).value for row_idx in range(2, 11)]
        rows = {
            summary_ws.cell(row_idx, 1).value: summary_ws.cell(row_idx, 2).value
            for row_idx in range(2, summary_ws.max_row + 1)
        }

        self.assertEqual(
            top_labels[:9],
            [
                "사용자명",
                "대상 기간",
                "생성일시",
                "검토 시작",
                "1차 검토 흐름",
                "건보·연금 상태",
                "즉시 재확인 항목 수",
                "공식자료 교차검증 상태",
                "증빙 확인 필요",
            ],
        )
        self.assertIn("09_건보_연금_요약.xlsx", str(rows["1차 검토 흐름"]))
        self.assertIn("01_사업_상태_요약.xlsx", str(rows["1차 검토 흐름"]))

        review_ws = load_workbook(io.BytesIO(archive.read(f"{root}/06_세무사_확인필요목록.xlsx")))["세무사_확인필요목록"]
        review_headers = {cell.value: idx + 1 for idx, cell in enumerate(review_ws[1])}
        ordered_types = [review_ws.cell(row_idx, review_headers["항목유형"]).value for row_idx in range(2, 7)]
        self.assertEqual(
            ordered_types[:5],
            ["건보자료누락", "연금자료누락", "건보연금상태확인", "공식자료교차검증재확인", "공식자료재확인"],
        )

    def test_source_labels_support_new_bank_sync_provider_shape(self) -> None:
        self.assertEqual(_source_labels("bank_sync", "popbill"), ("자동연동", "팝빌"))

    def test_source_labels_keep_legacy_popbill_rows_compatible(self) -> None:
        self.assertEqual(_source_labels("popbill", None), ("자동연동", "팝빌"))

    def test_reference_material_type_rules_classify_known_patterns(self) -> None:
        self.assertEqual(_classify_reference_material_type("2025년 연 수익 정리", "", ""), "연 수익표")
        self.assertEqual(_classify_reference_material_type("3월 월 수익표", "", ""), "월 수익표")
        self.assertEqual(_classify_reference_material_type("비용 정리 메모", "", ""), "비용 정리표")
        self.assertEqual(_classify_reference_material_type("추가 설명", "비고 메모", "추가설명"), "설명 메모")
        self.assertEqual(_classify_reference_material_type("기타 전달자료", "", ""), "기타")

    def test_reference_transaction_comparison_supports_income_keywords(self) -> None:
        basis, target, total = _reference_transaction_comparison(
            "3월 월 수익표",
            [
                {"direction": "in", "amount_krw": 3200000},
                {"direction": "out", "amount_krw": 180000},
            ],
        )
        self.assertEqual(basis, "거래 합계 대비")
        self.assertEqual(target, "월간 수입 합계")
        self.assertEqual(total, 3200000)

    def test_reference_material_comparison_prefers_official_summary_over_transaction_total(self) -> None:
        result = _resolve_reference_material_comparison(
            reference_type="설명 메모",
            reported_period="2026-03",
            reported_amount=152000,
            linked_official_doc_type="홈택스 납부내역",
            official_documents=self.snapshot_with_official.official_documents,
            transaction_basis="거래 합계 대비",
            transaction_target="월간 수입 합계",
            transaction_total=152000,
            package_month_key="2026-03",
        )
        self.assertEqual(result["comparison_basis"], "공식자료 요약 대비")
        self.assertEqual(result["comparison_target"], "홈택스 납부내역")
        self.assertEqual(result["link_status"], "공식자료 요약과 대체로 일치")
        self.assertEqual(result["needs_review"], "아니오")

    def test_reference_material_comparison_uses_transaction_total_as_secondary_basis(self) -> None:
        result = _resolve_reference_material_comparison(
            reference_type="월 수익표",
            reported_period="2026-03",
            reported_amount=3200000,
            linked_official_doc_type="",
            official_documents=self.snapshot_with_official.official_documents,
            transaction_basis="거래 합계 대비",
            transaction_target="월간 수입 합계",
            transaction_total=3180000,
            package_month_key="2026-03",
        )
        self.assertEqual(result["comparison_basis"], "거래 합계 대비")
        self.assertEqual(result["comparison_target"], "월간 수입 합계")
        self.assertEqual(result["link_status"], "거래 합계와 대체로 일치")
        self.assertEqual(result["needs_review"], "아니오")

    def test_reference_material_comparison_keeps_annual_or_weak_links_conservative(self) -> None:
        annual_result = _resolve_reference_material_comparison(
            reference_type="연 수익표",
            reported_period="2026",
            reported_amount=12000000,
            linked_official_doc_type="",
            official_documents=[],
            transaction_basis="거래 합계 대비",
            transaction_target="월간 수입 합계",
            transaction_total=3200000,
            package_month_key="2026-03",
        )
        self.assertEqual(annual_result["link_status"], "비교 기준 없음")
        self.assertEqual(annual_result["comparison_basis"], "비교 기준 없음")
        self.assertEqual(annual_result["needs_review"], "예")
        self.assertIn("연간 기준 참고자료", str(annual_result["difference_description"]))

        weak_result = _resolve_reference_material_comparison(
            reference_type="설명 메모",
            reported_period="2026-03",
            reported_amount="",
            linked_official_doc_type="홈택스 납부내역",
            official_documents=self.snapshot_with_official.official_documents,
            transaction_basis="거래 합계 대비",
            transaction_target="월간 수입 합계",
            transaction_total=3200000,
            package_month_key="2026-03",
        )
        self.assertEqual(weak_result["link_status"], "참고용")
        self.assertEqual(weak_result["comparison_basis"], "공식자료 요약 대비")
        self.assertEqual(weak_result["needs_review"], "예")

    def test_reference_material_review_items_expand_for_difference_and_missing_basis(self) -> None:
        review_items = _extend_review_items(
            review_items=[],
            official_documents=[],
            business_status_rows=[
                {
                    "user_type": "프리랜서(3.3)",
                    "health_insurance_status": "지역가입자",
                    "vat_status": "과세사업자/부가세 대상이에요",
                }
            ],
            withholding_summary_rows=[],
            vat_summary_rows=[],
            nhis_pension_summary_rows=[],
            reference_material_rows=[
                {
                    "reference_material_id": 8101,
                    "reference_type": "월 수익표",
                    "needs_review": "예",
                    "link_status_key": "official_difference",
                    "link_status": "공식자료 요약과 차이 있음",
                    "comparison_target": "홈택스 납부내역",
                    "difference_description": "기재 금액과 연결된 공식자료 요약값 차이를 확인해 주세요",
                },
                {
                    "reference_material_id": 8102,
                    "reference_type": "연 수익표",
                    "needs_review": "예",
                    "link_status_key": "no_comparison",
                    "link_status": "비교 기준 없음",
                    "difference_description": "연간 기준 참고자료라 대상 월 패키지와 직접 비교하지 않았습니다",
                },
            ],
        )
        by_id = {row["관련번호"]: row for row in review_items if row["항목유형"] == "참고자료검토"}
        self.assertEqual(by_id[8101]["요약설명"], "참고자료와 공식자료 금액 차이 확인 필요")
        self.assertEqual(by_id[8101]["우선확인순서"], 5)
        self.assertEqual(by_id[8101]["우선순위"], "낮음")
        self.assertIn("공식자료", str(by_id[8101]["메모"]))
        self.assertEqual(by_id[8102]["요약설명"], "참고자료 비교 기준 확인 필요")
        self.assertEqual(by_id[8102]["현재상태"], "비교 기준 없음")
        self.assertIn("연간 기준 참고자료", str(by_id[8102]["메모"]))

    def test_extend_review_items_adds_official_cross_validation_review_rows_conservatively(self) -> None:
        review_items = _extend_review_items(
            review_items=[],
            official_documents=[
                {
                    "자료번호": 7001,
                    "문서종류": "홈택스 납부내역",
                    "교차검증재확인필요여부": "예",
                    "교차검증사유": "같은 성격의 거래 1건이 있지만 금액 또는 날짜가 다릅니다.",
                    "_cross_validation_status_key": "mismatch",
                    "재확인필요여부": "아니오",
                },
                {
                    "자료번호": 7002,
                    "문서종류": "건강보험 납부확인서",
                    "교차검증재확인필요여부": "예",
                    "교차검증사유": "비교 가능한 거래나 참고자료가 부족해 재확인이 필요합니다.",
                    "_cross_validation_status_key": "review_needed",
                    "재확인필요여부": "아니오",
                },
                {
                    "자료번호": 7003,
                    "문서종류": "건강보험 자격 관련 문서",
                    "교차검증재확인필요여부": "아니오",
                    "교차검증사유": "비교 가능한 공식자료 범위가 아니어서 교차검증 v1 비교를 하지 않았습니다.",
                    "_cross_validation_status_key": "reference_only",
                    "재확인필요여부": "아니오",
                },
            ],
            business_status_rows=[
                {
                    "user_type": "프리랜서(3.3)",
                    "health_insurance_status": "지역가입자",
                    "vat_status": "과세사업자/부가세 대상이에요",
                }
            ],
            withholding_summary_rows=[],
            vat_summary_rows=[],
            nhis_pension_summary_rows=[],
            reference_material_rows=[],
        )
        cross_validation_rows = [row for row in review_items if row["항목유형"] == "공식자료교차검증재확인"]
        self.assertEqual(len(cross_validation_rows), 2)
        by_id = {row["관련번호"]: row for row in cross_validation_rows}
        self.assertEqual(by_id[7001]["현재상태"], "비교 가능한 거래와 차이 있음")
        self.assertEqual(by_id[7001]["우선확인순서"], 3)
        self.assertEqual(by_id[7002]["현재상태"], "비교 가능한 거래 없음")
        self.assertNotIn(7003, by_id)
