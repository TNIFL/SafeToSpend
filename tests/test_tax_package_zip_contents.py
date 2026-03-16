from __future__ import annotations

import io
import re
import tempfile
import unittest
import zipfile
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import load_workbook

from domain.models import EvidenceItem, ExpenseLabel, IncomeLabel, Transaction
from services.tax_package import _build_attachment_zip_path, build_tax_package_zip


class _RowsQuery:
    def __init__(self, rows):
        self._rows = rows

    def outerjoin(self, *args, **kwargs):
        return self

    def filter(self, *args, **kwargs):
        return self

    def order_by(self, *args, **kwargs):
        return self

    def all(self):
        return list(self._rows)


class TaxPackageZipContentsTest(unittest.TestCase):
    def _build_zip(self) -> tuple[bytes, str]:
        tx = SimpleNamespace(
            id=101,
            occurred_at=datetime(2026, 3, 14, 13, 24, 55),
            direction="out",
            amount_krw=15_800,
            bank_account_id=None,
            counterparty="스타벅스",
            memo="아메리카노",
            source="csv",
            external_hash="txhash-101",
        )
        ev = SimpleNamespace(
            requirement="required",
            status="attached",
            note="",
            file_key="evidence/key",
            original_filename="receipt.jpg",
            mime_type="image/jpeg",
            size_bytes=1234,
            sha256="abcd",
            uploaded_at=datetime(2026, 3, 14, 14, 0, 0),
            retention_until=None,
            deleted_at=None,
        )

        with tempfile.TemporaryDirectory() as td:
            abs_path = Path(td) / "receipt.jpg"
            abs_path.write_bytes(b"fake-jpg-bytes")

            rows = [(tx, None, None, ev)]

            def _query_side_effect(*entities):
                if entities == (Transaction, IncomeLabel, ExpenseLabel, EvidenceItem):
                    return _RowsQuery(rows)
                raise AssertionError(f"unexpected query entities: {entities!r}")

            tax_est = SimpleNamespace(
                tax_rate=0.15,
                buffer_total_krw=100_000,
                buffer_target_krw=200_000,
                buffer_shortage_krw=100_000,
                tax_due_est_krw=200_000,
                tax_calculation_mode="official_exact",
                official_calculable=True,
                is_limited_estimate=False,
                official_block_reason="",
                taxable_income_input_source="profile_taxable_income",
            )

            with (
                patch("services.tax_package.ensure_can_download_package"),
                patch("services.tax_package._ensure_month_evidence_items"),
                patch("services.tax_package.tax_profile_summary", return_value={}),
                patch("services.tax_package.resolve_file_path", return_value=abs_path),
                patch("services.tax_package.db.session.query", side_effect=_query_side_effect),
                patch("services.risk.compute_tax_estimate", return_value=tax_est),
            ):
                zip_io, filename = build_tax_package_zip(user_pk=1, month_key="2026-03")

        return zip_io.getvalue(), filename

    def test_zip_attachments_use_human_readable_rule_name(self) -> None:
        zip_bytes, filename = self._build_zip()
        self.assertTrue(filename.startswith("SafeToSpend_TaxPackage_2026-03"))

        with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as zf:
            names = zf.namelist()
        attachment_paths = [n for n in names if "/03_증빙첨부(attachments)/attachments/" in n and not n.endswith("/")]
        self.assertEqual(len(attachment_paths), 1)
        self.assertRegex(
            attachment_paths[0],
            r"20260314_132455_15800원_스타벅스_영수증_001\.jpg$",
        )

    def test_attachment_index_paths_match_zip_internal_file(self) -> None:
        zip_bytes, _filename = self._build_zip()
        with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as zf:
            names = zf.namelist()
            attachment_paths = [n for n in names if "/03_증빙첨부(attachments)/attachments/" in n and not n.endswith("/")]
            self.assertEqual(len(attachment_paths), 1)
            attachment_full_path = attachment_paths[0]
            root_prefix = attachment_full_path.split("/03_증빙첨부(attachments)/attachments/")[0]
            rel_path = attachment_full_path[len(root_prefix) + 1 :]

            idx_bytes = zf.read(f"{root_prefix}/03_증빙첨부(attachments)/attachments_index.xlsx")
            ev_bytes = zf.read(f"{root_prefix}/02_원장_원본데이터(raw)/evidence_index.xlsx")

        wb_idx = load_workbook(io.BytesIO(idx_bytes), read_only=True, data_only=True)
        ws_idx = wb_idx["attachments_index"]
        idx_headers = [str(cell.value or "") for cell in next(ws_idx.iter_rows(min_row=1, max_row=1))]
        idx_row = [cell.value for cell in next(ws_idx.iter_rows(min_row=2, max_row=2))]
        idx_col = next(i for i, h in enumerate(idx_headers) if "첨부 경로" in h)
        idx_path = idx_row[idx_col]

        wb_ev = load_workbook(io.BytesIO(ev_bytes), read_only=True, data_only=True)
        ws_ev = wb_ev["evidence_index"]
        ev_headers = [str(cell.value or "") for cell in next(ws_ev.iter_rows(min_row=1, max_row=1))]
        ev_row = [cell.value for cell in next(ws_ev.iter_rows(min_row=2, max_row=2))]
        ev_col = next(i for i, h in enumerate(ev_headers) if "첨부 경로" in h)
        ev_path = ev_row[ev_col]

        self.assertEqual(idx_path, rel_path)
        self.assertEqual(ev_path, rel_path)

    def test_duplicate_names_are_safely_incremented(self) -> None:
        tx = SimpleNamespace(
            id=200,
            occurred_at=datetime(2026, 3, 14, 13, 24, 55),
            amount_krw=15_800,
            counterparty="스타벅스",
            memo="",
        )
        ev = SimpleNamespace(
            original_filename="receipt.jpg",
            mime_type="image/jpeg",
            note="",
        )
        seq_by_tx: dict[int, int] = {}
        registry: set[str] = set()

        first = _build_attachment_zip_path(tx=tx, ev=ev, sequence_by_tx=seq_by_tx, name_registry=registry)
        seq_by_tx[tx.id] = 0
        second = _build_attachment_zip_path(tx=tx, ev=ev, sequence_by_tx=seq_by_tx, name_registry=registry)

        self.assertNotEqual(first, second)
        self.assertTrue(first.endswith("_001.jpg"))
        self.assertTrue(second.endswith("_002.jpg"))
        self.assertTrue(re.search(r"/attachments/.+_00[12]\.jpg$", first))


if __name__ == "__main__":
    unittest.main()
