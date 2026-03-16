# services/import_csv.py
from __future__ import annotations

import csv
import hashlib
import os
import re
import secrets
from dataclasses import dataclass
from datetime import datetime, date
from typing import Dict, List, Optional, Tuple

from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from zoneinfo import ZoneInfo

from core.extensions import db
from core.time import utcnow
from services.risk import refresh_recurring_candidates
from domain.models import (
    ImportJob,
    Transaction,
    IncomeLabel,
    ExpenseLabel,
    EvidenceItem,
    CounterpartyRule,
    CounterpartyExpenseRule,
    CsvFormatMapping,
)

# ----------------------------
# Errors / Results
# ----------------------------

class CsvImportError(Exception):
    pass


def normalize_csv_import_error(raw_message: str | None) -> str:
    text = str(raw_message or "").strip()
    low = text.lower()
    if not text:
        return "CSV 처리 중 문제가 발생했어요. 파일 형식을 확인하고 다시 시도해주세요."
    if ("openpyxl" in low) or ("xlrd" in low):
        return "엑셀 파일 처리 설정이 아직 준비되지 않았어요(개발용 설정 필요). CSV 파일로 업로드하거나 잠시 후 다시 시도해주세요."
    if ("파일 저장 실패" in text) or ("엑셀 변환 실패" in text):
        return "파일 준비 중 문제가 발생했어요. 다시 업로드해 주세요."
    if "날짜 파싱 실패" in text:
        return "날짜 형식을 읽지 못했어요. 거래일시 형식을 확인해 주세요."
    if ("지원하지 않는 파일 형식" in text) or ("지원 형식:" in text):
        return "지원 형식을 확인해 주세요. CSV 또는 엑셀(.xlsx/.xls)만 업로드할 수 있어요."
    if ("파일이 없습니다" in text) or ("업로드 파일을 찾을 수 없습니다" in text):
        return "업로드 파일을 찾을 수 없어요. 파일을 다시 선택해 주세요."
    if ("표 헤더를 찾지 못했습니다" in text) or ("거래내역 표를 찾지 못했습니다" in text):
        return "거래내역 표를 찾지 못했어요. 거래일시/금액 컬럼이 포함된 파일인지 확인해 주세요."
    short = text[:220]
    return short + ("…" if len(text) > 220 else "")


@dataclass
class CsvImportResult:
    import_job_id: int
    total_rows: int
    inserted_rows: int
    duplicate_rows: int
    failed_rows: int


kst = ZoneInfo("Asia/Seoul")

# ----------------------------
# File upload / preview
# ----------------------------

_TMP_BASE = "/tmp/safetospend_uploads"
_PREVIEW_ROWS = 20
_SNIFF_BYTES = 50_000

_DELIMS = [",", "\t", ";", "|"]

# Common Korean/English patterns for mapping
_PAT_DATE = re.compile(r"(일자|거래일|승인일|날짜|date|datetime|거래일시|승인일시)", re.I)
_PAT_AMOUNT = re.compile(r"(금액|거래금액|승인금액|amount|amt|금액\(원\))", re.I)
_PAT_IN_AMOUNT = re.compile(r"(입금|수입|credit|입금액|수입액)", re.I)
_PAT_OUT_AMOUNT = re.compile(r"(출금|지출|debit|출금액|지출액)", re.I)
_PAT_DIR = re.compile(r"(구분|입출|입\/출|입출금|type|dr\/cr|direction)", re.I)
_PAT_CP = re.compile(r"(거래처|가맹점|상호|merchant|counterparty|가게|업체|사용처)", re.I)
_PAT_MEMO = re.compile(r"(적요|내용|메모|비고|description|memo|내역|상세)", re.I)
_PAT_BALANCE = re.compile(r"(잔액|balance)", re.I)
_PAT_CP = re.compile(r"(거래처|가맹점|상대|상대방|보낸분|받는분|보낸분/받는분|입금자|출금계좌|수취인|송금인|상호|사용처)", re.I)
_PAT_MEMO = re.compile(r"(적요|내용|메모|비고|내역|송금메모|거래내용|거래구분)", re.I)


def _ensure_user_tmp_dir(user_pk: int) -> str:
    path = os.path.join(_TMP_BASE, str(user_pk))
    os.makedirs(path, exist_ok=True)
    return path


def _excel_to_csv(src_path: str, dst_path: str) -> None:
    ext = os.path.splitext(src_path)[1].lower()

    def _cell_to_str(v):
        if v is None:
            return ""
        if isinstance(v, (datetime, date)):
            # dateutil 파서가 잘 읽는 형태
            try:
                return v.isoformat(sep=" ")
            except TypeError:
                return v.isoformat()
        s = str(v).strip()
        return s

    # .xlsx
    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except Exception:
            raise CsvImportError("엑셀(.xlsx) 처리를 위해 openpyxl이 필요합니다. requirements 설치를 확인해주세요.")

        wb = load_workbook(src_path, read_only=True, data_only=True)
        ws = wb.active  # 첫 시트

        with open(dst_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            wrote_any = False
            for row in ws.iter_rows(values_only=True):
                if not row:
                    continue
                out = [_cell_to_str(v) for v in row]
                if all(not x for x in out):
                    continue
                w.writerow(out)
                wrote_any = True

        if not wrote_any:
            raise CsvImportError("엑셀 파일이 비어있습니다. (첫 시트에 데이터가 있어야 합니다.)")
        return

    # .xls (구형)
    if ext == ".xls":
        try:
            import xlrd  # type: ignore
        except Exception:
            raise CsvImportError("구형 엑셀(.xls)은 추가 모듈(xlrd)이 필요합니다. .xlsx 또는 CSV로 저장 후 업로드해주세요.")

        wb = xlrd.open_workbook(src_path)
        sh = wb.sheet_by_index(0)  # 첫 시트

        with open(dst_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            wrote_any = False
            for r in range(sh.nrows):
                out = []
                empty = True
                for c in range(sh.ncols):
                    cell = sh.cell(r, c)
                    v = cell.value

                    # 날짜 셀 처리
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            dt = xlrd.xldate.xldate_as_datetime(v, wb.datemode)
                            s = dt.strftime("%Y-%m-%d %H:%M:%S")
                        except Exception:
                            s = ""
                    else:
                        s = _cell_to_str(v)

                    out.append(s)
                    if s:
                        empty = False

                if empty:
                    continue
                w.writerow(out)
                wrote_any = True

        if not wrote_any:
            raise CsvImportError("엑셀 파일이 비어있습니다. (첫 시트에 데이터가 있어야 합니다.)")
        return

    raise CsvImportError("지원하지 않는 파일 형식입니다. CSV 또는 Excel(.xlsx/.xls)을 업로드해주세요.")


def save_temp_upload(file: FileStorage | None, user_pk: int) -> Tuple[str, str, str]:
    """Save uploaded CSV/XLSX/XLS into /tmp for preview/import.
    Excel files are converted to CSV and then processed by the existing pipeline.
    Returns (token, filepath(csv), original_filename).
    """
    if not file or not file.filename:
        raise CsvImportError("파일이 없습니다.")

    filename = secure_filename(file.filename)
    lower = filename.lower()

    ext = ".csv"
    if lower.endswith(".csv"):
        ext = ".csv"
    elif lower.endswith(".xlsx"):
        ext = ".xlsx"
    elif lower.endswith(".xls"):
        ext = ".xls"
    else:
        raise CsvImportError("지원 형식: CSV(.csv), 엑셀(.xlsx, .xls)")

    token = secrets.token_hex(16)
    user_dir = _ensure_user_tmp_dir(user_pk)

    # 원본 저장 경로(엑셀은 원본 먼저 저장)
    raw_path = os.path.join(user_dir, f"{token}{ext}")
    csv_path = os.path.join(user_dir, f"{token}.csv")

    try:
        file.save(raw_path)
    except Exception as e:
        raise CsvImportError(f"파일 저장 실패: {e}")

    # 엑셀이면 CSV로 변환
    if ext in (".xlsx", ".xls"):
        try:
            _excel_to_csv(raw_path, csv_path)
        except CsvImportError:
            raise
        except Exception as e:
            raise CsvImportError(f"엑셀 변환 실패: {e}")
        finally:
            # 원본 파일은 정리(필요하면 남겨도 되지만 기본은 삭제)
            try:
                os.remove(raw_path)
            except Exception:
                pass
        return token, csv_path, filename

    # CSV면 그대로 사용
    return token, raw_path, filename


def _read_bytes_sample(filepath: str) -> bytes:
    with open(filepath, "rb") as f:
        return f.read(_SNIFF_BYTES)


def _decode_best_effort(raw: bytes) -> str:
    # Try common encodings for KR CSV exports
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    # last resort: replace
    return raw.decode("utf-8", errors="replace")


def _sniff_delimiter(text_sample: str) -> str:
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text_sample, delimiters=_DELIMS)
        return dialect.delimiter
    except Exception:
        # fallback: choose the delimiter with most occurrences in first non-empty line
        lines = [ln for ln in text_sample.splitlines() if ln.strip()]
        if not lines:
            return ","
        line = lines[0]
        best = ","
        best_count = -1
        for d in _DELIMS:
            c = line.count(d)
            if c > best_count:
                best = d
                best_count = c
        return best


# ----------------------------
# CSV table detection (은행/카드사 엑셀/CSV 상단 메타행 자동 스킵)
# ----------------------------

def _row_is_empty(row: List[str]) -> bool:
    return (not row) or all(not (str(c).strip()) for c in row)


def _clean_row(row: List[str]) -> List[str]:
    out = [(c.strip() if isinstance(c, str) else "") for c in (row or [])]
    while out and not out[-1]:
        out.pop()
    return out


def _detect_header_row_index(rows: List[List[str]]) -> int:
    """상단 메타행(조회기간/계좌번호 등)을 건너뛰고, '거래일시+금액(또는 입금/출금)'이 있는 표 헤더 행을 찾는다."""
    best_idx = -1
    best_score = -1

    scan_n = min(len(rows), 80)
    bal_pat = globals().get("_PAT_BALANCE", None)

    for i in range(scan_n):
        r = _clean_row(rows[i])
        if _row_is_empty(r):
            continue

        date_hits = sum(1 for c in r if c and _PAT_DATE.search(c))
        amt_hits = sum(1 for c in r if c and (_PAT_AMOUNT.search(c) or _PAT_IN_AMOUNT.search(c) or _PAT_OUT_AMOUNT.search(c)))
        cp_hits = sum(1 for c in r if c and _PAT_CP.search(c))
        memo_hits = sum(1 for c in r if c and _PAT_MEMO.search(c))
        dir_hits = sum(1 for c in r if c and _PAT_DIR.search(c))

        if date_hits <= 0 or amt_hits <= 0:
            continue

        nonempty = sum(1 for c in r if c)
        if nonempty < 3:
            continue

        header_idx = _build_header_index(r)

        # date col (첫 번째 date 키워드 컬럼)
        date_col = None
        for c in r:
            if c and _PAT_DATE.search(c):
                date_col = c
                break

        # amount cols (입금/출금 우선, 없으면 금액)
        amount_cols: List[str] = []
        for c in r:
            if c and _PAT_IN_AMOUNT.search(c):
                amount_cols.append(c)
        for c in r:
            if c and _PAT_OUT_AMOUNT.search(c):
                amount_cols.append(c)
        if not amount_cols:
            for c in r:
                if not c:
                    continue
                if not _PAT_AMOUNT.search(c):
                    continue
                if bal_pat and bal_pat.search(c):
                    continue
                amount_cols.append(c)

        look_n = 12
        ok_date = 0
        ok_amt = 0
        seen = 0

        for j in range(i + 1, min(len(rows), i + 1 + look_n)):
            rr = _clean_row(rows[j])
            if _row_is_empty(rr):
                continue
            seen += 1

            try:
                if date_col:
                    dt_raw = _get_cell(rr, header_idx, date_col)
                    _ = _parse_datetime_kst_to_utc(dt_raw)
                    ok_date += 1
            except Exception:
                pass

            try:
                amt_ok = False
                for ac in amount_cols:
                    v = _get_cell(rr, header_idx, ac)
                    a = _parse_int_amount(v)
                    if a is not None and abs(int(a)) > 0:
                        amt_ok = True
                        break
                if amt_ok:
                    ok_amt += 1
            except Exception:
                pass

        if seen <= 0:
            continue

        date_rate = ok_date / seen
        amt_rate = ok_amt / seen

        score = (
            40 * date_hits
            + 35 * amt_hits
            + 5 * cp_hits
            + 5 * memo_hits
            + 5 * dir_hits
            + int(100 * (0.5 * date_rate + 0.5 * amt_rate))
        )

        if score > best_score:
            best_score = score
            best_idx = i

    if best_idx >= 0:
        return best_idx

    # fallback: 첫 번째 non-empty 행
    for i in range(len(rows)):
        r = _clean_row(rows[i])
        if not _row_is_empty(r):
            return i
    return 0


def _read_csv_table(filepath: str) -> Tuple[List[str], List[List[str]], str, int]:
    """(headers, data_rows, delimiter, header_row_index)"""
    raw = _read_bytes_sample(filepath)
    sample_text = _decode_best_effort(raw)
    delimiter = _sniff_delimiter(sample_text)

    with open(filepath, "rb") as f:
        full_text = _decode_best_effort(f.read())

    full_text = full_text.replace("\r\n", "\n").replace("\r", "\n")
    lines = full_text.split("\n")
    reader = csv.reader(lines, delimiter=delimiter)

    raw_rows: List[List[str]] = []
    for row in reader:
        r = _clean_row([c if isinstance(c, str) else "" for c in row])
        if _row_is_empty(r):
            continue
        raw_rows.append(r)

    if not raw_rows:
        raise CsvImportError("파일이 비어있습니다.")

    hdr_idx = _detect_header_row_index(raw_rows)
    headers = _clean_row(raw_rows[hdr_idx])

    if not headers or all(not h for h in headers):
        raise CsvImportError("표 헤더를 찾지 못했습니다. (거래일시/금액 컬럼이 있는 거래내역 표를 업로드해주세요)")

    data_rows = raw_rows[hdr_idx + 1 :]

    if not data_rows:
        raise CsvImportError("거래내역 표를 찾지 못했습니다. 엑셀 첫 시트에 거래내역 표가 있는지 확인해주세요.")

    return headers, data_rows, delimiter, hdr_idx


def read_csv_preview(filepath: str) -> Tuple[List[str], List[List[str]], str]:
    """표 헤더 자동 탐지 후, 데이터 첫 N행 미리보기"""
    headers, data_rows, delimiter, _hdr_idx = _read_csv_table(filepath)
    return headers, data_rows[:_PREVIEW_ROWS], delimiter


def suggest_mapping(headers: List[str]) -> Dict[str, str]:
    """Heuristic mapping suggestion from headers."""
    def pick(pattern: re.Pattern) -> str:
        for h in headers:
            if pattern.search(h or ""):
                return h
        return ""

    mapping = {
        "date": pick(_PAT_DATE),
        "amount": pick(_PAT_AMOUNT),
        "in_amount": pick(_PAT_IN_AMOUNT),
        "out_amount": pick(_PAT_OUT_AMOUNT),
        "direction": pick(_PAT_DIR),
        "counterparty": pick(_PAT_CP),
        "memo": pick(_PAT_MEMO),
    }

    # If both in/out exist, often amount should be left blank
    if mapping["in_amount"] and mapping["out_amount"]:
        mapping["amount"] = ""
        
    if mapping["amount"] and _PAT_BALANCE.search(mapping["amount"]):
        mapping["amount"] = ""

    return mapping
    

_AUTO_IMPORT_MIN_CONF = 90
_AUTO_IMPORT_MIN_RATE = 0.90

def temp_upload_path(user_pk: int, token: str) -> str:
    return os.path.join(_TMP_BASE, str(user_pk), f"{token}.csv")

def compute_format_signature(headers: List[str], delimiter: str) -> str:
    norm = []
    for h in headers:
        h2 = (h or "").strip().lower()
        h2 = re.sub(r"\s+", "", h2)
        norm.append(h2)
    base = f"{delimiter}|{len(headers)}|" + "|".join(norm)
    return hashlib.sha256(base.encode("utf-8")).hexdigest()

def _safe_db_call(fn, default=None):
    try:
        return fn()
    except Exception:
        db.session.rollback()
        return default

def load_cached_mapping(user_pk: int, signature: str) -> Optional[Dict[str, str]]:
    def _q():
        row = CsvFormatMapping.query.filter_by(user_pk=user_pk, signature=signature).first()
        return dict(row.mapping) if row and row.mapping else None
    return _safe_db_call(_q, default=None)

def save_cached_mapping(user_pk: int, signature: str, delimiter: str, mapping: Dict[str, str], meta: Optional[dict] = None) -> None:
    def _upsert():
        row = CsvFormatMapping.query.filter_by(user_pk=user_pk, signature=signature).first()
        if not row:
            row = CsvFormatMapping(user_pk=user_pk, signature=signature, mapping=mapping, delimiter=delimiter, meta=meta)
            db.session.add(row)
        else:
            row.mapping = mapping
            row.delimiter = delimiter
            row.meta = meta
        db.session.commit()
    _safe_db_call(_upsert, default=None)

def score_mapping(headers: List[str], rows: List[List[str]], mapping: Dict[str, str]) -> Tuple[int, float, float]:
    map_date = (mapping.get("date") or "").strip()
    map_amount = (mapping.get("amount") or "").strip()
    map_in = (mapping.get("in_amount") or "").strip()
    map_out = (mapping.get("out_amount") or "").strip()
    map_dir = (mapping.get("direction") or "").strip()

    if not map_date:
        return 0, 0.0, 0.0
    if not map_amount and not (map_in or map_out):
        return 0, 0.0, 0.0
    if map_amount and _PAT_BALANCE.search(map_amount):
        return 0, 0.0, 0.0

    header_idx = _build_header_index(headers)
    total = 0
    ok_date = 0
    ok_amt = 0

    for row in rows[:_PREVIEW_ROWS]:
        if not row or all(not str(c).strip() for c in row):
            continue
        total += 1
        try:
            dt_raw = _get_cell(row, header_idx, map_date)
            _ = _parse_datetime_kst_to_utc(dt_raw)
            ok_date += 1
        except Exception:
            pass

        try:
            amount_krw = None
            if map_in or map_out:
                in_amt = _parse_int_amount(_get_cell(row, header_idx, map_in)) if map_in else None
                out_amt = _parse_int_amount(_get_cell(row, header_idx, map_out)) if map_out else None
                if in_amt is not None and in_amt != 0 and (out_amt is None or out_amt == 0):
                    amount_krw = abs(int(in_amt))
                elif out_amt is not None and out_amt != 0 and (in_amt is None or in_amt == 0):
                    amount_krw = abs(int(out_amt))
                elif in_amt is not None and out_amt is not None and (in_amt != 0 or out_amt != 0):
                    amount_krw = abs(int(in_amt)) if abs(in_amt) >= abs(out_amt) else abs(int(out_amt))
            elif map_amount:
                amt = _parse_int_amount(_get_cell(row, header_idx, map_amount))
                if amt is not None:
                    amount_krw = abs(int(amt))

            if amount_krw is not None and amount_krw > 0:
                ok_amt += 1
        except Exception:
            pass

    if total <= 0:
        return 0, 0.0, 0.0

    date_rate = ok_date / total
    amt_rate = ok_amt / total

    base = 100.0 * (0.55 * date_rate + 0.45 * amt_rate)

    if (mapping.get("counterparty") or "").strip():
        base += 2
    if (mapping.get("memo") or "").strip():
        base += 1
    if map_dir:
        base += 2

    conf = int(max(0, min(100, round(base))))
    return conf, date_rate, amt_rate

def detect_mapping_for_preview(user_pk: int, headers: List[str], rows: List[List[str]], delimiter: str) -> Tuple[str, Dict[str, str], int, float, float, str]:
    sig = compute_format_signature(headers, delimiter)

    cached = load_cached_mapping(user_pk, sig)
    if cached:
        conf, dr, ar = score_mapping(headers, rows, cached)
        return sig, cached, max(conf, 95), max(dr, 0.95), max(ar, 0.95), "saved"

    m = suggest_mapping(headers)
    conf, dr, ar = score_mapping(headers, rows, m)
    return sig, m, conf, dr, ar, "heuristic"


# ----------------------------
# Import core
# ----------------------------

def _normalize_space(s: Optional[str]) -> str:
    if not s:
        return ""
    s = str(s).strip()
    s = " ".join(s.split())
    return s


def _parse_int_amount(v: Optional[str]) -> Optional[int]:
    """Parse KRW integer from strings like '1,234', '₩1,234', '(1,234)', '-1234'."""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None

    # parentheses as negative
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1].strip()

    s = s.replace("원", "").replace("₩", "").replace(",", "").replace(" ", "")
    # keep digits, sign, dot
    s = re.sub(r"[^0-9\.\-\+]", "", s)
    if not s or s in ("-", "+", ".", "-.", "+."):
        return None

    try:
        # Some CSVs export amounts as "1234.00"
        val = float(s)
    except ValueError:
        return None

    iv = int(round(val))
    if neg:
        iv = -abs(iv)
    return iv


def _parse_direction(v: Optional[str]) -> Optional[str]:
    if not v:
        return None
    s = _normalize_space(v).lower()

    # Korean
    if any(k in s for k in ("입금", "수입", "입", "credit", "cr")):
        if "출" not in s and "지출" not in s and "debit" not in s and "dr" not in s:
            return "in"
    if any(k in s for k in ("출금", "지출", "출", "debit", "dr")):
        return "out"

    # English common words
    if s in ("in", "income", "credit"):
        return "in"
    if s in ("out", "expense", "debit"):
        return "out"

    return None


def _parse_datetime_kst_to_utc(dt_str: str) -> datetime:
    """Parse a date/datetime string assuming KST when tz missing, store UTC."""
    if not dt_str or not str(dt_str).strip():
        raise CsvImportError("거래일시 값이 비어있습니다.")

    from dateutil import parser
    from zoneinfo import ZoneInfo

    kst = ZoneInfo("Asia/Seoul")
    try:
        dt = parser.parse(str(dt_str))
    except Exception as e:
        raise CsvImportError(f"날짜 파싱 실패: '{dt_str}' ({e})")

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=kst)

    return dt.astimezone(ZoneInfo("UTC"))


def _compute_external_hash(
    occurred_at_utc: datetime,
    direction: str,
    amount_krw: int,
    counterparty: str,
    memo: str,
) -> str:
    # Keep it stable and tolerant
    base = f"{occurred_at_utc.isoformat()}|{direction}|{amount_krw}|{counterparty.lower()}|{memo.lower()}"
    return hashlib.sha256(base.encode("utf-8")).hexdigest()


def _build_header_index(headers: List[str]) -> Dict[str, int]:
    idx: Dict[str, int] = {}
    for i, h in enumerate(headers):
        h2 = str(h).strip()
        if h2 and h2 not in idx:
            idx[h2] = i
    return idx


def _get_cell(row: List[str], header_idx: Dict[str, int], col: str) -> str:
    if not col:
        return ""
    i = header_idx.get(col)
    if i is None:
        return ""
    if i >= len(row):
        return ""
    v = row[i]
    return v if v is not None else ""


def _chunked(iterable: List[str], size: int) -> List[List[str]]:
    return [iterable[i:i + size] for i in range(0, len(iterable), size)]


def _load_income_rules(user_pk: int) -> Dict[str, str]:
    rules = (
        CounterpartyRule.query
        .filter_by(user_pk=user_pk, active=True)
        .all()
    )
    # key: normalized counterparty_key, val: "income"/"non_income"
    return {r.counterparty_key: r.rule for r in rules}


def _load_expense_rules(user_pk: int) -> Dict[str, str]:
    rules = (
        CounterpartyExpenseRule.query
        .filter_by(user_pk=user_pk, active=True)
        .all()
    )
    # key: normalized counterparty_key, val: "business"/"personal"
    return {r.counterparty_key: r.rule for r in rules}


def _norm_counterparty_key(counterparty: str) -> str:
    return _normalize_space(counterparty).lower()


def _apply_income_rule(counterparty: str, income_rules: Dict[str, str]) -> Tuple[str, int]:
    """Return (status, confidence)."""
    key = _norm_counterparty_key(counterparty)
    rule = income_rules.get(key)
    if rule in ("income", "non_income"):
        return rule, 80
    return "unknown", 0


def _apply_expense_rule(counterparty: str, expense_rules: Dict[str, str]) -> Tuple[str, int]:
    key = _norm_counterparty_key(counterparty)
    rule = expense_rules.get(key)
    if rule in ("business", "personal"):
        return rule, 80
    return "unknown", 0


def import_csv_to_db(
    user_pk: int,
    filepath: str,
    filename: str,
    mapping: Dict[str, str],
    bank_account_id: int | None = None,
) -> CsvImportResult:
    """Import CSV file into DB as Transactions + default labels/evidence.
    mapping keys:
      date (required),
      amount OR (in_amount/out_amount) (at least one strategy required),
      direction (optional),
      counterparty (optional),
      memo (optional)
    """
    if not os.path.exists(filepath):
        raise CsvImportError("업로드 파일을 찾을 수 없습니다. 다시 업로드해주세요.")

    map_date = (mapping.get("date") or "").strip()
    map_amount = (mapping.get("amount") or "").strip()
    map_in = (mapping.get("in_amount") or "").strip()
    map_out = (mapping.get("out_amount") or "").strip()
    map_dir = (mapping.get("direction") or "").strip()
    map_cp = (mapping.get("counterparty") or "").strip()
    map_memo = (mapping.get("memo") or "").strip()

    if not map_date:
        raise CsvImportError("거래일시 컬럼은 필수입니다.")
    if not map_amount and not (map_in or map_out):
        raise CsvImportError("금액 컬럼이 필요합니다. (단일 금액 또는 입금/출금 컬럼을 지정하세요)")

    # Read full CSV text
    # Read CSV table (auto-skip meta rows)
    headers, data_rows, delimiter, header_row_idx = _read_csv_table(filepath)

    header_idx = _build_header_index(headers)

    # Validate mapping columns exist in headers (when provided)
    for key, col in (
        ("거래일시", map_date),
        ("금액", map_amount),
        ("입금 금액", map_in),
        ("출금 금액", map_out),
        ("입/출 구분", map_dir),
        ("거래처", map_cp),
        ("메모", map_memo),
    ):
        if col and col not in header_idx:
            raise CsvImportError(f"{key}로 지정한 컬럼을 CSV에서 찾지 못했습니다: '{col}'")

    # Create import job
    job = ImportJob(
        user_pk=user_pk,
        source="csv",
        filename=filename,
        total_rows=0,
        inserted_rows=0,
        duplicate_rows=0,
        failed_rows=0,
        error_summary={},
        started_at=utcnow(),
    )
    db.session.add(job)
    db.session.commit()  # get job.id

    income_rules = _load_income_rules(user_pk)
    expense_rules = _load_expense_rules(user_pk)

    # Parse rows
    parsed: List[dict] = []
    errors: List[dict] = []

    row_num = header_row_idx + 1  # header line number (1-indexed)
    for row in data_rows:
        row_num += 1
        # skip empty lines
        if not row or all(not str(c).strip() for c in row):
            continue

        job.total_rows += 1

        try:

            dt_raw = _get_cell(row, header_idx, map_date)

            occurred_at_utc = _parse_datetime_kst_to_utc(dt_raw)
            from zoneinfo import ZoneInfo
            kst = ZoneInfo("Asia/Seoul")
            occurred_at = occurred_at_utc.astimezone(kst).replace(tzinfo=None)

            cp = _normalize_space(_get_cell(row, header_idx, map_cp)) if map_cp else ""
            memo = _normalize_space(_get_cell(row, header_idx, map_memo)) if map_memo else ""
            
            # 기존: cp = ... , memo = ...
            cp = (cp or "").strip()
            memo = (memo or "").strip()

            # (1) 거래처가 비어있으면 메모에서라도 채움(제목 표시용 fallback)
            if not cp and memo:
                cp = memo

            # (2) KB류 파일에서 '적요'와 '송금메모'가 분리되어 들어오는 경우 대비:
            # mapping에 memo로 들어온 값이 '적요'일 수 있으니, 다른 컬럼이 있다면 합치기
            # (너가 이미 매핑 UI에서 map_memo를 "적요"로 잡았을 가능성이 큼)
            # -> 아래는 "송금메모" 같은 컬럼이 존재하면 자동으로 더함
            try:
                # header_idx, row가 있는 스코프라고 가정
                # 송금메모/메모/내용 후보 컬럼이 있으면 붙임
                for h in header_idx.keys():
                    if "송금메모" in h or "송금 메모" in h:
                        extra = (_get_cell(row, header_idx, h) or "").strip()
                        if extra and extra not in memo:
                            memo = f"{memo} · {extra}" if memo else extra
            except Exception:
                pass

            # Determine direction + amount
            direction: Optional[str] = None
            amount_krw: Optional[int] = None

            if map_dir:
                direction = _parse_direction(_get_cell(row, header_idx, map_dir))

            if map_in or map_out:
                in_amt = _parse_int_amount(_get_cell(row, header_idx, map_in)) if map_in else None
                out_amt = _parse_int_amount(_get_cell(row, header_idx, map_out)) if map_out else None

                # choose the non-empty one
                if in_amt is not None and in_amt != 0 and (out_amt is None or out_amt == 0):
                    direction = direction or "in"
                    amount_krw = abs(int(in_amt))
                elif out_amt is not None and out_amt != 0 and (in_amt is None or in_amt == 0):
                    direction = direction or "out"
                    amount_krw = abs(int(out_amt))
                elif in_amt is not None and out_amt is not None and (in_amt != 0 or out_amt != 0):
                    # some exports put both; treat net direction by larger abs
                    if abs(in_amt) >= abs(out_amt):
                        direction = direction or "in"
                        amount_krw = abs(int(in_amt))
                    else:
                        direction = direction or "out"
                        amount_krw = abs(int(out_amt))

            if amount_krw is None and map_amount:
                amt = _parse_int_amount(_get_cell(row, header_idx, map_amount))
                if amt is None:
                    raise CsvImportError("금액을 파싱할 수 없습니다.")
                # If direction unknown, infer from sign
                if direction is None:
                    if amt < 0:
                        direction = "out"
                        amount_krw = abs(amt)
                    else:
                        direction = "in"
                        amount_krw = abs(amt)
                else:
                    amount_krw = abs(int(amt))

            if direction not in ("in", "out"):
                # last inference: if amount provided and sign unknown, assume out? no, assume in for non-negative.
                if direction is None:
                    direction = "in"
                else:
                    raise CsvImportError("입/출 구분을 판단할 수 없습니다.")

            if amount_krw is None or amount_krw <= 0:
                raise CsvImportError("금액이 비어있거나 0입니다.")

            external_hash = _compute_external_hash(
                occurred_at_utc=occurred_at_utc,   # ✅ 여기 중요
                direction=direction,
                amount_krw=amount_krw,
                counterparty=cp,
                memo=memo,
            )

            parsed.append(
                {
                    "occurred_at": occurred_at,     # ✅ KST naive 저장
                    "direction": direction,
                    "amount_krw": amount_krw,
                    "counterparty": cp,
                    "memo": memo,
                    "external_hash": external_hash,
                }
            )
        except Exception as e:
            job.failed_rows += 1
            errors.append({"row": row_num, "error": str(e)})
            continue

    # Save job parsing stats early
    if errors:
        job.error_summary = {"errors": errors[:50]}  # cap to avoid huge json
    db.session.commit()

    if not parsed:
        job.finished_at = utcnow()
        db.session.commit()
        raise CsvImportError("가져올 거래가 없습니다. (CSV 매핑/내용을 확인해주세요)")

    # Dedup: find existing hashes
    hashes = [p["external_hash"] for p in parsed]
    existing: set[str] = set()
    for chunk in _chunked(hashes, 1000):
        rows = (
            db.session.query(Transaction.external_hash)
            .filter(Transaction.user_pk == user_pk, Transaction.external_hash.in_(chunk))
            .all()
        )
        existing.update([r[0] for r in rows])

    to_insert = [p for p in parsed if p["external_hash"] not in existing]
    job.duplicate_rows = len(parsed) - len(to_insert)

    if not to_insert:
        job.finished_at = utcnow()
        db.session.commit()
        return CsvImportResult(
            import_job_id=job.id,
            total_rows=job.total_rows,
            inserted_rows=0,
            duplicate_rows=job.duplicate_rows,
            failed_rows=job.failed_rows,
        )

    # Insert Transactions
    tx_objs: List[Transaction] = []
    tx_bank_account_id = int(bank_account_id) if isinstance(bank_account_id, int) and int(bank_account_id) > 0 else None
    for p in to_insert:
        tx_objs.append(
            Transaction(
                user_pk=user_pk,
                import_job_id=job.id,
                occurred_at=p["occurred_at"],
                direction=p["direction"],
                amount_krw=int(p["amount_krw"]),
                counterparty=p["counterparty"] or None,
                memo=p["memo"] or None,
                source="csv",
                bank_account_id=tx_bank_account_id,
                external_hash=p["external_hash"],
                created_at=utcnow(),
            )
        )

    db.session.bulk_save_objects(tx_objs)
    db.session.commit()

    # Fetch inserted transactions to attach labels/evidence (need IDs)
    inserted_hashes = [p["external_hash"] for p in to_insert]
    inserted_txs: List[Transaction] = []
    for chunk in _chunked(inserted_hashes, 1000):
        inserted_txs.extend(
            Transaction.query
            .filter(Transaction.user_pk == user_pk, Transaction.external_hash.in_(chunk))
            .all()
        )

    # Attach IncomeLabel / ExpenseLabel / EvidenceItem
    income_labels: List[IncomeLabel] = []
    expense_labels: List[ExpenseLabel] = []
    evidences: List[EvidenceItem] = []

    now = utcnow()

    for tx in inserted_txs:
        cp = tx.counterparty or ""

        if tx.direction == "in":
            status, conf = _apply_income_rule(cp, income_rules)
            income_labels.append(
                IncomeLabel(
                    user_pk=user_pk,
                    transaction_id=tx.id,
                    status=status,
                    confidence=conf,
                    labeled_by="auto",
                    rule_version=1,
                    decided_at=(now if status != "unknown" else None),
                    note=None,
                    created_at=now,
                    updated_at=now,
                )
            )
        else:
            estatus, conf = _apply_expense_rule(cp, expense_rules)
            expense_labels.append(
                ExpenseLabel(
                    user_pk=user_pk,
                    transaction_id=tx.id,
                    status=estatus,
                    confidence=conf,
                    labeled_by="auto",
                    rule_version=1,
                    decided_at=(now if estatus != "unknown" else None),
                    note=None,
                    created_at=now,
                    updated_at=now,
                )
            )

            # Evidence default based on expense label
            if estatus == "business":
                requirement = "required"
                ev_status = "missing"
            elif estatus == "personal":
                requirement = "not_needed"
                ev_status = "not_needed"
            else:
                requirement = "maybe"
                ev_status = "missing"

            evidences.append(
                EvidenceItem(
                    user_pk=user_pk,
                    transaction_id=tx.id,
                    requirement=requirement,
                    status=ev_status,
                    note=None,
                    created_at=now,
                    updated_at=now,
                )
            )

    if income_labels:
        db.session.bulk_save_objects(income_labels)
    if expense_labels:
        db.session.bulk_save_objects(expense_labels)
    if evidences:
        db.session.bulk_save_objects(evidences)

    job.inserted_rows = len(inserted_txs)
    job.finished_at = utcnow()
    db.session.commit()

    # 정기 거래 후보 갱신은 가져오기 성공 흐름을 막지 않도록 best-effort로 처리
    try:
        refresh_recurring_candidates(user_pk=user_pk, lookback_days=90, min_samples=3)
    except Exception:
        db.session.rollback()

    return CsvImportResult(
        import_job_id=job.id,
        total_rows=job.total_rows,
        inserted_rows=job.inserted_rows,
        duplicate_rows=job.duplicate_rows,
        failed_rows=job.failed_rows,
    )
