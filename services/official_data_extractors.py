from __future__ import annotations

import csv
import hashlib
import io
import mimetypes
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".pdf"}
SUPPORTED_MIME_TYPES = {
    ".csv": {"text/csv", "application/csv", "application/vnd.ms-excel"},
    ".xlsx": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    ".pdf": {"application/pdf"},
}
DEFAULT_MAX_BYTES = 10 * 1024 * 1024


@dataclass(slots=True)
class OfficialDataFileEnvelope:
    filename: str
    ext: str
    mime_type: str
    size_bytes: int
    sha256: str
    raw_bytes: bytes


class OfficialDataFileError(ValueError):
    pass


def _safe_filename(raw: str | None) -> str:
    filename = secure_filename(str(raw or "").strip())
    return filename or "official_data_upload"


def _normalize_mime_type(filename: str, content_type: str | None) -> str:
    guessed, _ = mimetypes.guess_type(filename)
    mime = str(content_type or guessed or "application/octet-stream").split(";", 1)[0].strip().lower()
    if mime:
        return mime
    ext = Path(filename).suffix.lower()
    if ext == ".csv":
        return "text/csv"
    if ext == ".xlsx":
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if ext == ".pdf":
        return "application/pdf"
    return "application/octet-stream"


def build_upload_envelope(uploaded: FileStorage, *, max_bytes: int = DEFAULT_MAX_BYTES) -> OfficialDataFileEnvelope:
    if uploaded is None or not getattr(uploaded, "filename", None):
        raise OfficialDataFileError("파일을 먼저 선택해 주세요.")
    filename = _safe_filename(uploaded.filename)
    ext = Path(filename).suffix.lower()
    try:
        if hasattr(uploaded.stream, "seek"):
            uploaded.stream.seek(0)
        raw_bytes = uploaded.read() or b""
    finally:
        try:
            if hasattr(uploaded.stream, "seek"):
                uploaded.stream.seek(0)
        except Exception:
            pass
    size_bytes = int(len(raw_bytes))
    if size_bytes <= 0:
        raise OfficialDataFileError("비어 있는 파일은 처리할 수 없어요.")
    if size_bytes > int(max_bytes or DEFAULT_MAX_BYTES):
        raise OfficialDataFileError("파일이 너무 커요. 지원 범위 안의 원본 파일만 올려주세요.")
    mime_type = _normalize_mime_type(filename, getattr(uploaded, "mimetype", None))
    return OfficialDataFileEnvelope(
        filename=filename,
        ext=ext,
        mime_type=mime_type,
        size_bytes=size_bytes,
        sha256=hashlib.sha256(raw_bytes).hexdigest(),
        raw_bytes=raw_bytes,
    )


def build_envelope_from_path(path: str | Path, *, mime_type: str | None = None) -> OfficialDataFileEnvelope:
    p = Path(path)
    raw_bytes = p.read_bytes()
    filename = p.name
    ext = p.suffix.lower()
    return OfficialDataFileEnvelope(
        filename=filename,
        ext=ext,
        mime_type=_normalize_mime_type(filename, mime_type),
        size_bytes=len(raw_bytes),
        sha256=hashlib.sha256(raw_bytes).hexdigest(),
        raw_bytes=raw_bytes,
    )


def is_supported_extension(ext: str) -> bool:
    return str(ext or "").lower() in SUPPORTED_EXTENSIONS


def is_supported_mime_for_extension(ext: str, mime_type: str) -> bool:
    ext_key = str(ext or "").lower()
    allowed = SUPPORTED_MIME_TYPES.get(ext_key, set())
    if not allowed:
        return False
    if ext_key == ".csv" and mime_type in {"text/plain", "application/octet-stream"}:
        return True
    return mime_type in allowed or mime_type == "application/octet-stream"


def decode_text_bytes(raw_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr", "latin-1"):
        try:
            return raw_bytes.decode(encoding)
        except Exception:
            continue
    return raw_bytes.decode("utf-8", errors="ignore")


def read_csv_matrix(raw_bytes: bytes) -> list[list[str]]:
    text = decode_text_bytes(raw_bytes)
    sample = text[:5000]
    delimiter = ","
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except Exception:
        delimiter = ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows: list[list[str]] = []
    for row in reader:
        cleaned = [str(cell or "").strip() for cell in row]
        if any(cleaned):
            rows.append(cleaned)
    return rows


def read_xlsx_matrix(raw_bytes: bytes) -> list[list[str]]:
    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        cleaned = [str(cell).strip() if cell is not None else "" for cell in row]
        if any(cleaned):
            rows.append(cleaned)
    wb.close()
    return rows


def matrix_to_preview_text(matrix: Iterable[Iterable[str]], *, limit: int = 12) -> str:
    lines: list[str] = []
    for idx, row in enumerate(matrix):
        if idx >= limit:
            break
        joined = " | ".join(str(cell or "").strip() for cell in row if str(cell or "").strip())
        if joined:
            lines.append(joined)
    return "\n".join(lines)


def extract_pdf_text(raw_bytes: bytes) -> str:
    text = decode_text_bytes(raw_bytes)
    lines: list[str] = []
    seen: set[str] = set()
    for raw_line in re.split(r"\r?\n", text):
        line = str(raw_line or "").strip()
        if not line:
            continue
        if line.startswith("%"):
            line = line.lstrip("% ").strip()
        if not line:
            continue
        if line in {"stream", "endstream", "endobj", "obj", "xref", "trailer", "startxref", "%%EOF"}:
            continue
        if not any(("가" <= ch <= "힣") or ch.isalpha() or ch.isdigit() for ch in line):
            continue
        if line not in seen:
            seen.add(line)
            lines.append(line)
    return "\n".join(lines)


def limit_detection_text(text: str, *, max_chars: int = 2000) -> str:
    body = str(text or "").strip()
    if len(body) <= max_chars:
        return body
    return body[:max_chars].rstrip()


def pdf_is_encrypted(raw_bytes: bytes) -> bool:
    return b"/Encrypt" in raw_bytes


def pdf_looks_like_scanned_image(raw_bytes: bytes) -> bool:
    if b"/Subtype /Image" not in raw_bytes:
        return False
    preview = extract_pdf_text(raw_bytes)
    normalized = re.sub(r"[A-Za-z0-9_./:-]", "", preview).strip()
    return len(normalized) < 8


def extract_preview_text(envelope: OfficialDataFileEnvelope) -> str:
    if envelope.ext == ".csv":
        return matrix_to_preview_text(read_csv_matrix(envelope.raw_bytes), limit=16)
    if envelope.ext == ".xlsx":
        return matrix_to_preview_text(read_xlsx_matrix(envelope.raw_bytes), limit=16)
    if envelope.ext == ".pdf":
        return limit_detection_text(extract_pdf_text(envelope.raw_bytes))
    return ""


def extract_matrix(envelope: OfficialDataFileEnvelope) -> list[list[str]]:
    if envelope.ext == ".csv":
        return read_csv_matrix(envelope.raw_bytes)
    if envelope.ext == ".xlsx":
        return read_xlsx_matrix(envelope.raw_bytes)
    return []
