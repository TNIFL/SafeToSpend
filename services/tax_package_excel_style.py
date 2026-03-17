# SafeToSpend/services/tax_package_excel_style.py
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from typing import Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


# =========================
# Header label (KO)
# =========================
HEADER_KO: dict[str, str] = {
    # transactions
    "tx_id": "거래 ID",
    "occurred_at_kst": "거래시각(KST)",
    "date_kst": "거래일(KST)",
    "direction": "구분(입·출금)",
    "amount_krw": "금액(원)",
    "bank_account": "계좌",
    "counterparty": "거래처",
    "memo": "메모",
    "source": "출처",
    "external_hash": "외부 해시",
    "income_label_status": "수입 분류 상태",
    "income_label_confidence": "수입 분류 신뢰도",
    "income_labeled_by": "수입 분류자",
    "expense_label_status": "지출 분류 상태",
    "expense_label_confidence": "지출 분류 신뢰도",
    "expense_labeled_by": "지출 분류자",
    "evidence_requirement": "증빙 필요",
    "evidence_status": "증빙 상태",
    "evidence_note": "증빙 메모",
    "evidence_original_filename": "증빙 파일명",
    "evidence_sha256": "증빙 SHA256",
    "evidence_uploaded_at_kst": "증빙 업로드(KST)",
    "attachment_zip_path": "첨부 경로(ZIP)",

    # evidence index
    "requirement": "요구",
    "status": "상태",
    "note": "메모",
    "file_key": "파일 키",
    "original_filename": "원본 파일명",
    "mime_type": "MIME 타입",
    "size_bytes": "크기(byte)",
    "sha256": "SHA256",
    "uploaded_at_kst": "업로드(KST)",
    "retention_until": "보관 만료",
    "deleted_at_kst": "삭제(KST)",

    # missing list
    "priority": "우선순위",
    "why": "사유",
    "next_action": "다음 액션",

    # checklist
    "group": "분류",
    "required": "필수 여부",
    "item": "항목",
    "where": "발급/준비처",
}


@dataclass
class StyleConfig:
    header_row: int = 1
    header_row_height: float = 34.0  # ✅ 헤더 잘림 방지
    min_col_width: float = 10.0
    max_col_width: float = 70.0
    sample_rows_for_width: int = 600
    apply_cell_borders: bool = True  # ✅ 내부 격자(얇은 테두리)
    table_style_name: str = "TableStyleMedium2"


# theme
_HEADER_FILL = PatternFill("solid", fgColor="111827")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_TEXT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=False)
_WRAP_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
_RIGHT_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=False)

_THIN = Side(style="thin", color="D1D5DB")
_MEDIUM = Side(style="medium", color="111827")


def _sheet_has_real_data(ws: Worksheet) -> bool:
    dim = ws.calculate_dimension()
    if dim == "A1:A1" and (ws["A1"].value is None):
        return False
    return True


def _used_bounds(ws: Worksheet) -> Tuple[int, int, int, int]:
    dim = ws.calculate_dimension()
    start, end = dim.split(":")

    def split_cell(a1: str) -> Tuple[str, int]:
        col = "".join([c for c in a1 if c.isalpha()])
        row = int("".join([c for c in a1 if c.isdigit()]))
        return col, row

    s_col, s_row = split_cell(start)
    e_col, e_row = split_cell(end)

    from openpyxl.utils.cell import column_index_from_string
    return s_row, e_row, column_index_from_string(s_col), column_index_from_string(e_col)


def apply_table_borders(
    ws: Worksheet,
    *,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    inner_grid: bool,
) -> None:
    """
    ✅ 표 영역에만 테두리:
    - 내부: 얇은 격자(thin) (inner_grid=True일 때)
    - 외곽: 두꺼운 테두리(medium) (항상)
    """
    cache: dict[tuple[bool, bool, bool, bool, bool], Border] = {}

    def _border(top: bool, bottom: bool, left: bool, right: bool, inner: bool) -> Border:
        key = (top, bottom, left, right, inner)
        if key in cache:
            return cache[key]

        # 내부 격자 ON이면 기본 thin, OFF면 내부는 None(겹치지 않게)
        base = _THIN if inner else Side(style=None)

        b = Border(
            left=_MEDIUM if left else base,
            right=_MEDIUM if right else base,
            top=_MEDIUM if top else base,
            bottom=_MEDIUM if bottom else base,
        )
        cache[key] = b
        return b

    for r in range(min_row, max_row + 1):
        is_top = (r == min_row)
        is_bottom = (r == max_row)
        for c in range(min_col, max_col + 1):
            is_left = (c == min_col)
            is_right = (c == max_col)
            ws.cell(row=r, column=c).border = _border(is_top, is_bottom, is_left, is_right, inner_grid)


def _localize_header(ws: Worksheet, header_row: int, min_col: int, max_col: int) -> None:
    for c in range(min_col, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        v = cell.value
        if isinstance(v, str) and v in HEADER_KO:
            cell.value = HEADER_KO[v]


def _format_cell_by_value(cell) -> None:
    v = cell.value
    if v is None:
        return

    if isinstance(v, datetime):
        if v.hour or v.minute or v.second:
            cell.number_format = "yyyy-mm-dd hh:mm"
        else:
            cell.number_format = "yyyy-mm-dd"
        cell.alignment = _CENTER_ALIGN
        return

    if isinstance(v, date):
        cell.number_format = "yyyy-mm-dd"
        cell.alignment = _CENTER_ALIGN
        return

    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if (cell.number_format or "General") == "General":
            if isinstance(v, float) and (abs(v - int(v)) > 1e-9):
                cell.number_format = "#,##0.00"
            else:
                cell.number_format = "#,##0"
        cell.alignment = _RIGHT_ALIGN
        return

    # text
    cell.alignment = _TEXT_ALIGN


def _auto_width(ws: Worksheet, header_row: int, max_row: int, min_col: int, max_col: int, cfg: StyleConfig) -> None:
    data_start = header_row + 1
    data_end = min(max_row, data_start + cfg.sample_rows_for_width)

    for col in range(min_col, max_col + 1):
        letter = get_column_letter(col)
        header_val = ws.cell(row=header_row, column=col).value
        header = "" if header_val is None else str(header_val)

        max_len = len(header)
        for r in range(data_start, data_end + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            s = str(v)
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                s = f"{int(v):,}" if float(v).is_integer() else f"{v:,}"
            elif isinstance(v, (date, datetime)):
                s = s[:16]
            max_len = max(max_len, len(s))

        width = max(cfg.min_col_width, min(cfg.max_col_width, max_len * 1.2 + 4))
        ws.column_dimensions[letter].width = width


def _ensure_table(ws: Worksheet, ref: str, table_name: str, style_name: str) -> None:
    if ws.tables:
        return

    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def style_sheet(ws: Worksheet, *, table_prefix: str, cfg: Optional[StyleConfig] = None) -> None:
    cfg = cfg or StyleConfig()

    if not _sheet_has_real_data(ws):
        ws.sheet_view.showGridLines = False
        return

    min_row, max_row, min_col, max_col = _used_bounds(ws)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 110

    if max_row < cfg.header_row or max_col < 1:
        return

    # ✅ 헤더 한글화(키 -> 라벨)
    _localize_header(ws, cfg.header_row, min_col, max_col)

    # header height(잘림 방지)
    ws.row_dimensions[cfg.header_row].height = cfg.header_row_height

    # filter + freeze
    ref = f"{get_column_letter(min_col)}{cfg.header_row}:{get_column_letter(max_col)}{max_row}"
    ws.auto_filter.ref = ref
    ws.freeze_panes = ws.cell(row=cfg.header_row + 1, column=min_col).coordinate

    # header style
    for col in range(min_col, max_col + 1):
        c = ws.cell(row=cfg.header_row, column=col)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = _HEADER_ALIGN

    # body formats
    for r in range(cfg.header_row + 1, max_row + 1):
        for col in range(min_col, max_col + 1):
            _format_cell_by_value(ws.cell(row=r, column=col))

    # table
    safe_name = f"{table_prefix}_{(ws.title or 'sheet')}".replace(" ", "_").replace("-", "_")
    safe_name = "".join(ch for ch in safe_name if (ch.isalnum() or ch == "_"))[:50] or f"{table_prefix}_sheet"
    _ensure_table(ws, ref, safe_name, cfg.table_style_name)

    # ✅ 내부 격자 + 외곽 두껍게(표 영역에만)
    apply_table_borders(
        ws,
        min_row=cfg.header_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        inner_grid=cfg.apply_cell_borders,
    )

    # widths
    _auto_width(ws, cfg.header_row, max_row, min_col, max_col, cfg)


def style_workbook(wb: Workbook, *, table_prefix: str, cfg: Optional[StyleConfig] = None) -> None:
    cfg = cfg or StyleConfig()
    for idx, ws in enumerate(wb.worksheets, start=1):
        style_sheet(ws, table_prefix=f"{table_prefix}{idx}", cfg=cfg)
