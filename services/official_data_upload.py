from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from core.extensions import db
from domain.models import OfficialDataDocument
from services.official_data_parser_registry import DOCUMENT_TYPE_LABELS, identify_official_data_document
from services.official_data_parsers import PARSER_VERSION, parse_official_data_document
from services.official_data_store import (
    StoredOfficialDataFile,
    delete_official_data_file,
    resolve_official_data_path,
    store_official_data_file,
)


USER_STATUS_LABELS = {
    "parsed": "반영 가능",
    "needs_review": "검토 필요",
    "unsupported": "미지원 형식",
    "failed": "읽기 실패",
}

VERIFICATION_LABELS = {
    "not_verified": "검증 미실시",
    "verified": "검증 완료",
    "verification_failed": "검증 실패",
}

TRUST_GRADE_LABELS = {
    "A": "신뢰도 높음",
    "B": "구조 확인됨",
    "C": "검토 우선",
    "D": "반영 보류",
}

STRUCTURE_LABELS = {
    "passed": "구조 확인됨",
    "needs_review": "구조 검토 필요",
    "unsupported": "지원 범위 아님",
    "failed": "읽기 실패",
    "unknown": "구조 미확인",
}


@dataclass(frozen=True)
class OfficialDataUploadResult:
    document: OfficialDataDocument
    status_label: str
    status_reason: str


def _extension(filename: str) -> str:
    return Path(filename).suffix.lower()


def _read_csv_rows(path: Path) -> list[list[str]]:
    raw = path.read_bytes()
    decoded = None
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            decoded = raw.decode(encoding)
            break
        except Exception:
            continue
    if decoded is None:
        decoded = raw.decode("utf-8", errors="replace")
    reader = csv.reader(decoded.replace("\r\n", "\n").replace("\r", "\n").split("\n"))
    return [[cell.strip() for cell in row] for row in reader]


def _read_xlsx_rows(path: Path) -> list[list[str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if value is None else str(value).strip() for value in row])
    return rows


def _extract_pdf_text(path: Path) -> str:
    raw = path.read_bytes()
    text = raw.decode("utf-8", errors="ignore")
    if len(re.sub(r"\s+", "", text)) < 20:
        text = raw.decode("latin-1", errors="ignore")
    return re.sub(r"\s+", " ", text).strip()


def _load_parse_inputs(stored: StoredOfficialDataFile) -> tuple[list[list[str]] | None, str]:
    ext = _extension(stored.original_filename)
    path = stored.abs_path
    if ext == ".csv":
        return _read_csv_rows(path), ""
    if ext == ".xlsx":
        return _read_xlsx_rows(path), ""
    if ext == ".pdf":
        return None, _extract_pdf_text(path)
    return None, ""


def _parse_status_reason(parse_status: str, fallback: str | None = None) -> str:
    if fallback:
        return fallback
    defaults = {
        "parsed": "핵심 항목을 읽어 반영 후보로 저장했습니다.",
        "needs_review": "핵심 항목이 불완전해 검토가 필요합니다.",
        "unsupported": "현재 지원 범위 밖 문서 또는 형식입니다.",
        "failed": "문서를 읽는 중 오류가 발생했습니다.",
    }
    return defaults.get(parse_status, "처리 결과를 확인해 주세요.")


def _friendly_failure_reason(exc: Exception) -> str:
    text = (str(exc) or "").strip()
    lowered = text.lower()
    if "zip" in lowered and "file" in lowered:
        return "엑셀 파일 형식을 읽지 못했습니다."
    if "openpyxl" in lowered:
        return "엑셀 파일을 읽는 중 오류가 발생했습니다."
    if "pdf" in lowered:
        return "PDF 문서를 읽는 중 오류가 발생했습니다."
    return "문서를 읽는 중 오류가 발생했습니다."


def _trust_grade_for(parse_status: str, structure_validation_status: str) -> str:
    if parse_status == "parsed" and structure_validation_status == "passed":
        return "B"
    if parse_status == "needs_review":
        return "C"
    return "D"


def _summary_payload(summary: dict[str, Any] | None, reason: str) -> dict[str, Any]:
    payload = dict(summary or {})
    payload.setdefault("status_reason", reason)
    payload.setdefault("display_summary", [])
    return payload


def process_official_data_upload(*, user_pk: int, uploaded_file) -> OfficialDataUploadResult:
    stored = store_official_data_file(user_pk=user_pk, file=uploaded_file)
    document: OfficialDataDocument | None = None
    try:
        rows, extracted_text = _load_parse_inputs(stored)
        decision = identify_official_data_document(
            extension=_extension(stored.original_filename),
            rows=rows,
            extracted_text=extracted_text,
        )

        parse_status = "unsupported"
        structure_validation_status = "unsupported"
        reference_date: date | None = None
        summary: dict[str, Any] | None = None

        if decision.registry_status == "identified" and decision.document_type:
            parsed = parse_official_data_document(
                document_type=decision.document_type,
                rows=rows,
                extracted_text=extracted_text,
            )
            parse_status = parsed["parse_status"]
            structure_validation_status = parsed["structure_validation_status"]
            reference_date = parsed.get("reference_date")
            summary = parsed.get("summary") or {}
            status_reason = _parse_status_reason(parse_status, summary.get("status_reason"))
        elif decision.registry_status == "needs_review":
            parse_status = "needs_review"
            structure_validation_status = "needs_review"
            status_reason = decision.reason
            summary = {"display_summary": [], "status_reason": status_reason}
        else:
            parse_status = "unsupported"
            structure_validation_status = "unsupported"
            status_reason = decision.reason
            summary = {"display_summary": [], "status_reason": status_reason}

        trust_grade = _trust_grade_for(parse_status, structure_validation_status)

        document = OfficialDataDocument(
            user_pk=user_pk,
            document_type=decision.document_type,
            source_authority=decision.source_authority,
            raw_file_key=stored.raw_file_key,
            original_filename=stored.original_filename,
            mime_type=stored.mime_type,
            size_bytes=stored.size_bytes,
            sha256=stored.sha256,
            reference_date=reference_date,
            parse_status=parse_status,
            verification_status="not_verified",
            structure_validation_status=structure_validation_status,
            trust_grade=trust_grade,
            extracted_key_summary_json=_summary_payload(summary, status_reason),
            parser_version=PARSER_VERSION,
        )
        db.session.add(document)
        db.session.commit()
        return OfficialDataUploadResult(document=document, status_label=USER_STATUS_LABELS[parse_status], status_reason=status_reason)
    except ValueError:
        raise
    except Exception as exc:
        db.session.rollback()
        status_reason = _friendly_failure_reason(exc)
        failed_document = OfficialDataDocument(
            user_pk=user_pk,
            document_type=None,
            source_authority=None,
            raw_file_key=stored.raw_file_key,
            original_filename=stored.original_filename,
            mime_type=stored.mime_type,
            size_bytes=stored.size_bytes,
            sha256=stored.sha256,
            reference_date=None,
            parse_status="failed",
            verification_status="not_verified",
            structure_validation_status="failed",
            trust_grade="D",
            extracted_key_summary_json={"display_summary": [], "status_reason": status_reason},
            parser_version=PARSER_VERSION,
        )
        db.session.add(failed_document)
        db.session.commit()
        return OfficialDataUploadResult(document=failed_document, status_label=USER_STATUS_LABELS["failed"], status_reason=status_reason)


def official_data_document_to_view_model(document: OfficialDataDocument) -> dict[str, Any]:
    summary = dict(document.extracted_key_summary_json or {})
    display_summary = summary.get("display_summary")
    if not isinstance(display_summary, list):
        display_summary = []
    return {
        "id": int(document.id),
        "document_type": document.document_type,
        "document_type_label": DOCUMENT_TYPE_LABELS.get(document.document_type or "", "문서 판별 전"),
        "source_authority": document.source_authority or "확인 전",
        "reference_date": document.reference_date.isoformat() if document.reference_date else "확인 전",
        "parse_status": document.parse_status,
        "parse_status_label": USER_STATUS_LABELS.get(document.parse_status, "처리 결과 확인"),
        "verification_status": document.verification_status,
        "verification_status_label": VERIFICATION_LABELS.get(document.verification_status, "검증 미실시"),
        "structure_validation_status": document.structure_validation_status,
        "structure_validation_label": STRUCTURE_LABELS.get(document.structure_validation_status, "구조 미확인"),
        "trust_grade": document.trust_grade,
        "trust_grade_label": TRUST_GRADE_LABELS.get(document.trust_grade, "반영 보류"),
        "original_filename": document.original_filename,
        "summary_items": display_summary,
        "status_reason": summary.get("status_reason") or _parse_status_reason(document.parse_status),
        "parser_version": document.parser_version,
        "created_at": document.created_at.strftime("%Y-%m-%d %H:%M") if document.created_at else "",
    }


def query_official_data_documents(*, user_pk: int, limit: int = 50) -> list[OfficialDataDocument]:
    return (
        OfficialDataDocument.query.filter_by(user_pk=user_pk)
        .order_by(OfficialDataDocument.created_at.desc(), OfficialDataDocument.id.desc())
        .limit(limit)
        .all()
    )


def list_official_data_documents(*, user_pk: int, limit: int = 50) -> list[dict[str, Any]]:
    rows = query_official_data_documents(user_pk=user_pk, limit=limit)
    return [official_data_document_to_view_model(row) for row in rows]


def get_official_data_document_for_user(*, user_pk: int, document_id: int) -> OfficialDataDocument | None:
    return OfficialDataDocument.query.filter_by(id=document_id, user_pk=user_pk).first()


def get_official_data_download_path(*, document: OfficialDataDocument) -> Path:
    return resolve_official_data_path(document.raw_file_key)


def delete_official_data_document_file(*, document: OfficialDataDocument) -> None:
    delete_official_data_file(document.raw_file_key)
