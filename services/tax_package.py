from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func, or_

from core.extensions import db
from core.time import utcnow
from domain.models import (
    BankAccountLink,
    EvidenceItem,
    ExpenseLabel,
    ImportJob,
    IncomeLabel,
    OfficialDataDocument,
    ReferenceMaterialItem,
    SafeToSpendSettings,
    TaxBufferLedger,
    Transaction,
    User,
)
from services.onboarding import build_onboarding_reflection
from services.official_data_upload import official_data_document_to_view_model
from services.evidence_vault import resolve_file_path
from services.reference_material_upload import reference_material_to_view_model
from services.cross_validation import (
    build_cross_validation_context,
    build_official_document_cross_validation,
)
from services.transaction_origin import (
    get_transaction_provider_label,
    get_transaction_source_label,
    resolve_transaction_origin,
)


KST = ZoneInfo("Asia/Seoul")
HEADER_FILL = PatternFill("solid", fgColor="E8EEF8")
HEADER_FONT = Font(bold=True)
GOOD_FILL = PatternFill("solid", fgColor="E7F5EA")
WARN_FILL = PatternFill("solid", fgColor="FFF4E5")
BAD_FILL = PatternFill("solid", fgColor="FDECEC")
TOP_ALIGN = Alignment(vertical="top", wrap_text=True)
PACKAGE_VERSION = "세무사 패키지 v2 4차"
EVIDENCE_ATTACHMENTS_DIR = "attachments/evidence"
DEFAULT_TAX_PACKAGE_PROFILE_CODE = "common"
ALL_WORKBOOK_KEYS = (
    "summary",
    "business_status",
    "transactions",
    "evidence",
    "withholding",
    "review",
    "attachments",
    "vat",
    "nhis_pension",
    "reference",
)
WORKBOOK_FILENAME_MAP = {
    "summary": "00_패키지요약.xlsx",
    "business_status": "01_사업_상태_요약.xlsx",
    "transactions": "03_거래원장.xlsx",
    "evidence": "04_증빙상태표.xlsx",
    "withholding": "05_원천징수_기납부세액_요약.xlsx",
    "review": "06_세무사_확인필요목록.xlsx",
    "attachments": "07_첨부인덱스.xlsx",
    "vat": "08_부가세_자료_요약.xlsx",
    "nhis_pension": "09_건보_연금_요약.xlsx",
    "reference": "10_참고자료_요약.xlsx",
}


def _build_type_rank(*item_types: str) -> dict[str, int]:
    return {item_type: idx for idx, item_type in enumerate(item_types)}


@dataclass(frozen=True)
class TaxPackageProfile:
    code: str
    display_name: str
    archive_label: str
    page_description: str
    event_summary_note: str
    included_workbooks: tuple[str, ...]
    workbook_badge_order: tuple[str, ...]
    workbook_badge_tones: dict[str, str]
    summary_row_order: tuple[str, ...]
    review_flow_lines: tuple[str, ...]
    page_flow_lines: tuple[str, ...]
    review_type_rank: dict[str, int]
    review_priority_overrides: dict[str, tuple[int, str, str]] = field(default_factory=dict)
    official_keyword_groups: tuple[tuple[str, ...], ...] = ()
    reference_keyword_groups: tuple[tuple[str, ...], ...] = ()


TAX_PACKAGE_PROFILES: dict[str, TaxPackageProfile] = {
    "common": TaxPackageProfile(
        code="common",
        display_name="공통형",
        archive_label="",
        page_description="거래·증빙·원천징수·부가세·건보/연금·참고자료를 함께 담는 공통 세무사 전달 패키지",
        event_summary_note="기본 공통형입니다. 종합 검토를 기준으로 요약과 확인필요목록을 함께 봅니다.",
        included_workbooks=ALL_WORKBOOK_KEYS,
        workbook_badge_order=("summary", "review", "business_status", "withholding", "vat", "nhis_pension", "transactions", "evidence", "attachments", "reference"),
        workbook_badge_tones={"summary": "good", "review": "warn"},
        summary_row_order=(
            "user_name",
            "period",
            "generated_at",
            "review_start",
            "primary_review_flow",
            "immediate_recheck_count",
            "evidence_review",
            "withholding_status",
            "vat_status",
            "nhis_pension_status",
            "reference_review_status",
            "cross_validation_notice",
            "cross_validation_match",
            "cross_validation_partial",
            "cross_validation_review_needed",
            "cross_validation_mismatch",
            "cross_validation_unavailable",
            "tx_total",
            "sum_in_total",
            "sum_out_total",
            "expense_business_total",
            "evidence_attached_count",
            "official_data_total",
            "reference_count",
            "official_data_parsed_count",
            "official_data_review_count",
            "reference_note",
        ),
        review_flow_lines=(
            "- 1) 00_패키지요약.xlsx : 검토 시작 / 핵심 상태 / 공식자료 교차검증 요약 확인",
            "- 2) 06_세무사_확인필요목록.xlsx : 우선확인순서 기준으로 먼저 연락할 항목 확인",
            "- 3) 01_사업_상태_요약.xlsx : 상태값과 출처/확인 수준 확인",
            "- 4) 05_원천징수_기납부세액_요약.xlsx : 누락 여부와 합계 기준 확인",
            "- 5) 08_부가세_자료_요약.xlsx / 09_건보_연금_요약.xlsx : 해당 월 추가 요청 포인트 확인",
            "- 6) 03_거래원장.xlsx / 04_증빙상태표.xlsx / 07_첨부인덱스.xlsx : 거래·증빙·첨부 순서로 상세 확인",
            "- 7) 10_참고자료_요약.xlsx : 보조 설명과 차이 설명 마지막 확인",
            "- 8) 00_패키지요약.xlsx 내부 공식자료 시트 : 공식자료 목록/상태/핵심값 최종 확인",
        ),
        page_flow_lines=(
            "1) ZIP을 내려받아 압축 해제",
            "2) 00_패키지요약.xlsx에서 검토 시작 / 핵심 재확인 상태 확인",
            "3) 06_세무사_확인필요목록.xlsx → 01_사업_상태_요약.xlsx → 05_원천징수/기납부세액 요약 순서로 확인",
            "4) 08_부가세 / 09_건보·연금 → 03_거래원장 / 04_증빙상태표 / 07_첨부인덱스 → 10_참고자료 순서로 상세 확인",
        ),
        review_type_rank=_build_type_rank(
            "거래검토",
            "부가세자료누락",
            "부가세재확인",
            "원천징수자료누락",
            "기납부세액자료누락",
            "공식자료교차검증재확인",
            "공식자료재확인",
            "건보자료누락",
            "연금자료누락",
            "증빙누락",
            "증빙검토",
            "참고자료검토",
            "사용자상태확인",
            "건보연금상태확인",
        ),
    ),
    "comprehensive_income": TaxPackageProfile(
        code="comprehensive_income",
        display_name="종합소득세용",
        archive_label="종합소득세용",
        page_description="종합소득세 검토 흐름에 맞춰 원천징수·기납부세액, 공식자료 교차검증, 거래·증빙을 앞쪽에 두는 파생 패키지",
        event_summary_note="종합소득세 검토 보조용입니다. 원천징수·기납부세액과 공식자료 교차검증 재확인을 앞쪽에 배치합니다.",
        included_workbooks=ALL_WORKBOOK_KEYS,
        workbook_badge_order=("summary", "review", "business_status", "withholding", "transactions", "evidence", "attachments", "reference", "vat", "nhis_pension"),
        workbook_badge_tones={"summary": "good", "review": "warn", "business_status": "good", "withholding": "warn"},
        summary_row_order=(
            "user_name",
            "period",
            "generated_at",
            "review_start",
            "primary_review_flow",
            "immediate_recheck_count",
            "withholding_status",
            "cross_validation_overview",
            "evidence_review",
            "reference_review_status",
            "vat_status",
            "nhis_pension_status",
            "cross_validation_notice",
            "cross_validation_match",
            "cross_validation_partial",
            "cross_validation_review_needed",
            "cross_validation_mismatch",
            "cross_validation_unavailable",
            "tx_total",
            "sum_in_total",
            "sum_out_total",
            "expense_business_total",
            "evidence_attached_count",
            "official_data_total",
            "reference_count",
            "official_data_parsed_count",
            "official_data_review_count",
            "reference_note",
        ),
        review_flow_lines=(
            "- 1) 00_패키지요약.xlsx : 종합소득세 검토 시작 / 원천징수·기납부세액 / 공식자료 교차검증 상태 확인",
            "- 2) 06_세무사_확인필요목록.xlsx : 신고 영향이 큰 항목과 교차검증 재확인 항목부터 확인",
            "- 3) 01_사업_상태_요약.xlsx : 사용자 유형·건보·과세 상태의 출처와 확인 수준 확인",
            "- 4) 05_원천징수_기납부세액_요약.xlsx : 원천징수·기납부세액 누락 여부와 합계 기준 확인",
            "- 5) 03_거래원장.xlsx / 04_증빙상태표.xlsx / 07_첨부인덱스.xlsx : 수입·지출·증빙 흐름 상세 확인",
            "- 6) 10_참고자료_요약.xlsx : 보조 설명과 차이 설명 확인",
            "- 7) 08_부가세_자료_요약.xlsx / 09_건보_연금_요약.xlsx : 필요 시 보조 확인",
            "- 8) 00_패키지요약.xlsx 내부 공식자료 시트 : 문서별 교차검증 상태와 사유 최종 확인",
        ),
        page_flow_lines=(
            "1) ZIP을 내려받아 압축 해제",
            "2) 00_패키지요약.xlsx → 06_세무사_확인필요목록.xlsx 순서로 먼저 확인",
            "3) 01_사업_상태_요약.xlsx → 05_원천징수/기납부세액 요약에서 종합소득세 검토 축을 먼저 점검",
            "4) 03_거래원장 / 04_증빙상태표 / 07_첨부인덱스 → 10_참고자료 → 08/09 보조 시트 순서로 상세 확인",
        ),
        review_type_rank=_build_type_rank(
            "거래검토",
            "원천징수자료누락",
            "기납부세액자료누락",
            "공식자료교차검증재확인",
            "공식자료재확인",
            "증빙누락",
            "증빙검토",
            "참고자료검토",
            "부가세자료누락",
            "부가세재확인",
            "건보자료누락",
            "연금자료누락",
            "사용자상태확인",
            "건보연금상태확인",
        ),
        official_keyword_groups=(("홈택스", "원천징수", "납부"), ("건강보험", "연금")),
        reference_keyword_groups=(("원천징수", "기납부", "세액", "수익", "소득"), ("부가세", "매입", "매출")),
    ),
    "vat_review": TaxPackageProfile(
        code="vat_review",
        display_name="부가세용",
        archive_label="부가세용",
        page_description="부가세 검토 보조용으로 과세 상태, 부가세 자료 요약, 부가세 관련 확인필요항목을 앞쪽에 두는 파생 패키지",
        event_summary_note="부가세 검토 보조용입니다. 현재 지원 범위 안의 부가세 관련 요약과 재확인 포인트를 앞쪽에 배치합니다.",
        included_workbooks=tuple(key for key in ALL_WORKBOOK_KEYS if key != "nhis_pension"),
        workbook_badge_order=("summary", "review", "vat", "business_status", "transactions", "evidence", "attachments", "reference", "withholding", "nhis_pension"),
        workbook_badge_tones={"summary": "good", "review": "warn", "vat": "warn", "business_status": "good"},
        summary_row_order=(
            "user_name",
            "period",
            "generated_at",
            "review_start",
            "primary_review_flow",
            "vat_status",
            "vat_recheck_count",
            "vat_material_status",
            "immediate_recheck_count",
            "cross_validation_overview",
            "evidence_review",
            "reference_review_status",
            "withholding_status",
            "nhis_pension_status",
            "cross_validation_notice",
            "cross_validation_match",
            "cross_validation_partial",
            "cross_validation_review_needed",
            "cross_validation_mismatch",
            "cross_validation_unavailable",
            "tx_total",
            "sum_out_total",
            "expense_business_total",
            "official_data_total",
            "official_data_parsed_count",
            "official_data_review_count",
            "reference_note",
        ),
        review_flow_lines=(
            "- 1) 00_패키지요약.xlsx : 부가세 검토 시작 / 부가세 상태 / 부가세 재확인 항목 수 확인",
            "- 2) 06_세무사_확인필요목록.xlsx : 부가세자료누락, 부가세재확인, 공식자료 교차검증 재확인 항목부터 확인",
            "- 3) 08_부가세_자료_요약.xlsx : 최근 신고 여부와 세금계산서·카드·현금영수증 요약 확인",
            "- 4) 01_사업_상태_요약.xlsx : 과세 상태와 출처/확인 수준 확인",
            "- 5) 03_거래원장.xlsx / 04_증빙상태표.xlsx / 07_첨부인덱스.xlsx : 부가세 검토에 필요한 거래·증빙 흐름 상세 확인",
            "- 6) 10_참고자료_요약.xlsx : 부가세 관련 보조 설명과 차이 설명 확인",
            "- 7) 05_원천징수_기납부세액_요약.xlsx : 후순위 보조 확인",
            "- 8) 00_패키지요약.xlsx 내부 공식자료 시트 : 홈택스 계열 공식자료 상태 최종 확인",
        ),
        page_flow_lines=(
            "1) ZIP을 내려받아 압축 해제",
            "2) 00_패키지요약.xlsx → 06_세무사_확인필요목록.xlsx 순서로 부가세 재확인 포인트를 먼저 확인",
            "3) 08_부가세_자료_요약.xlsx → 01_사업_상태_요약.xlsx에서 과세 상태와 자료 요약을 확인",
            "4) 03_거래원장 / 04_증빙상태표 / 07_첨부인덱스 → 10_참고자료 → 05 보조 시트 순서로 상세 확인",
        ),
        review_type_rank=_build_type_rank(
            "부가세자료누락",
            "부가세재확인",
            "거래검토",
            "공식자료교차검증재확인",
            "공식자료재확인",
            "증빙누락",
            "증빙검토",
            "참고자료검토",
            "원천징수자료누락",
            "기납부세액자료누락",
            "건보자료누락",
            "연금자료누락",
            "사용자상태확인",
            "건보연금상태확인",
        ),
        official_keyword_groups=(("홈택스", "세금", "부가세", "원천징수"), ("건강보험", "연금")),
        reference_keyword_groups=(("부가세", "세금계산서", "매입", "매출", "현금영수증", "카드"), ("원천징수", "기납부")),
    ),
    "nhis_pension_check": TaxPackageProfile(
        code="nhis_pension_check",
        display_name="건보·연금 점검용",
        archive_label="건보연금점검용",
        page_description="건강보험·국민연금 관련 상태, 자료 존재 여부, 재확인 포인트를 먼저 점검하는 파생 패키지",
        event_summary_note="건강보험·국민연금 관련 자료와 재확인 포인트를 먼저 점검하는 패키지입니다. 현재 구조상 국민연금은 보조 수준일 수 있습니다.",
        included_workbooks=tuple(key for key in ALL_WORKBOOK_KEYS if key not in {"withholding", "vat"}),
        workbook_badge_order=("summary", "review", "nhis_pension", "business_status", "transactions", "evidence", "attachments", "reference", "withholding", "vat"),
        workbook_badge_tones={"summary": "good", "review": "warn", "nhis_pension": "warn", "business_status": "good"},
        summary_row_order=(
            "user_name",
            "period",
            "generated_at",
            "review_start",
            "primary_review_flow",
            "nhis_pension_status",
            "immediate_recheck_count",
            "cross_validation_overview",
            "evidence_review",
            "reference_review_status",
            "withholding_status",
            "vat_status",
            "cross_validation_notice",
            "cross_validation_match",
            "cross_validation_partial",
            "cross_validation_review_needed",
            "cross_validation_mismatch",
            "cross_validation_unavailable",
            "tx_total",
            "sum_out_total",
            "expense_business_total",
            "official_data_total",
            "official_data_parsed_count",
            "official_data_review_count",
            "reference_note",
        ),
        review_flow_lines=(
            "- 1) 00_패키지요약.xlsx : 건보·연금 점검 시작 / 건보·연금 상태 / 공식자료 교차검증 보조 상태 확인",
            "- 2) 06_세무사_확인필요목록.xlsx : 건보자료누락, 연금자료누락, 건보연금상태확인 항목부터 확인",
            "- 3) 09_건보_연금_요약.xlsx : 건보·연금 자료 존재 여부와 기준 자료를 먼저 확인",
            "- 4) 01_사업_상태_요약.xlsx : 건강보험 상태와 값 출처·확인 수준 확인",
            "- 5) 00_패키지요약.xlsx 내부 공식자료 시트 : 건보 관련 공식자료 상태와 교차검증 사유 확인",
            "- 6) 03_거래원장.xlsx / 04_증빙상태표.xlsx / 07_첨부인덱스.xlsx : 필요 시 거래·증빙·첨부 흐름 상세 확인",
            "- 7) 10_참고자료_요약.xlsx : 건보·연금 관련 보조 설명과 차이 설명 확인",
        ),
        page_flow_lines=(
            "1) ZIP을 내려받아 압축 해제",
            "2) 00_패키지요약.xlsx → 06_세무사_확인필요목록.xlsx 순서로 건보·연금 재확인 포인트를 먼저 확인",
            "3) 09_건보_연금_요약.xlsx → 01_사업_상태_요약.xlsx에서 건강보험 상태와 기준 자료를 확인",
            "4) 00 내부 공식자료 시트 → 03_거래원장 / 04_증빙상태표 / 07_첨부인덱스 → 10_참고자료 순서로 상세 확인",
        ),
        review_type_rank=_build_type_rank(
            "건보자료누락",
            "연금자료누락",
            "건보연금상태확인",
            "공식자료교차검증재확인",
            "공식자료재확인",
            "거래검토",
            "증빙누락",
            "증빙검토",
            "참고자료검토",
            "원천징수자료누락",
            "기납부세액자료누락",
            "부가세자료누락",
            "부가세재확인",
            "사용자상태확인",
        ),
        review_priority_overrides={
            "건보자료누락": (1, "높음", "건강보험 자료 누락 또는 납부 상태 확인 필요"),
            "연금자료누락": (1, "높음", "국민연금 자료 누락 또는 납부 상태 확인 필요"),
            "건보연금상태확인": (2, "중간", "건강보험·국민연금 상태 재확인 필요"),
            "공식자료교차검증재확인": (3, "중간", "건보·연금 관련 공식자료 교차검증 재확인 필요"),
            "공식자료재확인": (4, "중간", "건보·연금 관련 공식자료 상태 재확인 필요"),
            "거래검토": (5, "중간", "관련 거래나 지출 흐름 보조 확인"),
            "증빙누락": (6, "중간", "증빙 누락 또는 증빙 불충분 검토"),
            "증빙검토": (6, "중간", "증빙 누락 또는 증빙 불충분 검토"),
            "참고자료검토": (7, "낮음", "보조 설명 자료 검토"),
            "원천징수자료누락": (8, "낮음", "후순위 보조 자료 확인"),
            "기납부세액자료누락": (8, "낮음", "후순위 보조 자료 확인"),
            "부가세자료누락": (9, "낮음", "후순위 보조 자료 확인"),
            "부가세재확인": (9, "낮음", "후순위 보조 자료 확인"),
            "사용자상태확인": (10, "낮음", "기타 사용자 상태 재확인"),
        },
        official_keyword_groups=(("건강보험", "건보", "국민연금", "연금", "자격", "납부"), ("홈택스", "원천징수", "세금")),
        reference_keyword_groups=(("건강보험", "건보", "국민연금", "연금"), ("원천징수", "기납부", "부가세")),
    ),
}


@dataclass(frozen=True)
class PackageStats:
    month_key: str
    period_start_kst: str
    period_end_kst: str
    generated_at_kst: str
    tx_total: int
    tx_in_count: int
    tx_out_count: int
    sum_in_total: int
    sum_out_total: int
    income_included_total: int
    income_excluded_non_income_total: int
    income_unknown_count: int
    expense_business_total: int
    expense_personal_total: int
    expense_mixed_total: int
    expense_unknown_total: int
    evidence_missing_required_count: int
    evidence_missing_required_amount: int
    evidence_missing_maybe_count: int
    evidence_missing_maybe_amount: int
    evidence_attached_count: int
    review_needed_count: int
    tax_rate: float
    tax_buffer_total: int
    tax_buffer_target: int
    tax_buffer_shortage: int
    official_data_total: int = 0
    official_data_parsed_count: int = 0
    official_data_review_count: int = 0
    official_data_unsupported_count: int = 0
    official_data_failed_count: int = 0


@dataclass(frozen=True)
class PackageSnapshot:
    root_name: str
    download_name: str
    display_name: str
    stats: PackageStats
    transactions: list[dict[str, Any]]
    evidences: list[dict[str, Any]]
    review_items: list[dict[str, Any]]
    evidence_missing_items: list[dict[str, Any]]
    review_trade_items: list[dict[str, Any]]
    included_source_labels: list[str]
    official_documents: list[dict[str, Any]] = field(default_factory=list)
    business_status_rows: list[dict[str, Any]] = field(default_factory=list)
    withholding_summary_rows: list[dict[str, Any]] = field(default_factory=list)
    vat_summary_rows: list[dict[str, Any]] = field(default_factory=list)
    nhis_pension_summary_rows: list[dict[str, Any]] = field(default_factory=list)
    reference_material_rows: list[dict[str, Any]] = field(default_factory=list)


def get_tax_package_profile(profile_code: str | None) -> TaxPackageProfile:
    code = (profile_code or "").strip() or DEFAULT_TAX_PACKAGE_PROFILE_CODE
    return TAX_PACKAGE_PROFILES.get(code, TAX_PACKAGE_PROFILES[DEFAULT_TAX_PACKAGE_PROFILE_CODE])


def list_tax_package_profiles() -> list[TaxPackageProfile]:
    return [TAX_PACKAGE_PROFILES[key] for key in ("common", "comprehensive_income", "vat_review", "nhis_pension_check")]


def _workbook_filename(workbook_key: str) -> str:
    return WORKBOOK_FILENAME_MAP[workbook_key]


def describe_tax_package_profile(profile_code: str | None) -> dict[str, Any]:
    profile = get_tax_package_profile(profile_code)
    return {
        "code": profile.code,
        "display_name": profile.display_name,
        "page_description": profile.page_description,
        "event_summary_note": profile.event_summary_note,
        "included_workbooks": [
            {
                "key": key,
                "filename": _workbook_filename(key),
                "tone": profile.workbook_badge_tones.get(key, "default"),
            }
            for key in profile.workbook_badge_order
            if key in profile.included_workbooks
        ],
        "page_flow_lines": list(profile.page_flow_lines),
    }


def _month_range_kst_naive(month_key: str) -> tuple[datetime, datetime]:
    y, m = month_key.split("-")
    y = int(y)
    m = int(m)
    start = datetime(y, m, 1, 0, 0, 0)
    if m == 12:
        end = datetime(y + 1, 1, 1, 0, 0, 0)
    else:
        end = datetime(y, m + 1, 1, 0, 0, 0)
    return start, end


def _get_settings(user_pk: int) -> SafeToSpendSettings | None:
    return SafeToSpendSettings.query.get(user_pk)


def _tax_rate(settings: SafeToSpendSettings | None) -> float:
    rate = float(getattr(settings, "default_tax_rate", 0.15) or 0.15)
    if rate > 1:
        rate = rate / 100.0
    return max(0.0, min(rate, 0.95))


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except Exception:
        return default


def _krw(value: int) -> str:
    return f"{int(value or 0):,}원"


def _to_kst(dt: datetime | None) -> datetime | None:
    if not dt:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=KST)
    return dt.astimezone(KST)


def _fmt_kst(dt: datetime | None, fmt: str) -> str:
    converted = _to_kst(dt)
    return converted.strftime(fmt) if converted else ""


def _safe_package_label(value: str | None, fallback: str) -> str:
    text = (value or "").strip() or fallback
    text = re.sub(r'[\\/:*?"<>|]+', "_", text)
    text = re.sub(r"\s+", "_", text)
    text = text.strip("._")
    return text or fallback


def _safe_attachment_name(filename: str | None, fallback: str) -> str:
    text = (filename or "").strip() or fallback
    text = re.sub(r'[\\/:*?"<>|]+', "_", text)
    text = text.replace("..", ".")
    return text[:160] or fallback


def _month_date_range(month_key: str) -> tuple[date, date]:
    start_dt, end_dt = _month_range_kst_naive(month_key)
    return start_dt.date(), (end_dt.date())


def _official_trust_label(view: dict[str, Any], document: OfficialDataDocument) -> str:
    label = view.get("trust_grade_label") or "반영 보류"
    grade = (document.trust_grade or "").strip()
    return f"{label} ({grade})" if grade else label


def _official_recheck_label(document: OfficialDataDocument) -> str:
    return "예" if document.parse_status != "parsed" or document.verification_status != "verified" else "아니오"


def _official_note(view: dict[str, Any], document: OfficialDataDocument) -> str:
    notes: list[str] = []
    if document.verification_status != "verified":
        notes.append(view.get("verification_status_label") or "검증 미실시")
    if document.parse_status != "parsed":
        reason = (view.get("status_reason") or "").strip()
        if reason:
            notes.append(reason)
    if document.raw_file_key:
        notes.append("원본 파일은 기본 패키지에 포함하지 않습니다")
    return " / ".join(dict.fromkeys(note for note in notes if note))


def _official_summary_text(view: dict[str, Any]) -> str:
    parts: list[str] = []
    for item in view.get("summary_items") or []:
        label = (item.get("label") or "").strip()
        value = (item.get("value") or "").strip()
        if label and value:
            parts.append(f"{label}: {value}")
    return " / ".join(parts)


def _package_cross_validation_status_label(status: str | None) -> str:
    normalized = (status or "").strip()
    return {
        "match": "일치",
        "partial_match": "부분 일치",
        "review_needed": "재확인 필요",
        "mismatch": "불일치",
        "reference_only": "비교 불가",
    }.get(normalized, "비교 불가")


def _package_cross_validation_reason(status: str | None, reason: str | None) -> str:
    normalized = (status or "").strip()
    text = (reason or "").strip()
    if normalized == "reference_only":
        return "비교 가능한 공식자료 범위가 아니어서 교차검증 v1 비교를 하지 않았습니다."
    if normalized == "mismatch":
        return text or "비교 가능한 거래와 차이가 있어 세무사 재확인이 필요합니다."
    if normalized == "review_needed":
        return text or "비교 기준이 부족해 세무사 재확인이 필요합니다."
    if normalized == "partial_match":
        return text or "일부 기준만 맞아 보조 확인이 필요할 수 있습니다."
    if normalized == "match":
        return text or "비교 가능한 거래 기준에서 대체로 일치했습니다."
    return text or "비교 기준이 부족해 세무사 재확인이 필요합니다."


def _package_cross_validation_recheck_label(status: str | None) -> str:
    return "예" if (status or "").strip() in {"review_needed", "mismatch"} else "아니오"


def _official_cross_validation_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    counts = {
        "일치": 0,
        "부분 일치": 0,
        "재확인 필요": 0,
        "불일치": 0,
        "비교 불가": 0,
    }
    for row in rows:
        label = str(row.get("교차검증상태") or "비교 불가").strip() or "비교 불가"
        if label not in counts:
            label = "비교 불가"
        counts[label] += 1
    return counts


def _review_type_sort_order(item_type: str) -> int:
    return {
        "거래검토": 10,
        "부가세자료누락": 20,
        "부가세재확인": 21,
        "원천징수자료누락": 30,
        "기납부세액자료누락": 31,
        "원천징수기준기간확인": 32,
        "원천징수지급처참조미확인": 33,
        "총지급액추출불가": 34,
        "원천징수세액추출불가": 35,
        "기납부세액추출불가": 36,
        "공식자료교차검증재확인": 40,
        "공식자료재확인": 41,
        "건보자료누락": 50,
        "연금자료누락": 51,
        "증빙누락": 60,
        "증빙검토": 61,
        "참고자료검토": 70,
        "사용자상태확인": 80,
        "건보연금상태확인": 81,
    }.get(item_type, 999)


def _profile_keyword_sort_key(text: str, groups: tuple[tuple[str, ...], ...]) -> int:
    normalized = str(text or "")
    if not groups:
        return 9
    for idx, keywords in enumerate(groups):
        if any(keyword in normalized for keyword in keywords):
            return idx
    return len(groups) + 1


def _review_related_no_sort_key(value: Any) -> tuple[int, Any]:
    if value in (None, ""):
        return (1, "")
    try:
        return (0, int(value))
    except Exception:
        return (0, str(value))


def _review_item_sort_key(row: dict[str, Any], profile: TaxPackageProfile | None = None) -> tuple[Any, ...]:
    type_rank = _review_type_sort_order(str(row.get("항목유형", "")))
    if profile is not None:
        type_rank = profile.review_type_rank.get(str(row.get("항목유형", "")), type_rank)
    return (
        int(row.get("우선확인순서") or 99),
        type_rank,
        _review_related_no_sort_key(row.get("관련번호")),
        str(row.get("요약설명", "")),
    )


def _official_document_sort_key(row: dict[str, Any], profile: TaxPackageProfile | None = None) -> tuple[Any, ...]:
    status_order = {
        "불일치": 0,
        "재확인 필요": 1,
        "부분 일치": 2,
        "비교 불가": 3,
        "일치": 4,
    }
    read_order = {
        "검토 필요": 0,
        "읽기 실패": 1,
        "미지원 형식": 2,
        "반영 가능": 3,
    }
    focus_text = " ".join(
        [
            str(row.get("문서종류") or ""),
            str(row.get("기관명") or ""),
            str(row.get("교차검증 상태") or row.get("교차검증상태") or ""),
        ]
    )
    return (
        0 if row.get("교차검증 재확인 필요") == "예" or row.get("교차검증재확인필요여부") == "예" else 1,
        status_order.get(str(row.get("교차검증 상태") or row.get("교차검증상태") or ""), 9),
        0 if row.get("재확인필요여부") == "예" else 1,
        _profile_keyword_sort_key(focus_text, profile.official_keyword_groups if profile is not None else ()),
        read_order.get(str(row.get("읽기상태") or ""), 9),
        str(row.get("문서종류") or ""),
        str(row.get("기준일") or ""),
        _review_related_no_sort_key(row.get("자료번호")),
    )


def _reference_material_sort_key(row: dict[str, Any], profile: TaxPackageProfile | None = None) -> tuple[Any, ...]:
    status_order = {
        "official_difference": 0,
        "transaction_difference": 1,
        "no_comparison": 2,
        "reference_only": 3,
        "official_match": 4,
        "transaction_match": 5,
    }
    focus_text = " ".join(
        [
            str(row.get("title") or ""),
            str(row.get("reference_type") or ""),
            str(row.get("linked_official_doc_type") or ""),
            str(row.get("comparison_basis") or ""),
        ]
    )
    return (
        0 if row.get("needs_review") == "예" else 1,
        status_order.get(str(row.get("link_status_key") or ""), 9),
        _profile_keyword_sort_key(focus_text, profile.reference_keyword_groups if profile is not None else ()),
        str(row.get("reported_period") or ""),
        _review_related_no_sort_key(row.get("reference_material_id")),
    )


def _attachment_index_sort_key(row: dict[str, Any]) -> tuple[Any, ...]:
    package_order = {"포함": 0, "기본 제외": 1}
    sensitive_order = {"낮음": 0, "중간": 1, "높음": 2}
    return (
        package_order.get(str(row.get("package_status") or ""), 9),
        sensitive_order.get(str(row.get("contains_sensitive_info") or ""), 9),
        str(row.get("document_type") or ""),
        _review_related_no_sort_key(row.get("related_transaction_id")),
        str(row.get("display_file_name") or ""),
    )


def _evidence_linked_sort_key(row: dict[str, Any]) -> tuple[Any, ...]:
    status_order = {
        "필수 누락": 0,
        "확인 필요": 1,
        "첨부됨": 2,
        "불필요": 3,
    }
    return (
        status_order.get(str(row.get("증빙상태") or ""), 9),
        _review_related_no_sort_key(row.get("거래번호")),
    )


def _withholding_overview_label(source: dict[str, Any]) -> str:
    has_withholding = str(source.get("has_withholding_data", "") or "")
    has_paid = str(source.get("has_paid_tax_data", "") or "")
    other_income = str(source.get("other_income_flag", "") or "")
    if str(source.get("needs_review", "") or "") == "예":
        return "원천징수·기납부세액 재확인 필요"
    if has_withholding == "예" and has_paid == "예":
        return "원천징수·기납부세액 자료 있음"
    if other_income.startswith("예"):
        return "원천징수·기납부세액 재확인 필요"
    if has_withholding == "예" or has_paid == "예":
        return "일부 자료 있음"
    return "자료 미확인"


def _vat_overview_label(source: dict[str, Any]) -> str:
    if str(source.get("needs_review", "") or "") == "예":
        return "부가세 재확인 필요"
    if str(source.get("recent_vat_filing_status", "") or "") == "예":
        return "부가세 신고 자료 확인됨"
    if str(source.get("vat_status", "") or "") == "미확인":
        return "과세 상태 미확인"
    return "부가세 자료 확인 필요"


def _nhis_pension_overview_label(source: dict[str, Any]) -> str:
    if str(source.get("needs_review", "") or "") == "예":
        return "건보·연금 재확인 필요"
    if str(source.get("has_nhis_data", "") or "") == "예" and str(source.get("has_pension_data", "") or "") == "예":
        return "건보·연금 자료 있음"
    if str(source.get("health_insurance_status", "") or "") == "미확인":
        return "건강보험 상태 미확인"
    return "건보·연금 자료 확인 필요"


def _cross_validation_overview_label(counts: dict[str, int]) -> str:
    mismatch = int(counts.get("불일치", 0) or 0)
    review_needed = int(counts.get("재확인 필요", 0) or 0)
    partial = int(counts.get("부분 일치", 0) or 0)
    matched = int(counts.get("일치", 0) or 0)
    unavailable = int(counts.get("비교 불가", 0) or 0)
    if mismatch > 0:
        return f"불일치 {mismatch}건 / 세무사 재확인 필요"
    if review_needed > 0:
        return f"재확인 필요 {review_needed}건"
    if partial > 0:
        return f"부분 일치 {partial}건"
    if matched > 0:
        return f"일치 {matched}건"
    if unavailable > 0:
        return f"비교 불가 {unavailable}건"
    return "비교 가능한 공식자료 없음"


def _vat_material_overview_label(source: dict[str, Any]) -> str:
    parts: list[str] = []
    sales = source.get("tax_invoice_sales_total_krw")
    purchase = source.get("tax_invoice_purchase_total_krw")
    card = source.get("card_purchase_total_krw")
    cash = source.get("cash_receipt_purchase_total_krw")
    if sales not in ("", None):
        parts.append(f"세금계산서 매출 {_krw(int(sales))}")
    if purchase not in ("", None):
        parts.append(f"세금계산서 매입 {_krw(int(purchase))}")
    if card not in ("", None):
        parts.append(f"카드 매입 {_krw(int(card))}")
    if cash not in ("", None):
        parts.append(f"현금영수증 매입 {_krw(int(cash))}")
    if not parts:
        return "세금계산서·매입자료 관련 요약 부족"
    return " / ".join(parts)


def _count_review_items_by_type(review_items: list[dict[str, Any]], item_types: set[str]) -> int:
    return len([row for row in review_items if str(row.get("항목유형") or "") in item_types])


def _build_profiled_package_names(snapshot: PackageSnapshot, profile: TaxPackageProfile) -> tuple[str, str]:
    if profile.code == DEFAULT_TAX_PACKAGE_PROFILE_CODE:
        return snapshot.root_name, snapshot.download_name
    package_label = _safe_package_label(snapshot.display_name, "user")
    root_name = f"세무사전달패키지_{profile.archive_label}_{snapshot.stats.month_key}_{package_label}"
    return root_name, f"{root_name}.zip"


def _label_or_unconfirmed(value: str | None) -> str:
    return (value or "").strip() or "미확인"


def _parse_krw_text(value: Any) -> int | None:
    text = str(value or "").strip()
    if not text:
        return None
    digits = re.sub(r"[^0-9-]", "", text)
    if not digits or digits == "-":
        return None
    try:
        return int(digits)
    except ValueError:
        return None


def _sum_official_amounts(
    documents: list[dict[str, Any]],
    *,
    document_keywords: tuple[str, ...],
    item_keywords: tuple[str, ...],
) -> tuple[bool, int]:
    total = 0
    found = False
    for row in documents:
        document_type = str(row.get("문서종류", ""))
        if not any(keyword in document_type for keyword in document_keywords):
            continue
        for item in row.get("_summary_items") or []:
            label = str(item.get("label") or "")
            if not any(keyword in label for keyword in item_keywords):
                continue
            amount = _parse_krw_text(item.get("value"))
            if amount is None:
                continue
            total += amount
            found = True
    return found, total


def _summary_values(row: dict[str, Any]) -> dict[str, Any]:
    values = row.get("_summary_values")
    return values if isinstance(values, dict) else {}


def _summary_amount(row: dict[str, Any], *keys: str) -> int | None:
    values = _summary_values(row)
    for key in keys:
        if key not in values:
            continue
        amount = _parse_krw_text(values.get(key))
        if amount is not None:
            return amount
    return None


def _has_official_document(documents: list[dict[str, Any]], keywords: tuple[str, ...]) -> bool:
    for row in documents:
        document_type = str(row.get("문서종류", ""))
        if any(keyword in document_type for keyword in keywords):
            return True
    return False


def _collect_period_basis(documents: list[dict[str, Any]], keywords: tuple[str, ...]) -> str:
    period_values: list[str] = []
    for row in documents:
        document_type = str(row.get("문서종류", ""))
        if not any(keyword in document_type for keyword in keywords):
            continue
        period = str(row.get("_period_basis") or row.get("기준일") or "").strip()
        if period and period not in period_values:
            period_values.append(period)
    if not period_values:
        return "미확인"
    if len(period_values) == 1:
        return period_values[0]
    return f"{period_values[0]} 외 {len(period_values) - 1}건"


def _matching_official_documents(documents: list[dict[str, Any]], keywords: tuple[str, ...]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in documents:
        document_type = str(row.get("문서종류", ""))
        if any(keyword in document_type for keyword in keywords):
            rows.append(row)
    return rows


def _collapse_distinct_texts(values: list[str], *, empty: str = "미확인", multi_suffix: str = "건") -> str:
    compact = [value for value in dict.fromkeys(str(value or "").strip() for value in values) if value]
    if not compact:
        return empty
    if len(compact) == 1:
        return compact[0]
    return f"{compact[0]} 외 {len(compact) - 1}{multi_suffix}"


def _collapse_payer_reference(values: list[str]) -> str:
    compact = [value for value in dict.fromkeys(str(value or "").strip() for value in values) if value]
    if not compact:
        return "미확인"
    if len(compact) == 1:
        return compact[0]
    return "복수 지급처 참조"


def _withholding_period_basis(row: dict[str, Any]) -> str:
    values = _summary_values(row)
    period_start = str(values.get("period_start") or "").strip()
    period_end = str(values.get("period_end") or "").strip()
    if period_start and period_end:
        return f"{period_start} ~ {period_end}"
    period_summary = str(values.get("period_summary") or "").strip()
    if period_summary:
        return period_summary
    return str(row.get("_period_basis") or row.get("기준일") or "").strip() or "미확인"


def _sum_transactions_by_direction(transactions: list[dict[str, Any]], direction: str) -> tuple[bool, int]:
    total = 0
    found = False
    for row in transactions:
        if row.get("direction") != direction:
            continue
        total += int(row.get("amount_krw", 0) or 0)
        found = True
    return found, total


def _extract_single_month_token(*texts: str) -> str:
    matches: set[str] = set()
    for text in texts:
        for year, month in re.findall(r"(20\d{2})[./-]?\s?(0?[1-9]|1[0-2])", text or ""):
            matches.add(f"{year}-{int(month):02d}")
    if len(matches) == 1:
        return next(iter(matches))
    return "미확인"


def _extract_single_amount_token(*texts: str) -> int | str:
    matches: set[int] = set()
    for text in texts:
        for raw in re.findall(r"(?<!\d)(\d{1,3}(?:,\d{3})+|\d{4,})(?!\d)\s*원?", text or ""):
            amount = _parse_krw_text(raw)
            if amount is not None:
                matches.add(amount)
    if len(matches) == 1:
        return next(iter(matches))
    return ""


def _classify_reference_material_type(title: str, note: str, material_kind_label: str) -> str:
    text = f"{title} {note}".lower()
    if any(keyword in text for keyword in ("연 수익", "연수익", "연간 수익", "연 매출", "연매출", "연간 매출")):
        return "연 수익표"
    if any(keyword in text for keyword in ("월 수익", "월수익", "월 매출", "월매출", "매출표", "수익표")):
        return "월 수익표"
    if any(keyword in text for keyword in ("비용 정리", "지출 정리", "경비 정리", "비용표", "매입 정리")):
        return "비용 정리표"
    if material_kind_label == "추가설명" or any(keyword in text for keyword in ("메모", "설명", "사유", "비고")):
        return "설명 메모"
    return "기타"


def _link_official_document_type(text: str, documents: list[dict[str, Any]]) -> str:
    lowered = (text or "").lower()
    keyword_groups = (
        ("원천징수", "원천세"),
        ("홈택스 납부", "기납부", "납부세액", "세금 납부"),
        ("건강보험", "건보", "자격"),
        ("국민연금", "연금"),
        ("세금계산서", "전자세금계산서"),
        ("사업용카드", "카드 매입"),
        ("현금영수증",),
        ("부가세", "부가가치세"),
    )
    for keywords in keyword_groups:
        if not any(keyword in lowered for keyword in keywords):
            continue
        for row in documents:
            document_type = str(row.get("문서종류", ""))
            if any(keyword in document_type.lower() for keyword in keywords):
                return document_type
    return ""


def _official_document_total_for_link(document_type: str, documents: list[dict[str, Any]]) -> int | None:
    if "원천징수" in document_type:
        found, total = _sum_official_amounts(
            documents,
            document_keywords=("원천징수",),
            item_keywords=("원천징수세액",),
        )
        return total if found else None
    if "홈택스 납부" in document_type:
        found, total = _sum_official_amounts(
            documents,
            document_keywords=("홈택스 납부",),
            item_keywords=("납부세액",),
        )
        return total if found else None
    if "건강보험" in document_type or "건보" in document_type:
        found, total = _sum_official_amounts(
            documents,
            document_keywords=("건강보험", "건보"),
            item_keywords=("납부", "보험료", "합계"),
        )
        return total if found else None
    if "국민연금" in document_type or document_type == "연금":
        found, total = _sum_official_amounts(
            documents,
            document_keywords=("국민연금", "연금"),
            item_keywords=("납부", "연금", "합계"),
        )
        return total if found else None
    if "세금계산서" in document_type:
        item_keywords = ("매출", "공급가액") if "매출" in document_type else ("매입", "공급가액")
        found, total = _sum_official_amounts(
            documents,
            document_keywords=("세금계산서", "전자세금계산서"),
            item_keywords=item_keywords,
        )
        return total if found else None
    if "사업용카드" in document_type:
        found, total = _sum_official_amounts(
            documents,
            document_keywords=("사업용카드",),
            item_keywords=("매입", "사용금액", "합계"),
        )
        return total if found else None
    if "현금영수증" in document_type:
        found, total = _sum_official_amounts(
            documents,
            document_keywords=("현금영수증",),
            item_keywords=("매입", "사용금액", "합계"),
        )
        return total if found else None
    return None


def _reference_transaction_comparison(
    text: str,
    transactions: list[dict[str, Any]],
) -> tuple[str, str, int | None]:
    lowered = (text or "").lower()
    if any(keyword in lowered for keyword in ("매출", "수입", "수익", "입금", "용역", "매출액")):
        found, total = _sum_transactions_by_direction(transactions, "in")
        return "거래 합계 대비", "월간 수입 합계", total if found else None
    if any(keyword in lowered for keyword in ("매입", "지출", "경비", "비용", "카드", "현금영수증")):
        found, total = _sum_transactions_by_direction(transactions, "out")
        return "거래 합계 대비", "월간 지출 합계", total if found else None
    return "", "", None


def _is_reference_difference_small(reported_amount: int, comparison_amount: int) -> bool:
    tolerance = max(10_000, int(abs(comparison_amount) * 0.05))
    return abs(reported_amount - comparison_amount) <= tolerance


def _resolve_reference_material_comparison(
    *,
    reference_type: str,
    reported_period: str,
    reported_amount: int | str,
    linked_official_doc_type: str,
    official_documents: list[dict[str, Any]],
    transaction_basis: str,
    transaction_target: str,
    transaction_total: int | None,
    package_month_key: str,
) -> dict[str, Any]:
    official_amount = _official_document_total_for_link(linked_official_doc_type, official_documents) if linked_official_doc_type else None
    result = {
        "link_status_key": "no_comparison",
        "link_status": "비교 기준 없음",
        "comparison_basis": "비교 기준 없음",
        "comparison_target": "연결 가능한 공식자료 또는 거래 합계 없음",
        "difference_krw": "",
        "difference_description": "구조화된 비교 기준이 없어 참고용으로만 전달합니다",
        "needs_review": "예",
    }

    if reference_type == "연 수익표":
        result["difference_description"] = "연간 기준 참고자료라 대상 월 패키지와 직접 비교하지 않았습니다"
        return result

    is_month_scoped = reference_type in {"월 수익표", "비용 정리표"}
    if is_month_scoped and reported_period not in {"", "미확인", package_month_key}:
        result["comparison_target"] = package_month_key
        result["difference_description"] = "참고자료 기준 기간이 패키지 대상 월과 달라 직접 비교하지 않았습니다"
        return result

    if linked_official_doc_type and reported_amount != "" and official_amount is not None:
        difference = int(reported_amount) - int(official_amount)
        result.update(
            {
                "link_status_key": "official_match" if _is_reference_difference_small(int(reported_amount), int(official_amount)) else "official_difference",
                "link_status": "공식자료 요약과 대체로 일치" if _is_reference_difference_small(int(reported_amount), int(official_amount)) else "공식자료 요약과 차이 있음",
                "comparison_basis": "공식자료 요약 대비",
                "comparison_target": linked_official_doc_type,
                "difference_krw": difference,
                "difference_description": "기재 금액과 연결된 공식자료 요약값이 대체로 일치합니다" if _is_reference_difference_small(int(reported_amount), int(official_amount)) else "기재 금액과 연결된 공식자료 요약값 차이를 확인해 주세요",
                "needs_review": "아니오" if _is_reference_difference_small(int(reported_amount), int(official_amount)) else "예",
            }
        )
        return result

    transaction_allowed = reference_type in {"월 수익표", "비용 정리표", "설명 메모", "기타"}
    if transaction_allowed and reported_amount != "" and transaction_basis and transaction_total is not None:
        difference = int(reported_amount) - int(transaction_total)
        result.update(
            {
                "link_status_key": "transaction_match" if _is_reference_difference_small(int(reported_amount), int(transaction_total)) else "transaction_difference",
                "link_status": "거래 합계와 대체로 일치" if _is_reference_difference_small(int(reported_amount), int(transaction_total)) else "거래 합계와 차이 있음",
                "comparison_basis": transaction_basis,
                "comparison_target": transaction_target,
                "difference_krw": difference,
                "difference_description": "기재 금액과 대상 월 거래 합계가 대체로 일치합니다" if _is_reference_difference_small(int(reported_amount), int(transaction_total)) else "기재 금액과 대상 월 거래 합계 차이를 확인해 주세요",
                "needs_review": "아니오" if _is_reference_difference_small(int(reported_amount), int(transaction_total)) else "예",
            }
        )
        return result

    if linked_official_doc_type or transaction_basis:
        result.update(
            {
                "link_status_key": "reference_only",
                "link_status": "참고용",
                "comparison_basis": "공식자료 요약 대비" if linked_official_doc_type else transaction_basis,
                "comparison_target": linked_official_doc_type or transaction_target,
                "difference_description": "비교 가능한 연결 축은 찾았지만 구조화 금액 또는 비교 근거가 약해 참고용으로만 전달합니다",
            }
        )
        return result

    if reference_type == "설명 메모":
        result["link_status_key"] = "reference_only"
        result["link_status"] = "참고용"
        result["difference_description"] = "설명 메모는 금액 비교보다 보조 설명 확인이 우선이라 참고용으로 전달합니다"
        return result

    return result


def _collect_business_status_rows(user_pk: int) -> list[dict[str, Any]]:
    reflection = build_onboarding_reflection(user_pk)
    active_links = (
        BankAccountLink.query.filter_by(user_pk=user_pk, is_active=True)
        .order_by(BankAccountLink.created_at.asc(), BankAccountLink.id.asc())
        .all()
    )
    business_aliases = [link.alias.strip() for link in active_links if (link.alias or "").strip() and "사업" in link.alias]

    if reflection["is_business_owner"]:
        business_registration_status = "예(입력값 기준)"
    else:
        business_registration_status = "미확인"

    if business_aliases:
        business_account_usage_status = "사업용 계좌 사용 흔적 있음"
    elif active_links:
        business_account_usage_status = "연결 계좌 있음 · 사업용 여부 미확인"
    else:
        business_account_usage_status = "미연결"

    basis_parts: list[str] = []
    if reflection["has_any_specific"]:
        basis_parts.append("온보딩 입력값")
    elif reflection.get("skipped_at"):
        basis_parts.append("온보딩 건너뛰기")
    else:
        basis_parts.append("온보딩 미설정")
    if active_links:
        basis_parts.append("계좌 별칭 참고")

    note_parts: list[str] = []
    if business_aliases:
        note_parts.append(f"사업용으로 보이는 계좌 별칭: {', '.join(business_aliases[:3])}")
    if not reflection["has_any_specific"]:
        note_parts.append("사용자 유형/건보/과세 상태 중 미확인 항목이 있습니다")

    return [
        {
            "user_type": _label_or_unconfirmed(reflection.get("user_type_label")),
            "user_type_source": "사용자 입력" if reflection["has_specific_user_type"] else "미확인",
            "user_type_confidence": "참고용" if reflection["has_specific_user_type"] else "미확인",
            "health_insurance_status": _label_or_unconfirmed(reflection.get("health_insurance_label")),
            "health_insurance_source": "사용자 입력" if reflection["has_specific_health_insurance"] else "미확인",
            "health_insurance_confidence": "참고용" if reflection["has_specific_health_insurance"] else "미확인",
            "vat_status": _label_or_unconfirmed(reflection.get("vat_status_label")),
            "vat_status_source": "사용자 입력" if reflection["has_specific_vat_status"] else "미확인",
            "vat_status_confidence": "참고용" if reflection["has_specific_vat_status"] else "미확인",
            "business_registration_status": business_registration_status,
            "business_registration_source": "사용자 입력" if reflection["is_business_owner"] else "미확인",
            "business_registration_confidence": "참고용" if reflection["is_business_owner"] else "미확인",
            "business_account_usage_status": business_account_usage_status,
            "business_account_usage_source": "계좌 별칭" if business_aliases else ("계좌 연결 정보" if active_links else "미확인"),
            "business_account_usage_confidence": "참고용" if business_aliases else "미확인",
            "business_card_usage_status": "미확인",
            "business_card_usage_source": "미확인",
            "business_card_usage_confidence": "미확인",
            "onboarding_basis": " / ".join(basis_parts),
            "note": " / ".join(note_parts),
        }
    ]


def _collect_withholding_summary_rows(user_pk: int, documents: list[dict[str, Any]]) -> list[dict[str, Any]]:
    reflection = build_onboarding_reflection(user_pk)
    withholding_documents = _matching_official_documents(documents, ("원천징수",))
    paid_tax_documents = _matching_official_documents(documents, ("홈택스 납부",))

    has_withholding_data = bool(withholding_documents)
    has_paid_tax_data = bool(paid_tax_documents)
    withholding_found, withholding_total = _sum_official_amounts(
        documents,
        document_keywords=("원천징수",),
        item_keywords=("원천징수세액",),
    )
    paid_found, paid_total = _sum_official_amounts(
        documents,
        document_keywords=("홈택스 납부",),
        item_keywords=("납부세액",),
    )
    gross_found = False
    gross_total = 0
    for row in withholding_documents:
        gross_amount = _summary_amount(row, "gross_pay_total_krw")
        if gross_amount is None:
            continue
        gross_total += gross_amount
        gross_found = True

    if reflection["is_freelancer"] or reflection["is_employee_sidejob"]:
        other_income_flag = "예(입력값 기준)"
    else:
        other_income_flag = "미확인"

    basis_parts = []
    if documents:
        basis_parts.append("공식자료 요약값")
    if reflection["has_specific_user_type"]:
        basis_parts.append("온보딩 입력값")

    material_kinds = [
        str(_summary_values(row).get("withholding_material_kind") or "").strip() or "원천징수 관련 자료"
        for row in withholding_documents
    ]
    material_scope = " / ".join(
        part
        for part in (
            _collapse_distinct_texts(material_kinds, multi_suffix="종") if material_kinds else "",
            "홈택스 납부내역" if has_paid_tax_data else "",
        )
        if part
    ) or "미확인"
    withholding_period_basis = _collapse_distinct_texts(
        [_withholding_period_basis(row) for row in withholding_documents],
        multi_suffix="건",
    )
    paid_tax_period_basis = _collapse_distinct_texts(
        [_withholding_period_basis(row) for row in paid_tax_documents],
        multi_suffix="건",
    )
    payer_reference = _collapse_payer_reference(
        [str(_summary_values(row).get("payer_reference") or "").strip() for row in withholding_documents]
    )

    note_parts: list[str] = []
    if has_withholding_data and not withholding_found:
        note_parts.append("원천징수세액 합계가 현재 구조화되지 않아 합계를 확정하지 못했습니다")
    if has_withholding_data and not gross_found:
        note_parts.append("총지급액 합계가 현재 구조화되지 않아 합계를 확정하지 못했습니다")
    if has_paid_tax_data and not paid_found:
        note_parts.append("납부세액 합계가 현재 구조화되지 않아 합계를 확정하지 못했습니다")
    if has_withholding_data and withholding_period_basis == "미확인":
        note_parts.append("원천징수 자료의 기준 기간을 확인하지 못했습니다")
    elif has_withholding_data and " 외 " in withholding_period_basis:
        note_parts.append("원천징수 자료가 복수 기간에 걸쳐 있어 대상 기간 재확인이 필요합니다")
    if has_paid_tax_data and paid_tax_period_basis == "미확인":
        note_parts.append("기납부세액 자료의 기준 기간을 확인하지 못했습니다")
    elif has_paid_tax_data and " 외 " in paid_tax_period_basis:
        note_parts.append("기납부세액 자료가 복수 기간에 걸쳐 있어 대상 기간 재확인이 필요합니다")
    if has_withholding_data and payer_reference == "미확인":
        note_parts.append("지급처 참조를 확인하지 못했습니다")
    elif payer_reference == "복수 지급처 참조":
        note_parts.append("복수 지급처 참조가 있어 단일 지급처로 요약하지 않았습니다")
    if not has_withholding_data and not has_paid_tax_data:
        note_parts.append("대상 월에 포함된 원천징수/기납부세액 공식자료가 없습니다")

    needs_review = False
    if other_income_flag.startswith("예") and (not has_withholding_data or not has_paid_tax_data):
        needs_review = True
    if has_withholding_data and (
        not withholding_found
        or not gross_found
        or withholding_period_basis == "미확인"
        or " 외 " in withholding_period_basis
        or payer_reference in {"미확인", "복수 지급처 참조"}
    ):
        needs_review = True
    if has_paid_tax_data and (not paid_found or paid_tax_period_basis == "미확인" or " 외 " in paid_tax_period_basis):
        needs_review = True

    return [
        {
            "document_scope": material_scope,
            "withholding_period_basis": withholding_period_basis,
            "paid_tax_period_basis": paid_tax_period_basis,
            "has_withholding_data": "예" if has_withholding_data else "아니오",
            "gross_pay_total_krw": gross_total if gross_found else "",
            "withholding_tax_total_krw": withholding_total if withholding_found else "",
            "has_paid_tax_data": "예" if has_paid_tax_data else "아니오",
            "paid_tax_total_krw": paid_total if paid_found else "",
            "payer_reference": payer_reference,
            "other_income_flag": other_income_flag,
            "source_basis": " / ".join(basis_parts) or "미확인",
            "needs_review": "예" if needs_review else "아니오",
            "note": " / ".join(note_parts),
        }
    ]


def _collect_vat_summary_rows(
    user_pk: int,
    documents: list[dict[str, Any]],
    transactions: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    reflection = build_onboarding_reflection(user_pk)
    vat_status = _label_or_unconfirmed(reflection.get("vat_status_label"))

    has_vat_docs = _has_official_document(
        documents,
        ("부가세", "부가가치세", "세금계산서", "전자세금계산서", "사업용카드", "현금영수증"),
    )
    has_vat_filing_doc = any(
        ("부가세" in str(row.get("문서종류", "")) or "부가가치세" in str(row.get("문서종류", "")))
        and "신고" in str(row.get("문서종류", ""))
        for row in documents
    )

    sales_found, sales_total = _sum_official_amounts(
        documents,
        document_keywords=("세금계산서", "전자세금계산서"),
        item_keywords=("매출", "공급가액"),
    )
    purchase_found, purchase_total = _sum_official_amounts(
        documents,
        document_keywords=("세금계산서", "전자세금계산서"),
        item_keywords=("매입", "공급가액"),
    )
    card_found, card_total = _sum_official_amounts(
        documents,
        document_keywords=("사업용카드",),
        item_keywords=("매입", "사용금액", "합계"),
    )
    cash_found, cash_total = _sum_official_amounts(
        documents,
        document_keywords=("현금영수증",),
        item_keywords=("매입", "사용금액", "합계"),
    )

    _, out_total = _sum_transactions_by_direction(transactions, "out")

    basis_parts: list[str] = []
    if reflection["has_specific_vat_status"]:
        basis_parts.append("사용자 입력")
    if has_vat_docs or has_vat_filing_doc:
        basis_parts.append("공식자료 요약값")
    if out_total > 0:
        basis_parts.append("거래 집계 참고")

    if vat_status == "아니에요":
        recent_vat_filing_status = "해당 없음(입력값 기준)"
    elif has_vat_filing_doc:
        recent_vat_filing_status = "예"
    elif has_vat_docs:
        recent_vat_filing_status = "미확인"
    else:
        recent_vat_filing_status = "자료 없음"

    needs_review = False
    note_parts: list[str] = []
    if vat_status == "미확인":
        needs_review = True
        note_parts.append("과세 상태가 미확인입니다")
    if vat_status.startswith("과세") and recent_vat_filing_status in {"미확인", "자료 없음"}:
        needs_review = True
        note_parts.append("과세 상태 입력 기준으로 최근 부가세 신고 여부 재확인이 필요합니다")
    if vat_status.startswith("과세") and not any((sales_found, purchase_found, card_found, cash_found)):
        needs_review = True
        note_parts.append("부가세 판단에 필요한 공식자료 요약값이 부족합니다")
    if out_total > 0 and not any((purchase_found, card_found, cash_found)):
        note_parts.append("거래 지출은 있으나 부가세 관련 공식자료 요약이 부족합니다")

    return [
        {
            "vat_status": vat_status,
            "recent_vat_filing_status": recent_vat_filing_status,
            "tax_invoice_sales_total_krw": sales_total if sales_found else "",
            "tax_invoice_purchase_total_krw": purchase_total if purchase_found else "",
            "card_purchase_total_krw": card_total if card_found else "",
            "cash_receipt_purchase_total_krw": cash_total if cash_found else "",
            "source_basis": " / ".join(dict.fromkeys(part for part in basis_parts if part)) or "미확인",
            "needs_review": "예" if needs_review else "아니오",
            "note": " / ".join(dict.fromkeys(part for part in note_parts if part)),
        }
    ]


def _collect_nhis_pension_summary_rows(
    user_pk: int,
    month_key: str,
    documents: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    reflection = build_onboarding_reflection(user_pk)
    health_insurance_status = _label_or_unconfirmed(reflection.get("health_insurance_label"))

    has_nhis_docs = _has_official_document(documents, ("건강보험", "건보"))
    has_pension_docs = _has_official_document(documents, ("국민연금", "연금"))
    nhis_found, nhis_total = _sum_official_amounts(
        documents,
        document_keywords=("건강보험", "건보"),
        item_keywords=("납부", "보험료", "합계"),
    )
    pension_found, pension_total = _sum_official_amounts(
        documents,
        document_keywords=("국민연금", "연금"),
        item_keywords=("납부", "연금", "합계"),
    )

    basis_parts: list[str] = []
    if reflection["has_specific_health_insurance"]:
        basis_parts.append("사용자 입력")
    if has_nhis_docs or has_pension_docs:
        basis_parts.append("공식자료 요약값")

    needs_review = False
    note_parts: list[str] = []
    if health_insurance_status == "미확인":
        needs_review = True
        note_parts.append("건강보험 상태가 미확인입니다")
    if health_insurance_status != "미확인" and not has_nhis_docs:
        needs_review = True
        note_parts.append("건강보험 관련 공식자료 요약이 없습니다")
    if (reflection["is_freelancer"] or reflection["is_business_owner"]) and not has_pension_docs:
        needs_review = True
        note_parts.append("국민연금 납부 자료 여부 재확인이 필요합니다")

    return [
        {
            "health_insurance_status": health_insurance_status,
            "period_basis": _collect_period_basis(documents, ("건강보험", "건보", "국민연금", "연금")) if (has_nhis_docs or has_pension_docs) else month_key,
            "nhis_total_krw": nhis_total if nhis_found else "",
            "has_nhis_data": "예" if has_nhis_docs else "아니오",
            "has_pension_data": "예" if has_pension_docs else "아니오",
            "pension_total_krw": pension_total if pension_found else "",
            "pension_check_expected": "예" if (reflection["is_freelancer"] or reflection["is_business_owner"]) else "아니오",
            "source_basis": " / ".join(dict.fromkeys(part for part in basis_parts if part)) or "미확인",
            "needs_review": "예" if needs_review else "아니오",
            "note": " / ".join(dict.fromkeys(part for part in note_parts if part)),
        }
    ]


def _collect_reference_material_rows(
    *,
    user_pk: int,
    start_dt: datetime,
    end_dt: datetime,
    official_documents: list[dict[str, Any]],
    transactions: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    package_month_key = start_dt.strftime("%Y-%m")
    items = (
        ReferenceMaterialItem.query.filter(ReferenceMaterialItem.user_pk == user_pk)
        .filter(ReferenceMaterialItem.created_at >= start_dt, ReferenceMaterialItem.created_at < end_dt)
        .order_by(ReferenceMaterialItem.created_at.asc(), ReferenceMaterialItem.id.asc())
        .all()
    )

    rows: list[dict[str, Any]] = []
    for item in items:
        view = reference_material_to_view_model(item)
        merged_text = " ".join(part for part in [view.get("title"), view.get("note")] if part)
        reference_type = _classify_reference_material_type(
            str(view.get("title") or ""),
            str(view.get("note") or ""),
            str(view.get("material_kind_label") or "참고자료"),
        )
        linked_official_doc_type = _link_official_document_type(merged_text, official_documents)
        reported_period = _extract_single_month_token(str(view.get("title") or ""), str(view.get("note") or ""))
        reported_amount = _extract_single_amount_token(str(view.get("title") or ""), str(view.get("note") or ""))
        transaction_basis, transaction_target, transaction_total = _reference_transaction_comparison(merged_text, transactions)
        comparison = _resolve_reference_material_comparison(
            reference_type=reference_type,
            reported_period=reported_period,
            reported_amount=reported_amount,
            linked_official_doc_type=linked_official_doc_type,
            official_documents=official_documents,
            transaction_basis=transaction_basis,
            transaction_target=transaction_target,
            transaction_total=transaction_total,
            package_month_key=package_month_key,
        )

        note_parts: list[str] = []
        if view.get("note"):
            note_parts.append(str(view["note"]))
        if comparison["link_status_key"] in {"reference_only", "no_comparison"}:
            note_parts.append("공식자료 대체가 아니라 보조 설명 자료로 전달합니다")
        elif comparison["link_status_key"] in {"official_difference", "transaction_difference"}:
            note_parts.append("연결된 기준값과 금액 차이가 있어 세무사 확인이 필요합니다")
        elif reported_amount == "":
            note_parts.append("금액은 구조화하지 못해 참고용으로만 표기했습니다")

        rows.append(
            {
                "reference_material_id": int(item.id),
                "title": view.get("title", ""),
                "reference_type": reference_type,
                "reported_period": reported_period,
                "reported_amount_krw": reported_amount,
                "linked_official_doc_type": linked_official_doc_type,
                "link_status": comparison["link_status"],
                "link_status_key": comparison["link_status_key"],
                "comparison_basis": comparison["comparison_basis"],
                "comparison_target": comparison["comparison_target"],
                "difference_krw": comparison["difference_krw"],
                "difference_description": comparison["difference_description"],
                "needs_review": comparison["needs_review"],
                "note": " / ".join(dict.fromkeys(part for part in note_parts if part)),
                "_original_filename": item.original_filename,
                "_attachment_index_key": f"reference-{int(item.id)}",
                "_period_basis": reported_period,
            }
        )

    return rows


def _append_review_item(rows: list[dict[str, Any]], *, item_type: str, related_kind: str, related_no: Any, summary: str, status: str, needed: str, priority: str, note: str = "") -> None:
    rows.append(
        {
            "항목번호": len(rows) + 1,
            "항목유형": item_type,
            "관련자료구분": related_kind,
            "관련번호": related_no,
            "요약설명": summary,
            "현재상태": status,
            "필요한확인내용": needed,
            "우선순위": priority,
            "메모": note,
        }
    )


def _cross_validation_review_profile(row: dict[str, Any]) -> tuple[str, str, str, str] | None:
    status_key = str(row.get("_cross_validation_status_key", "") or "").strip()
    reason = str(row.get("교차검증사유", "") or "").strip()
    if status_key == "mismatch":
        return (
            "공식자료 교차검증 차이 확인 필요",
            "비교 가능한 거래와 차이 있음",
            "공식자료 요약과 비교 가능한 거래 사이에 차이가 있어 세무사 재확인이 필요합니다.",
            reason,
        )
    if status_key != "review_needed":
        return None

    if "비교 가능한 거래나 참고자료가 부족" in reason:
        return (
            "공식자료 교차검증 대상 확인 필요",
            "비교 가능한 거래 없음",
            "비교 가능한 거래가 부족해 교차검증 v1 기준으로는 세무사 재확인이 필요합니다.",
            reason,
        )
    if "금액 또는 날짜가 부족" in reason:
        return (
            "공식자료 교차검증 기준 확인 필요",
            "비교 기준 부족",
            "비교에 필요한 금액 또는 날짜가 부족해 교차검증 v1 기준으로는 세무사 재확인이 필요합니다.",
            reason,
        )
    return (
        "공식자료 교차검증 재확인 필요",
        "재확인 필요",
        "교차검증 v1 기준으로 세무사 재확인이 필요합니다.",
        reason,
    )


def _review_priority_profile(item_type: str) -> tuple[int, str, str]:
    if item_type in {"거래검토", "부가세자료누락", "부가세재확인"}:
        return 1, "높음", "신고 누락 또는 세액 영향이 큰 항목"
    if item_type in {
        "원천징수자료누락",
        "기납부세액자료누락",
        "원천징수기준기간확인",
        "원천징수지급처참조미확인",
        "총지급액추출불가",
        "원천징수세액추출불가",
        "기납부세액추출불가",
    }:
        return 2, "높음", "원천징수·기납부세액 자료 누락 또는 구조 재확인"
    if item_type in {"공식자료재확인", "공식자료교차검증재확인", "건보자료누락", "연금자료누락"}:
        return 3, "중간", "공식자료 재확인 또는 건보·연금 자료 확인 필요"
    if item_type in {"증빙누락", "증빙검토"}:
        return 4, "중간", "증빙 누락 또는 증빙 불충분 검토"
    if item_type in {"참고자료검토"}:
        return 5, "낮음", "참고자료 보조 설명 검토"
    if item_type in {"사용자상태확인", "건보연금상태확인"}:
        return 6, "낮음", "사용자 상태값 미확인"
    return 9, "낮음", "기타 확인 필요 항목"


def _review_priority_profile_for_package(
    item_type: str,
    profile: TaxPackageProfile | None = None,
) -> tuple[int, str, str]:
    if profile is not None and item_type in profile.review_priority_overrides:
        return profile.review_priority_overrides[item_type]
    return _review_priority_profile(item_type)


def _finalize_review_items(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=1):
        priority_order, priority_label, priority_reason = _review_priority_profile(str(row.get("항목유형", "")))
        enriched = dict(row)
        enriched["_original_order"] = idx
        enriched["우선확인순서"] = priority_order
        enriched["우선순위"] = priority_label
        enriched["우선순위기준"] = priority_reason
        normalized.append(enriched)

    normalized.sort(key=lambda item: (_review_item_sort_key(item), item.get("_original_order", 0)))
    for idx, row in enumerate(normalized, start=1):
        row["항목번호"] = idx
        row.pop("_original_order", None)
    return normalized


def _extend_review_items(
    *,
    review_items: list[dict[str, Any]],
    official_documents: list[dict[str, Any]],
    business_status_rows: list[dict[str, Any]],
    withholding_summary_rows: list[dict[str, Any]],
    vat_summary_rows: list[dict[str, Any]],
    nhis_pension_summary_rows: list[dict[str, Any]],
    reference_material_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows = [dict(item) for item in review_items]

    for row in official_documents:
        if row.get("재확인필요여부") != "예":
            continue
        priority = "높음" if row.get("읽기상태") in {"읽기 실패", "미지원 형식"} else "보통"
        _append_review_item(
            rows,
            item_type="공식자료재확인",
            related_kind="공식자료",
            related_no=row.get("자료번호", ""),
            summary=f"{row.get('문서종류', '공식자료')} 재확인 필요",
            status=f"{row.get('읽기상태', '')} / {row.get('검증상태', '')}".strip(" /"),
            needed="공식자료 요약값과 원본 기준이 맞는지 다시 확인해 주세요",
            priority=priority,
            note=row.get("메모", ""),
        )

    for row in official_documents:
        if row.get("교차검증재확인필요여부") != "예":
            continue
        review_profile = _cross_validation_review_profile(row)
        if not review_profile:
            continue
        summary, status, needed, note = review_profile
        _append_review_item(
            rows,
            item_type="공식자료교차검증재확인",
            related_kind="공식자료",
            related_no=row.get("자료번호", ""),
            summary=summary,
            status=status,
            needed=needed,
            priority="보통",
            note=f"{row.get('문서종류', '')} / {note}".strip(" /"),
        )

    business_row = business_status_rows[0] if business_status_rows else {}
    missing_status_fields = [
        label
        for key, label in (
            ("user_type", "사용자 유형"),
            ("health_insurance_status", "건강보험 상태"),
            ("vat_status", "과세 상태"),
        )
        if business_row.get(key) == "미확인"
    ]
    if missing_status_fields:
        _append_review_item(
            rows,
            item_type="사용자상태확인",
            related_kind="내 상태 설정",
            related_no="",
            summary="사용자 상태 설정 추가 확인 필요",
            status=" / ".join(missing_status_fields),
            needed="내 상태 설정 또는 추가 안내를 통해 미확인 상태를 보완해 주세요",
            priority="보통",
            note=business_row.get("note", ""),
        )

    withholding_row = withholding_summary_rows[0] if withholding_summary_rows else {}
    other_income_flag = str(withholding_row.get("other_income_flag", ""))
    has_withholding_data = str(withholding_row.get("has_withholding_data", ""))
    has_paid_tax_data = str(withholding_row.get("has_paid_tax_data", ""))
    withholding_period_basis = str(withholding_row.get("withholding_period_basis", "") or "")
    paid_tax_period_basis = str(withholding_row.get("paid_tax_period_basis", "") or "")
    payer_reference = str(withholding_row.get("payer_reference", "") or "")
    document_scope = str(withholding_row.get("document_scope", "") or "")
    gross_pay_total = withholding_row.get("gross_pay_total_krw")
    withholding_tax_total = withholding_row.get("withholding_tax_total_krw")
    paid_tax_total = withholding_row.get("paid_tax_total_krw")
    if other_income_flag.startswith("예") and withholding_row.get("has_withholding_data") != "예":
        _append_review_item(
            rows,
            item_type="원천징수자료누락",
            related_kind="공식자료",
            related_no="",
            summary="원천징수 자료 추가 확인 필요",
            status="자료 미첨부",
            needed="원천징수 관련 공식자료가 있다면 추가 업로드가 필요합니다",
            priority="높음",
            note=str(withholding_row.get("source_basis", "")),
        )
    if other_income_flag.startswith("예") and withholding_row.get("has_paid_tax_data") != "예":
        _append_review_item(
            rows,
            item_type="기납부세액자료누락",
            related_kind="공식자료",
            related_no="",
            summary="기납부세액 자료 추가 확인 필요",
            status="자료 미첨부",
            needed="홈택스 납부내역 등 기납부세액 자료가 있으면 함께 확인해 주세요",
            priority="보통",
            note=str(withholding_row.get("source_basis", "")),
        )
    if (
        (has_withholding_data == "예" and (withholding_period_basis == "미확인" or " 외 " in withholding_period_basis))
        or (has_paid_tax_data == "예" and (paid_tax_period_basis == "미확인" or " 외 " in paid_tax_period_basis))
    ):
        period_parts = []
        if has_withholding_data == "예":
            period_parts.append(f"원천징수: {withholding_period_basis or '미확인'}")
        if has_paid_tax_data == "예":
            period_parts.append(f"기납부세액: {paid_tax_period_basis or '미확인'}")
        _append_review_item(
            rows,
            item_type="원천징수기준기간확인",
            related_kind="공식자료",
            related_no="",
            summary="원천징수·기납부세액 기준 기간 재확인 필요",
            status=" / ".join(period_parts) or "미확인",
            needed="귀속기간 또는 문서 기준일이 대상 월과 맞는지 다시 확인해 주세요",
            priority="높음",
            note=" / ".join(part for part in [document_scope, str(withholding_row.get("source_basis", ""))] if part),
        )
    if has_withholding_data == "예" and payer_reference in {"미확인", "복수 지급처 참조"}:
        _append_review_item(
            rows,
            item_type="원천징수지급처참조미확인",
            related_kind="공식자료",
            related_no="",
            summary="원천징수 자료 지급처 참조 확인 필요",
            status=payer_reference,
            needed="지급처 참조 정보가 있으면 함께 확인해 주세요",
            priority="높음",
            note=" / ".join(part for part in [document_scope, str(withholding_row.get("source_basis", ""))] if part),
        )
    if has_withholding_data == "예" and gross_pay_total in ("", None):
        _append_review_item(
            rows,
            item_type="총지급액추출불가",
            related_kind="공식자료",
            related_no="",
            summary="원천징수 자료 총지급액 재확인 필요",
            status="총지급액 추출 불가",
            needed="총지급액을 구조화하지 못해 문서 기준으로 다시 확인해 주세요",
            priority="보통",
            note=" / ".join(part for part in [document_scope, str(withholding_row.get("note", ""))] if part),
        )
    if has_withholding_data == "예" and withholding_tax_total in ("", None):
        _append_review_item(
            rows,
            item_type="원천징수세액추출불가",
            related_kind="공식자료",
            related_no="",
            summary="원천징수세액 재확인 필요",
            status="원천징수세액 추출 불가",
            needed="원천징수세액 합계를 구조화하지 못해 문서 기준으로 다시 확인해 주세요",
            priority="보통",
            note=" / ".join(part for part in [document_scope, str(withholding_row.get("note", ""))] if part),
        )
    if has_paid_tax_data == "예" and paid_tax_total in ("", None):
        _append_review_item(
            rows,
            item_type="기납부세액추출불가",
            related_kind="공식자료",
            related_no="",
            summary="기납부세액 재확인 필요",
            status="기납부세액 추출 불가",
            needed="기납부세액 합계를 구조화하지 못해 납부내역 기준으로 다시 확인해 주세요",
            priority="보통",
            note=str(withholding_row.get("note", "")),
        )

    vat_row = vat_summary_rows[0] if vat_summary_rows else {}
    vat_status = str(vat_row.get("vat_status", ""))
    if vat_status.startswith("과세") and str(vat_row.get("recent_vat_filing_status", "")) in {"자료 없음", "미확인"}:
        _append_review_item(
            rows,
            item_type="부가세자료누락",
            related_kind="공식자료",
            related_no="",
            summary="부가세 신고 자료 추가 확인 필요",
            status=str(vat_row.get("recent_vat_filing_status", "자료 없음")),
            needed="최근 부가세 신고 여부와 관련 공식자료를 다시 확인해 주세요",
            priority="높음",
            note=str(vat_row.get("source_basis", "")),
        )
    elif vat_status != "미확인" and vat_row.get("needs_review") == "예":
        _append_review_item(
            rows,
            item_type="부가세재확인",
            related_kind="공식자료",
            related_no="",
            summary="부가세 요약값 재확인 필요",
            status="재확인 필요",
            needed="세금계산서/카드/현금영수증 관련 요약값이 충분한지 다시 확인해 주세요",
            priority="높음",
            note=str(vat_row.get("note", "")),
        )

    nhis_row = nhis_pension_summary_rows[0] if nhis_pension_summary_rows else {}
    if nhis_row.get("health_insurance_status") == "미확인":
        _append_review_item(
            rows,
            item_type="건보연금상태확인",
            related_kind="내 상태 설정",
            related_no="",
            summary="건강보험/연금 상태 추가 확인 필요",
            status="건강보험 상태 미확인",
            needed="내 상태 설정 또는 공식자료 업로드로 건보/연금 상태를 보완해 주세요",
            priority="보통",
            note=str(nhis_row.get("note", "")),
        )
    elif nhis_row.get("has_nhis_data") != "예":
        _append_review_item(
            rows,
            item_type="건보자료누락",
            related_kind="공식자료",
            related_no="",
            summary="건강보험 자료 추가 확인 필요",
            status="자료 미첨부",
            needed="건강보험 납부 자료 또는 안전한 요약 자료가 있으면 추가 확인해 주세요",
            priority="보통",
            note=str(nhis_row.get("source_basis", "")),
        )
    if nhis_row.get("pension_check_expected") == "예" and nhis_row.get("has_pension_data") != "예":
        _append_review_item(
            rows,
            item_type="연금자료누락",
            related_kind="공식자료",
            related_no="",
            summary="국민연금 자료 추가 확인 필요",
            status="자료 미첨부",
            needed="국민연금 납부 자료 여부를 다시 확인해 주세요",
            priority="보통",
            note=str(nhis_row.get("source_basis", "")),
        )

    for row in reference_material_rows:
        if row.get("needs_review") != "예":
            continue
        if row.get("link_status_key") == "official_difference":
            summary = "참고자료와 공식자료 금액 차이 확인 필요"
            needed = "참고자료의 설명 금액과 공식자료 요약값 차이를 확인해 주세요"
            note = f"{row.get('comparison_target', '')} / {row.get('difference_description', '')}".strip(" /")
        elif row.get("link_status_key") == "transaction_difference":
            summary = "참고자료와 거래 합계 차이 확인 필요"
            needed = "참고자료 기재 금액과 대상 월 거래 합계 차이를 확인해 주세요"
            note = f"{row.get('comparison_target', '')} / {row.get('difference_description', '')}".strip(" /")
        elif row.get("link_status_key") == "no_comparison":
            summary = "참고자료 비교 기준 확인 필요"
            needed = "기준 기간과 비교 대상이 모호해 참고자료를 보조 설명으로만 전달했습니다. 비교 가능 기준을 확인해 주세요"
            note = f"{row.get('reference_type', '')} / {row.get('difference_description', '')}".strip(" /")
        else:
            summary = "참고자료 보조 설명 연결 확인 필요"
            needed = "공식자료 또는 거래 합계와 연결되는 설명인지, 참고용 메모인지 확인해 주세요"
            note = f"{row.get('reference_type', '')} / {row.get('difference_description', '')}".strip(" /")
        _append_review_item(
            rows,
            item_type="참고자료검토",
            related_kind="참고자료",
            related_no=row.get("reference_material_id", ""),
            summary=summary,
            status=str(row.get("link_status", "reference_only")),
            needed=needed,
            priority="보통",
            note=note or str(row.get("title", "")),
        )

    return _finalize_review_items(rows)


def _source_labels(source: str | None, provider: str | None = None) -> tuple[str, str]:
    return (
        get_transaction_source_label(source, provider),
        get_transaction_provider_label(source, provider),
    )


def _classification_labels(tx: dict[str, Any]) -> tuple[str, str, str, bool, str, str]:
    direction = tx.get("direction")
    reasons: list[str] = []

    if direction == "in":
        income_status = tx.get("income_label_status") or "unknown"
        if income_status == "income":
            classification = "수입"
            business = "해당없음"
            calculation = "예"
        elif income_status == "non_income":
            classification = "수입 아님"
            business = "해당없음"
            calculation = "아니오"
        else:
            classification = "미확정"
            business = "해당없음"
            calculation = "보류"
            reasons.append("수입 분류가 아직 확정되지 않았습니다")
    else:
        expense_status = tx.get("expense_label_status") or "unknown"
        evidence_requirement = tx.get("evidence_requirement") or ""
        evidence_status = tx.get("evidence_status") or ""

        if expense_status == "business":
            classification = "업무지출"
            business = "예"
            calculation = "예"
        elif expense_status == "personal":
            classification = "개인지출"
            business = "아니오"
            calculation = "아니오"
        elif expense_status == "mixed":
            classification = "혼합지출"
            business = "혼합"
            calculation = "보류"
            reasons.append("업무/개인 지출 구분이 혼합 상태입니다")
        else:
            classification = "미확정"
            business = "미확정"
            calculation = "보류"
            reasons.append("지출 분류가 아직 확정되지 않았습니다")

        if evidence_status == "missing" and evidence_requirement == "required":
            calculation = "보류"
            reasons.append("필수 증빙이 아직 첨부되지 않았습니다")
        elif evidence_status == "missing" and evidence_requirement == "maybe":
            calculation = "보류"
            reasons.append("증빙 확인이 필요한 거래입니다")

    recheck_required = bool(reasons)
    recheck_reason = " / ".join(reasons)

    if recheck_required:
        trust = "재확인필요"
    elif tx.get("source") == "manual":
        trust = "참고용"
    else:
        trust = "반영됨"

    return classification, business, calculation, recheck_required, recheck_reason, trust


def _evidence_status_label(requirement: str | None, status: str | None) -> str:
    requirement = (requirement or "").strip()
    status = (status or "").strip()
    if status == "attached":
        return "첨부됨"
    if requirement == "not_needed" or status == "not_needed":
        return "불필요"
    if status == "missing" and requirement == "required":
        return "필수 누락"
    if status == "missing" and requirement == "maybe":
        return "확인 필요"
    return "상태 확인 필요"


def _evidence_type_label(mime_type: str | None, filename: str | None) -> str:
    mime = (mime_type or "").strip().lower()
    ext = Path(filename or "").suffix.lower()
    if mime.startswith("image/") or ext in {".jpg", ".jpeg", ".png", ".webp", ".gif", ".heic", ".heif"}:
        return "이미지 증빙"
    if mime == "application/pdf" or ext == ".pdf":
        return "PDF 증빙"
    return "증빙파일"


def _build_review_items(transactions: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    combined: list[dict[str, Any]] = []
    evidence_missing: list[dict[str, Any]] = []
    review_trades: list[dict[str, Any]] = []

    item_no = 1
    for tx in transactions:
        reasons: list[str] = []
        needed: list[str] = []
        current: list[str] = []
        related_material = "거래내역"
        item_type = "거래검토"
        priority = "보통"

        if tx.get("direction") == "out":
            if tx.get("evidence_status") == "missing" and tx.get("evidence_requirement") == "required":
                reasons.append("필수 증빙이 누락되었습니다")
                needed.append("대표 증빙을 첨부하거나 불필요 여부를 다시 판단해 주세요")
                current.append("필수 누락")
                related_material = "증빙자료"
                item_type = "증빙누락"
                priority = "높음"
                evidence_missing.append(
                    {
                        "거래번호": tx["tx_id"],
                        "거래일시": tx["occurred_at_kst"],
                        "거래처": tx["counterparty"],
                        "금액": tx["amount_krw"],
                        "증빙상태": "필수 누락",
                        "필요한확인내용": needed[-1],
                        "우선순위": priority,
                    }
                )
            elif tx.get("evidence_status") == "missing" and tx.get("evidence_requirement") == "maybe":
                reasons.append("증빙 확인이 필요한 거래입니다")
                needed.append("업무 관련이면 증빙을 첨부하고, 아니면 불필요로 표시해 주세요")
                current.append("확인 필요")
                related_material = "증빙자료"
                item_type = "증빙검토"
                priority = "보통"
                evidence_missing.append(
                    {
                        "거래번호": tx["tx_id"],
                        "거래일시": tx["occurred_at_kst"],
                        "거래처": tx["counterparty"],
                        "금액": tx["amount_krw"],
                        "증빙상태": "확인 필요",
                        "필요한확인내용": needed[-1],
                        "우선순위": priority,
                    }
                )

            if tx.get("expense_label_status") in {"unknown", "mixed", ""}:
                reasons.append("지출 분류가 확정되지 않았습니다")
                needed.append("업무/개인/혼합 중 하나로 확정해 주세요")
                current.append("분류 미확정")
                review_trades.append(
                    {
                        "거래번호": tx["tx_id"],
                        "거래일시": tx["occurred_at_kst"],
                        "자료출처": tx["source_label"],
                        "거래처": tx["counterparty"],
                        "금액": tx["amount_krw"],
                        "현재상태": tx["classification_result_label"],
                        "재확인사유": "업무/개인 판단이 확정되지 않았습니다",
                        "필요한확인내용": "업무/개인/혼합 중 하나로 확정해 주세요",
                    }
                )
        else:
            if tx.get("income_label_status") in {"unknown", ""}:
                reasons.append("수입 분류가 확정되지 않았습니다")
                needed.append("수입인지 수입 아님인지 확인해 주세요")
                current.append("분류 미확정")
                review_trades.append(
                    {
                        "거래번호": tx["tx_id"],
                        "거래일시": tx["occurred_at_kst"],
                        "자료출처": tx["source_label"],
                        "거래처": tx["counterparty"],
                        "금액": tx["amount_krw"],
                        "현재상태": tx["classification_result_label"],
                        "재확인사유": "수입/비수입 판단이 확정되지 않았습니다",
                        "필요한확인내용": "수입인지 수입 아님인지 확인해 주세요",
                    }
                )

        if not reasons:
            continue

        combined.append(
            {
                "항목번호": item_no,
                "항목유형": item_type,
                "관련자료구분": related_material,
                "관련번호": tx["tx_id"],
                "요약설명": " / ".join(reasons),
                "현재상태": " / ".join(current) or tx["trust_label"],
                "필요한확인내용": " / ".join(needed),
                "우선순위": priority,
                "메모": tx.get("memo") or "",
            }
        )
        item_no += 1

    return combined, evidence_missing, review_trades


def _collect_package_snapshot(user_pk: int, month_key: str) -> PackageSnapshot:
    month_key = (month_key or "").strip()
    if not month_key or len(month_key) != 7 or month_key[4] != "-":
        month_key = utcnow().strftime("%Y-%m")

    start_dt, end_dt = _month_range_kst_naive(month_key)
    period_start_kst = start_dt.strftime("%Y-%m-%d")
    last_day = (end_dt - datetime.resolution).date()
    period_end_kst = last_day.strftime("%Y-%m-%d")

    user = User.query.get(user_pk)
    display_name = getattr(user, "nickname", None) or f"user{user_pk}"
    package_label = _safe_package_label(display_name, f"user{user_pk}")
    root_name = f"세무사전달패키지_{month_key}_{package_label}"
    download_name = f"{root_name}.zip"

    rows = (
        db.session.query(Transaction, IncomeLabel, ExpenseLabel, EvidenceItem)
        .outerjoin(IncomeLabel, and_(IncomeLabel.transaction_id == Transaction.id, IncomeLabel.user_pk == user_pk))
        .outerjoin(ExpenseLabel, and_(ExpenseLabel.transaction_id == Transaction.id, ExpenseLabel.user_pk == user_pk))
        .outerjoin(EvidenceItem, and_(EvidenceItem.transaction_id == Transaction.id, EvidenceItem.user_pk == user_pk))
        .filter(Transaction.user_pk == user_pk)
        .filter(Transaction.occurred_at >= start_dt, Transaction.occurred_at < end_dt)
        .order_by(Transaction.occurred_at.asc(), Transaction.id.asc())
        .all()
    )

    import_job_ids = sorted({int(tx.import_job_id) for tx, _il, _el, _ev in rows if tx.import_job_id})
    import_job_map: dict[int, ImportJob] = {}
    if import_job_ids:
        for job in ImportJob.query.filter(ImportJob.id.in_(import_job_ids)).all():
            import_job_map[int(job.id)] = job

    tx_rows: list[dict[str, Any]] = []
    evidence_rows: list[dict[str, Any]] = []
    tx_in_count = 0
    tx_out_count = 0
    sum_in_total = 0
    sum_out_total = 0
    income_included_total = 0
    income_excluded_non_income_total = 0
    income_unknown_count = 0
    expense_business_total = 0
    expense_personal_total = 0
    expense_mixed_total = 0
    expense_unknown_total = 0
    evidence_missing_required_count = 0
    evidence_missing_required_amount = 0
    evidence_missing_maybe_count = 0
    evidence_missing_maybe_amount = 0
    evidence_attached_count = 0
    source_labels: set[str] = set()

    for tx, income_label, expense_label, evidence in rows:
        amount = _safe_int(tx.amount_krw)
        resolved_source, resolved_provider = resolve_transaction_origin(tx.source, tx.provider)
        source_label, provider_label = _source_labels(tx.source, tx.provider)
        source_labels.add(source_label)
        import_job = import_job_map.get(int(tx.import_job_id)) if tx.import_job_id else None

        evidence_requirement = (evidence.requirement if evidence else "") or ""
        evidence_status = (evidence.status if evidence else "") or ""
        evidence_filename = (evidence.original_filename if evidence else "") or ""
        evidence_mime = (evidence.mime_type if evidence else "") or ""
        evidence_type = _evidence_type_label(evidence_mime, evidence_filename)
        evidence_zip_path = ""
        evidence_abs_path: Path | None = None
        evidence_count = 0

        if evidence and evidence.file_key and evidence.deleted_at is None:
            try:
                evidence_abs_path = resolve_file_path(evidence.file_key)
                if evidence_abs_path.exists() and evidence_abs_path.is_file():
                    safe_name = _safe_attachment_name(evidence_filename, evidence_abs_path.name)
                    evidence_zip_path = f"{EVIDENCE_ATTACHMENTS_DIR}/{tx.id}_{safe_name}"
                    evidence_count = 1
                    evidence_attached_count += 1
            except Exception:
                evidence_abs_path = None
                evidence_zip_path = ""
                evidence_count = 0

        tx_row = {
            "tx_id": int(tx.id),
            "occurred_at_kst": _fmt_kst(tx.occurred_at, "%Y-%m-%d %H:%M"),
            "date_kst": _fmt_kst(tx.occurred_at, "%Y-%m-%d"),
            "direction": tx.direction,
            "direction_label": "입금" if tx.direction == "in" else "출금",
            "amount_krw": amount,
            "counterparty": tx.counterparty or "",
            "memo": tx.memo or "",
            "source": resolved_source or "",
            "provider": resolved_provider or "",
            "source_label": source_label,
            "provider_label": provider_label,
            "external_hash": tx.external_hash or "",
            "import_job_id": int(tx.import_job_id) if tx.import_job_id else "",
            "import_filename": (import_job.filename if import_job else "") or "",
            "income_label_status": (income_label.status if income_label else "") or "",
            "income_label_confidence": _safe_int(income_label.confidence) if income_label else 0,
            "income_labeled_by": (income_label.labeled_by if income_label else "") or "",
            "expense_label_status": (expense_label.status if expense_label else "") or "",
            "expense_label_confidence": _safe_int(expense_label.confidence) if expense_label else 0,
            "expense_labeled_by": (expense_label.labeled_by if expense_label else "") or "",
            "evidence_id": int(evidence.id) if evidence else "",
            "evidence_requirement": evidence_requirement,
            "evidence_status": evidence_status,
            "evidence_status_label": _evidence_status_label(evidence_requirement, evidence_status),
            "evidence_note": (evidence.note if evidence else "") or "",
            "evidence_original_filename": evidence_filename,
            "evidence_mime_type": evidence_mime,
            "evidence_size_bytes": _safe_int(evidence.size_bytes) if evidence and evidence.size_bytes is not None else 0,
            "evidence_sha256": (evidence.sha256 if evidence else "") or "",
            "evidence_uploaded_at_kst": _fmt_kst(evidence.uploaded_at if evidence else None, "%Y-%m-%d %H:%M"),
            "evidence_deleted_at_kst": _fmt_kst(evidence.deleted_at if evidence else None, "%Y-%m-%d %H:%M"),
            "evidence_retention_until": evidence.retention_until.isoformat() if evidence and evidence.retention_until else "",
            "representative_evidence_type": evidence_type,
            "evidence_count": evidence_count,
            "evidence_zip_path": evidence_zip_path,
            "evidence_abs_path": evidence_abs_path,
        }

        (
            tx_row["classification_result_label"],
            tx_row["business_related_label"],
            tx_row["calculation_included_label"],
            tx_row["recheck_required"],
            tx_row["recheck_reason"],
            tx_row["trust_label"],
        ) = _classification_labels(tx_row)

        tx_row["recheck_required_label"] = "예" if tx_row["recheck_required"] else "아니오"

        if tx.direction == "in":
            tx_in_count += 1
            sum_in_total += amount
            if tx_row["income_label_status"] == "non_income":
                income_excluded_non_income_total += amount
            else:
                income_included_total += amount
                if tx_row["income_label_status"] in {"", "unknown"}:
                    income_unknown_count += 1
        else:
            tx_out_count += 1
            sum_out_total += amount
            if tx_row["expense_label_status"] == "business":
                expense_business_total += amount
            elif tx_row["expense_label_status"] == "personal":
                expense_personal_total += amount
            elif tx_row["expense_label_status"] == "mixed":
                expense_mixed_total += amount
            else:
                expense_unknown_total += amount

            if evidence_status == "missing" and evidence_requirement == "required":
                evidence_missing_required_count += 1
                evidence_missing_required_amount += amount
            elif evidence_status == "missing" and evidence_requirement == "maybe":
                evidence_missing_maybe_count += 1
                evidence_missing_maybe_amount += amount

        tx_rows.append(tx_row)

        if evidence_count == 1:
            evidence_rows.append(
                {
                    "증빙번호": tx_row["evidence_id"],
                    "연결거래번호": tx_row["tx_id"],
                    "거래일시": tx_row["occurred_at_kst"],
                    "거래처": tx_row["counterparty"],
                    "금액": tx_row["amount_krw"],
                    "증빙종류": evidence_type,
                    "파일명": evidence_filename,
                    "파일열기": ("열기", evidence_zip_path),
                    "저장위치": evidence_zip_path,
                    "업로드일시": tx_row["evidence_uploaded_at_kst"],
                    "신뢰구분": tx_row["trust_label"],
                    "계산반영여부": tx_row["calculation_included_label"],
                    "재확인필요여부": tx_row["recheck_required_label"],
                    "메모": tx_row["evidence_note"],
                    "_zip_path": evidence_zip_path,
                    "_abs_path": evidence_abs_path,
                }
            )

    review_items, evidence_missing_items, review_trade_items = _build_review_items(tx_rows)

    official_rows: list[dict[str, Any]] = []
    official_start_date, official_end_date = _month_date_range(month_key)
    official_docs = (
        OfficialDataDocument.query.filter(OfficialDataDocument.user_pk == user_pk)
        .filter(
            or_(
                and_(
                    OfficialDataDocument.reference_date.isnot(None),
                    OfficialDataDocument.reference_date >= official_start_date,
                    OfficialDataDocument.reference_date < official_end_date,
                ),
                and_(
                    OfficialDataDocument.reference_date.is_(None),
                    OfficialDataDocument.created_at >= start_dt,
                    OfficialDataDocument.created_at < end_dt,
                ),
            )
        )
        .order_by(OfficialDataDocument.created_at.asc(), OfficialDataDocument.id.asc())
        .all()
    )

    official_parsed_count = 0
    official_review_count = 0
    official_unsupported_count = 0
    official_failed_count = 0
    cross_validation_context = build_cross_validation_context(user_pk=user_pk)

    for document in official_docs:
        view = official_data_document_to_view_model(document)
        cross_validation = build_official_document_cross_validation(
            document=document,
            context=cross_validation_context,
        )
        cross_validation_status_key = str(cross_validation.get("status") or "").strip()
        cross_validation_status_label = _package_cross_validation_status_label(cross_validation_status_key)
        cross_validation_reason = _package_cross_validation_reason(
            cross_validation_status_key,
            str(cross_validation.get("reason") or ""),
        )

        if document.parse_status == "parsed":
            official_parsed_count += 1
        elif document.parse_status == "needs_review":
            official_review_count += 1
        elif document.parse_status == "unsupported":
            official_unsupported_count += 1
        else:
            official_failed_count += 1

        official_rows.append(
            {
                "자료번호": int(document.id),
                "기관명": view.get("source_authority", "확인 전"),
                "문서종류": view.get("document_type_label", "문서 판별 전"),
                "기준일": view.get("reference_date", "확인 전"),
                "원본파일명": document.original_filename,
                "원본첨부여부": "아니오",
                "읽기상태": view.get("parse_status_label", "처리 결과 확인"),
                "검증상태": view.get("verification_status_label", "검증 미실시"),
                "구조확인": view.get("structure_validation_label", "구조 미확인"),
                "신뢰등급": _official_trust_label(view, document),
                "핵심값요약": _official_summary_text(view),
                "교차검증상태": cross_validation_status_label,
                "교차검증사유": cross_validation_reason,
                "교차검증재확인필요여부": _package_cross_validation_recheck_label(cross_validation_status_key),
                "목록반영여부": "예",
                "재확인필요여부": _official_recheck_label(document),
                "메모": _official_note(view, document),
                "_summary_items": view.get("summary_items") or [],
                "_summary_values": dict(document.extracted_key_summary_json or {}),
                "_document_type_label": view.get("document_type_label", "문서 판별 전"),
                "_parse_status": document.parse_status,
                "_cross_validation_status_key": cross_validation_status_key,
                "_attachment_index_key": f"official-{int(document.id)}",
                "_period_basis": view.get("reference_date", "확인 전"),
            }
        )

    business_status_rows = _collect_business_status_rows(user_pk)
    withholding_summary_rows = _collect_withholding_summary_rows(user_pk, official_rows)
    vat_summary_rows = _collect_vat_summary_rows(user_pk, official_rows, tx_rows)
    nhis_pension_summary_rows = _collect_nhis_pension_summary_rows(user_pk, month_key, official_rows)
    reference_material_rows = _collect_reference_material_rows(
        user_pk=user_pk,
        start_dt=start_dt,
        end_dt=end_dt,
        official_documents=official_rows,
        transactions=tx_rows,
    )
    review_items = _extend_review_items(
        review_items=review_items,
        official_documents=official_rows,
        business_status_rows=business_status_rows,
        withholding_summary_rows=withholding_summary_rows,
        vat_summary_rows=vat_summary_rows,
        nhis_pension_summary_rows=nhis_pension_summary_rows,
        reference_material_rows=reference_material_rows,
    )

    settings = _get_settings(user_pk)
    rate = _tax_rate(settings)
    tax_buffer_total = (
        db.session.query(func.coalesce(func.sum(TaxBufferLedger.delta_amount_krw), 0))
        .filter(TaxBufferLedger.user_pk == user_pk)
        .scalar()
    ) or 0
    tax_buffer_target = int(int(income_included_total) * float(rate))
    tax_buffer_shortage = max(0, int(tax_buffer_target) - int(tax_buffer_total))

    stats = PackageStats(
        month_key=month_key,
        period_start_kst=period_start_kst,
        period_end_kst=period_end_kst,
        generated_at_kst=_fmt_kst(utcnow(), "%Y-%m-%d %H:%M") or utcnow().strftime("%Y-%m-%d %H:%M"),
        tx_total=len(tx_rows),
        tx_in_count=int(tx_in_count),
        tx_out_count=int(tx_out_count),
        sum_in_total=int(sum_in_total),
        sum_out_total=int(sum_out_total),
        income_included_total=int(income_included_total),
        income_excluded_non_income_total=int(income_excluded_non_income_total),
        income_unknown_count=int(income_unknown_count),
        expense_business_total=int(expense_business_total),
        expense_personal_total=int(expense_personal_total),
        expense_mixed_total=int(expense_mixed_total),
        expense_unknown_total=int(expense_unknown_total),
        evidence_missing_required_count=int(evidence_missing_required_count),
        evidence_missing_required_amount=int(evidence_missing_required_amount),
        evidence_missing_maybe_count=int(evidence_missing_maybe_count),
        evidence_missing_maybe_amount=int(evidence_missing_maybe_amount),
        evidence_attached_count=int(evidence_attached_count),
        review_needed_count=len(review_items),
        tax_rate=float(rate),
        tax_buffer_total=int(tax_buffer_total),
        tax_buffer_target=int(tax_buffer_target),
        tax_buffer_shortage=int(tax_buffer_shortage),
        official_data_total=len(official_rows),
        official_data_parsed_count=int(official_parsed_count),
        official_data_review_count=int(official_review_count),
        official_data_unsupported_count=int(official_unsupported_count),
        official_data_failed_count=int(official_failed_count),
    )

    return PackageSnapshot(
        root_name=root_name,
        download_name=download_name,
        display_name=display_name,
        stats=stats,
        transactions=tx_rows,
        evidences=evidence_rows,
        review_items=review_items,
        evidence_missing_items=evidence_missing_items,
        review_trade_items=review_trade_items,
        included_source_labels=sorted(source_labels),
        official_documents=official_rows,
        business_status_rows=business_status_rows,
        withholding_summary_rows=withholding_summary_rows,
        vat_summary_rows=vat_summary_rows,
        nhis_pension_summary_rows=nhis_pension_summary_rows,
        reference_material_rows=reference_material_rows,
    )


def _write_table_sheet(ws, headers: list[str], rows: list[dict[str, Any]], freeze: str = "A2") -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = TOP_ALIGN

    for row in rows:
        values = []
        links: list[tuple[int, str]] = []
        for idx, header in enumerate(headers, start=1):
            value = row.get(header, "")
            if isinstance(value, tuple) and len(value) == 2:
                display, target = value
                values.append(display)
                if target:
                    links.append((idx, target))
            else:
                values.append(value)
        ws.append(values)
        current_row = ws.max_row
        for col_idx, target in links:
            cell = ws.cell(current_row, col_idx)
            cell.hyperlink = target
            cell.style = "Hyperlink"
        for cell in ws[current_row]:
            cell.alignment = TOP_ALIGN

    ws.freeze_panes = freeze
    if rows:
        ws.auto_filter.ref = ws.dimensions
    _style_table_sheet(ws, headers, rows)
    _autosize(ws)


def _style_table_sheet(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    amount_keywords = ("금액", "합계", "총", "세액", "부족", "목표", "amount", "_krw", "total")
    wrap_keywords = ("메모", "적요", "사유", "설명", "요약", "내용", "파일명", "값", "기준", "대상", "출처", "수준", "note", "summary", "basis", "title", "path", "link")
    status_keywords = ("상태", "구분", "여부", "우선순위", "신뢰", "반영", "재확인", "민감정보", "status", "flag", "review", "contains_sensitive_info")

    for idx, header in enumerate(headers, start=1):
        column = get_column_letter(idx)
        if any(keyword in header for keyword in amount_keywords):
            for cell in ws[column][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
        if any(keyword in header for keyword in wrap_keywords):
            for cell in ws[column]:
                cell.alignment = TOP_ALIGN
        if any(keyword in header for keyword in status_keywords):
            for cell in ws[column][1:]:
                value = str(cell.value or "")
                if value in {"반영됨", "첨부됨", "예", "반영 가능", "포함", "확인됨", "일치", "공식자료 요약과 대체로 일치", "거래 합계와 대체로 일치", "자료 있음"}:
                    cell.fill = GOOD_FILL
                elif value in {
                    "재확인필요",
                    "재확인 필요",
                    "검토 필요",
                    "보류",
                    "확인 필요",
                    "기본 제외",
                    "자료 없음",
                    "미확인",
                    "참고용",
                    "부분 일치",
                    "비교 불가",
                    "비교 기준 없음",
                    "공식자료 요약과 차이 있음",
                    "거래 합계와 차이 있음",
                    "중간",
                    "낮음",
                }:
                    cell.fill = WARN_FILL
                elif value in {"필수 누락", "아니오", "미지원 형식", "읽기 실패", "불일치", "높음"}:
                    cell.fill = BAD_FILL

    ws.sheet_view.showGridLines = True


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        values = ["" if c.value is None else str(c.value) for c in column_cells]
        width = min(max((len(v) for v in values), default=10) + 2, 52)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(10, width)


def _workbook_bytes(builder) -> bytes:
    wb = builder()
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_summary_workbook(snapshot: PackageSnapshot, profile: TaxPackageProfile | None = None) -> bytes:
    profile = profile or get_tax_package_profile(DEFAULT_TAX_PACKAGE_PROFILE_CODE)
    stats = snapshot.stats
    reference_count = len([row for row in snapshot.reference_material_rows if row.get("reference_material_id")])
    reference_review_count = len([row for row in snapshot.reference_material_rows if row.get("needs_review") == "예"])
    cross_validation_counts = _official_cross_validation_counts(snapshot.official_documents)
    withholding_row = (snapshot.withholding_summary_rows or [{}])[0]
    vat_row = (snapshot.vat_summary_rows or [{}])[0]
    nhis_row = (snapshot.nhis_pension_summary_rows or [{}])[0]
    vat_review_count = _count_review_items_by_type(snapshot.review_items, {"부가세자료누락", "부가세재확인"})
    summary_row_map = {
        "user_name": {"항목명": "사용자명", "값": snapshot.display_name},
        "period": {"항목명": "대상 기간", "값": f"{stats.period_start_kst} ~ {stats.period_end_kst}"},
        "generated_at": {"항목명": "생성일시", "값": stats.generated_at_kst},
        "review_start": {"항목명": "검토 시작", "값": f"{_workbook_filename('summary')} → {_workbook_filename('review')}"},
        "primary_review_flow": {"항목명": "1차 검토 흐름", "값": ""},
        "immediate_recheck_count": {"항목명": "즉시 재확인 항목 수", "값": stats.review_needed_count},
        "evidence_review": {"항목명": "증빙 확인 필요", "값": f"필수 {stats.evidence_missing_required_count}건 / 확인 {stats.evidence_missing_maybe_count}건"},
        "withholding_status": {"항목명": "원천징수·기납부세액 상태", "값": _withholding_overview_label(withholding_row)},
        "vat_status": {"항목명": "부가세 상태", "값": _vat_overview_label(vat_row)},
        "nhis_pension_status": {"항목명": "건보·연금 상태", "값": _nhis_pension_overview_label(nhis_row)},
        "reference_review_status": {"항목명": "참고자료 검토 상태", "값": f"재확인 필요 {reference_review_count}건 / 총 {reference_count}건"},
        "cross_validation_notice": {"항목명": "교차검증 안내", "값": "교차검증 v1 기준 / 비교 가능한 공식자료만 상세 비교하고 나머지는 비교 불가로 집계했습니다."},
        "cross_validation_overview": {"항목명": "공식자료 교차검증 상태", "값": _cross_validation_overview_label(cross_validation_counts)},
        "vat_recheck_count": {"항목명": "부가세 재확인 항목 수", "값": vat_review_count},
        "vat_material_status": {"항목명": "세금계산서/매입자료 요약", "값": _vat_material_overview_label(vat_row)},
        "cross_validation_match": {"항목명": "교차검증 일치 문서 수", "값": cross_validation_counts["일치"]},
        "cross_validation_partial": {"항목명": "교차검증 부분 일치 문서 수", "값": cross_validation_counts["부분 일치"]},
        "cross_validation_review_needed": {"항목명": "교차검증 재확인 필요 문서 수", "값": cross_validation_counts["재확인 필요"]},
        "cross_validation_mismatch": {"항목명": "교차검증 불일치 문서 수", "값": cross_validation_counts["불일치"]},
        "cross_validation_unavailable": {"항목명": "교차검증 비교 불가 문서 수", "값": cross_validation_counts["비교 불가"]},
        "tx_total": {"항목명": "총 거래 수", "값": stats.tx_total},
        "sum_in_total": {"항목명": "총 수입", "값": stats.sum_in_total},
        "sum_out_total": {"항목명": "총 지출", "값": stats.sum_out_total},
        "expense_business_total": {"항목명": "업무 관련 지출 합계", "값": stats.expense_business_total},
        "evidence_attached_count": {"항목명": "증빙 첨부 수", "값": stats.evidence_attached_count},
        "official_data_total": {"항목명": "공식자료 수", "값": stats.official_data_total},
        "reference_count": {"항목명": "참고자료 수", "값": reference_count},
        "official_data_parsed_count": {"항목명": "읽기 가능한 공식자료 수", "값": stats.official_data_parsed_count},
        "official_data_review_count": {"항목명": "검토 필요 공식자료 수", "값": stats.official_data_review_count},
        "reference_note": {"항목명": "참고", "값": "공식자료/참고자료 원본은 기본 패키지에 포함하지 않고 요약값 중심으로 전달합니다."},
    }
    primary_review_flow = [
        _workbook_filename(key)
        for key in profile.workbook_badge_order
        if key in {"business_status", "withholding", "vat", "nhis_pension", "transactions", "evidence", "attachments", "reference"}
        and key in profile.included_workbooks
    ]
    if primary_review_flow:
        summary_row_map["primary_review_flow"]["값"] = " → ".join(primary_review_flow)

    def build() -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "패키지요약"
        summary_rows = [summary_row_map[key] for key in profile.summary_row_order if key in summary_row_map]
        _write_table_sheet(ws, ["항목명", "값"], summary_rows)

        ws_guide = wb.create_sheet("패키지안내")
        guide_rows = [{"안내": line} for line in _render_package_guide(snapshot, profile).splitlines() if line.strip()]
        _write_table_sheet(ws_guide, ["안내"], guide_rows)

        ws2 = wb.create_sheet("신뢰구분")
        _write_table_sheet(
            ws2,
            ["구분", "의미", "계산 반영 여부", "세무사 확인 필요 여부", "예시"],
            [
                {
                    "구분": "반영됨",
                    "의미": "구조화된 거래 또는 읽기 가능한 공식자료처럼 현재 패키지에서 반영 후보로 볼 수 있는 항목",
                    "계산 반영 여부": "예 또는 아니오",
                    "세무사 확인 필요 여부": "낮음",
                    "예시": "자동연동/수동업로드 거래 + 분류 완료, 또는 반영 가능으로 읽힌 공식자료",
                },
                {
                    "구분": "참고용",
                    "의미": "현재 자료는 있으나 구조화 근거 또는 검증 정보가 상대적으로 약한 항목",
                    "계산 반영 여부": "예/아니오를 함께 표기",
                    "세무사 확인 필요 여부": "보통",
                    "예시": "수동입력 거래, 또는 검증 미실시 상태의 보수적 공식자료 메타",
                },
                {
                    "구분": "재확인필요",
                    "의미": "분류 미확정, 필수 증빙 누락, 공식자료 검토 필요/미지원/읽기 실패처럼 추가 확인이 필요한 항목",
                    "계산 반영 여부": "보류 중심",
                    "세무사 확인 필요 여부": "높음",
                    "예시": "지출 분류 미확정 거래, 필수 증빙 미첨부 거래, 검토 필요 공식자료",
                },
            ],
        )

        tx_status = _count_by_trust(snapshot.transactions)
        ev_status = _count_by_trust(snapshot.evidences)
        review_status = _count_by_trust(snapshot.review_items)
        ws3 = wb.create_sheet("반영현황")
        _write_table_sheet(
            ws3,
            ["자료 구분", "개수", "반영 건수", "참고용 건수", "재확인 건수", "미지원/실패 건수"],
            [
                {"자료 구분": "거래내역", "개수": len(snapshot.transactions), "반영 건수": tx_status["반영됨"], "참고용 건수": tx_status["참고용"], "재확인 건수": tx_status["재확인필요"], "미지원/실패 건수": 0},
                {"자료 구분": "증빙자료", "개수": len(snapshot.evidences), "반영 건수": ev_status["반영됨"], "참고용 건수": ev_status["참고용"], "재확인 건수": ev_status["재확인필요"], "미지원/실패 건수": 0},
                {"자료 구분": "공식자료", "개수": stats.official_data_total, "반영 건수": stats.official_data_parsed_count, "참고용 건수": 0, "재확인 건수": stats.official_data_review_count, "미지원/실패 건수": stats.official_data_unsupported_count + stats.official_data_failed_count},
                {"자료 구분": "참고자료", "개수": reference_count, "반영 건수": 0, "참고용 건수": reference_count, "재확인 건수": len([row for row in snapshot.reference_material_rows if row.get("needs_review") == "예"]), "미지원/실패 건수": 0},
                {"자료 구분": "확인필요항목", "개수": len(snapshot.review_items), "반영 건수": review_status["반영됨"], "참고용 건수": review_status["참고용"], "재확인 건수": review_status["재확인필요"], "미지원/실패 건수": 0},
            ],
        )

        ws4 = wb.create_sheet("기본정보")
        _write_table_sheet(
            ws4,
            ["항목명", "값"],
            [
                {"항목명": "패키지 버전", "값": PACKAGE_VERSION},
                {"항목명": "생성 기준 월", "값": stats.month_key},
                {"항목명": "포함 파일 수", "값": len(profile.included_workbooks) + len(snapshot.evidences)},
                {"항목명": "포함 증빙 수", "값": len(snapshot.evidences)},
                {"항목명": "목록 반영 공식자료 수", "값": stats.official_data_total},
                {"항목명": "요약 반영 참고자료 수", "값": reference_count},
                {"항목명": "연동 포함 여부", "값": ", ".join(snapshot.included_source_labels) if snapshot.included_source_labels else "없음"},
            ],
        )

        _append_official_data_sheets(wb, snapshot, profile)
        return wb

    return _workbook_bytes(build)


def _count_by_trust(rows: list[dict[str, Any]]) -> dict[str, int]:
    counts = {"반영됨": 0, "참고용": 0, "재확인필요": 0}
    for row in rows:
        label = row.get("trust_label") or row.get("신뢰구분") or "재확인필요"
        if label not in counts:
            label = "재확인필요"
        counts[label] += 1
    return counts


def _build_transactions_workbook(snapshot: PackageSnapshot) -> bytes:
    stats = snapshot.stats

    def build() -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "거래원장"
        rows = []
        for tx in snapshot.transactions:
            rows.append(
                {
                    "거래번호": tx.get("tx_id", ""),
                    "거래일시": tx.get("occurred_at_kst", ""),
                    "입출금구분": tx.get("direction_label", ""),
                    "금액": tx.get("amount_krw", 0),
                    "거래처": tx.get("counterparty", ""),
                    "적요": tx.get("memo", ""),
                    "자료출처": tx.get("source_label", ""),
                    "연동공급자": tx.get("provider_label", ""),
                    "분류결과": tx.get("classification_result_label", ""),
                    "업무관련여부": tx.get("business_related_label", ""),
                    "증빙상태": tx.get("evidence_status_label", ""),
                    "대표증빙종류": tx.get("representative_evidence_type", ""),
                    "증빙개수": tx.get("evidence_count", 0),
                    "대표첨부열기": ("열기", tx.get("evidence_zip_path")) if tx.get("evidence_zip_path") else "",
                    "신뢰구분": tx.get("trust_label", "재확인필요"),
                    "계산반영여부": tx.get("calculation_included_label", "보류"),
                    "재확인필요여부": tx.get("recheck_required_label", "아니오"),
                    "재확인사유": tx.get("recheck_reason", ""),
                    "증빙메모": tx.get("evidence_note", ""),
                }
            )
        _write_table_sheet(
            ws,
            [
                "거래번호",
                "거래일시",
                "입출금구분",
                "금액",
                "거래처",
                "적요",
                "자료출처",
                "연동공급자",
                "분류결과",
                "업무관련여부",
                "증빙상태",
                "대표증빙종류",
                "증빙개수",
                "대표첨부열기",
                "신뢰구분",
                "계산반영여부",
                "재확인필요여부",
                "재확인사유",
                "증빙메모",
            ],
            rows,
        )

        ws2 = wb.create_sheet("원본메타")
        raw_rows = []
        for tx in snapshot.transactions:
            raw_rows.append(
                {
                    "거래번호": tx.get("tx_id", ""),
                    "원본자료유형": tx.get("source_label", ""),
                    "원본파일명(있으면)": tx.get("import_filename", ""),
                    "원본행번호(있으면)": "",
                    "원본거래일시": tx.get("occurred_at_kst", ""),
                    "원본금액": tx.get("amount_krw", 0),
                    "원본거래처": tx.get("counterparty", ""),
                    "정규화메모": tx.get("memo", ""),
                }
            )
        _write_table_sheet(
            ws2,
            ["거래번호", "원본자료유형", "원본파일명(있으면)", "원본행번호(있으면)", "원본거래일시", "원본금액", "원본거래처", "정규화메모"],
            raw_rows,
        )

        ws3 = wb.create_sheet("월별요약")
        _write_table_sheet(
            ws3,
            ["항목", "값"],
            [
                {"항목": "대상 월", "값": stats.month_key},
                {"항목": "총 거래 수", "값": stats.tx_total},
                {"항목": "총 수입", "값": stats.sum_in_total},
                {"항목": "총 지출", "값": stats.sum_out_total},
                {"항목": "업무 관련 지출", "값": stats.expense_business_total},
                {"항목": "첨부된 증빙 수", "값": stats.evidence_attached_count},
                {"항목": "확인 필요 항목 수", "값": stats.review_needed_count},
            ],
        )

        ws4 = wb.create_sheet("분류요약")
        summary_map: dict[tuple[str, str], dict[str, int]] = {}
        for tx in snapshot.transactions:
            key = (tx.get("classification_result_label", ""), tx.get("trust_label", "재확인필요"))
            bucket = summary_map.setdefault(key, {"count": 0, "amount": 0})
            bucket["count"] += 1
            bucket["amount"] += int(tx.get("amount_krw", 0))
        rows4 = []
        for (classification, trust), bucket in sorted(summary_map.items()):
            rows4.append(
                {
                    "분류결과": classification,
                    "신뢰구분": trust,
                    "거래 수": bucket["count"],
                    "금액 합계": bucket["amount"],
                }
            )
        _write_table_sheet(ws4, ["분류결과", "신뢰구분", "거래 수", "금액 합계"], rows4)
        return wb

    return _workbook_bytes(build)


def _build_evidence_workbook(snapshot: PackageSnapshot) -> bytes:
    def build() -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "증빙상태표"
        evidence_rows = []
        for evidence in snapshot.evidences:
            evidence_rows.append(
                {
                    "증빙번호": evidence.get("증빙번호", ""),
                    "연결거래번호": evidence.get("연결거래번호", ""),
                    "증빙종류": evidence.get("증빙종류", "증빙파일"),
                    "파일명": evidence.get("파일명", ""),
                    "첨부열기": evidence.get("파일열기", ""),
                    "저장위치": evidence.get("저장위치", ""),
                    "업로드일시": evidence.get("업로드일시", ""),
                    "신뢰구분": evidence.get("신뢰구분", "재확인필요"),
                    "계산반영여부": evidence.get("계산반영여부", "보류"),
                    "재확인필요여부": evidence.get("재확인필요여부", "아니오"),
                    "메모": evidence.get("메모", ""),
                }
            )
        evidence_rows.sort(
            key=lambda row: (
                0 if row.get("재확인필요여부") == "예" else 1,
                0 if row.get("신뢰구분") == "재확인필요" else 1,
                _review_related_no_sort_key(row.get("연결거래번호")),
            )
        )
        _write_table_sheet(
            ws,
            ["증빙번호", "연결거래번호", "증빙종류", "파일명", "첨부열기", "저장위치", "업로드일시", "신뢰구분", "계산반영여부", "재확인필요여부", "메모"],
            evidence_rows,
        )

        ws2 = wb.create_sheet("거래별대표첨부")
        linked_rows = []
        for tx in snapshot.transactions:
            linked_rows.append(
                {
                    "거래번호": tx.get("tx_id", ""),
                    "거래일시": tx.get("occurred_at_kst", ""),
                    "거래처": tx.get("counterparty", ""),
                    "금액": tx.get("amount_krw", 0),
                    "증빙상태": tx.get("evidence_status_label", ""),
                    "대표증빙종류": tx.get("representative_evidence_type", "") if tx.get("evidence_count") else "",
                    "증빙개수": tx.get("evidence_count", 0),
                    "첨부열기": ("열기", tx.get("evidence_zip_path")) if tx.get("evidence_zip_path") else "",
                }
            )
        linked_rows.sort(key=_evidence_linked_sort_key)
        _write_table_sheet(
            ws2,
            ["거래번호", "거래일시", "거래처", "금액", "증빙상태", "대표증빙종류", "증빙개수", "첨부열기"],
            linked_rows,
        )

        ws3 = wb.create_sheet("증빙요약")
        summary = {}
        for tx in snapshot.transactions:
            key = (
                tx.get("evidence_status_label", ""),
                tx.get("representative_evidence_type", "") if tx.get("evidence_count") else "미첨부",
            )
            summary[key] = summary.get(key, 0) + 1
        rows3 = []
        for (status, ev_type), count in sorted(summary.items()):
            rows3.append({"증빙상태": status, "증빙종류": ev_type, "개수": count})
        _write_table_sheet(ws3, ["증빙상태", "증빙종류", "개수"], rows3)
        return wb

    return _workbook_bytes(build)


def _build_review_workbook(snapshot: PackageSnapshot, profile: TaxPackageProfile | None = None) -> bytes:
    profile = profile or get_tax_package_profile(DEFAULT_TAX_PACKAGE_PROFILE_CODE)

    def build() -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "세무사_확인필요목록"
        prioritized_rows = []
        for row in snapshot.review_items:
            priority_order, priority_label, priority_reason = _review_priority_profile_for_package(
                str(row.get("항목유형", "")),
                profile,
            )
            has_explicit_priority = row.get("우선확인순서") not in (None, "")
            prioritized_rows.append(
                {
                    "항목번호": row.get("항목번호", ""),
                    "우선확인순서": row.get("우선확인순서", priority_order) if has_explicit_priority else priority_order,
                    "우선순위": row.get("우선순위", priority_label) if has_explicit_priority else priority_label,
                    "우선순위 기준": row.get("우선순위기준", priority_reason) if has_explicit_priority else priority_reason,
                    "항목유형": row.get("항목유형", ""),
                    "관련자료구분": row.get("관련자료구분", ""),
                    "관련번호": row.get("관련번호", ""),
                    "요약설명": row.get("요약설명", ""),
                    "현재상태": row.get("현재상태", ""),
                    "필요한확인내용": row.get("필요한확인내용", ""),
                    "메모": row.get("메모", ""),
                }
            )
        prioritized_rows.sort(key=lambda row: _review_item_sort_key(row, profile))
        _write_table_sheet(
            ws,
            ["항목번호", "우선확인순서", "우선순위", "요약설명", "현재상태", "필요한확인내용", "항목유형", "관련자료구분", "관련번호", "우선순위 기준", "메모"],
            prioritized_rows,
        )

        ws2 = wb.create_sheet("증빙누락")
        _write_table_sheet(
            ws2,
            ["거래번호", "거래일시", "거래처", "금액", "증빙상태", "필요한확인내용", "우선순위"],
            snapshot.evidence_missing_items,
        )

        ws3 = wb.create_sheet("검토필요거래")
        _write_table_sheet(
            ws3,
            ["거래번호", "거래일시", "자료출처", "거래처", "금액", "현재상태", "재확인사유", "필요한확인내용"],
            snapshot.review_trade_items,
        )
        return wb

    return _workbook_bytes(build)


def _build_business_status_workbook(snapshot: PackageSnapshot) -> bytes:
    def build() -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "사업 상태 요약"
        source = (snapshot.business_status_rows or [{}])[0]

        def inferred_source(key: str, default_source: str) -> str:
            explicit = str(source.get(f"{key}_source", "") or "").strip()
            if explicit:
                return explicit
            value = str(source.get(key, "") or "").strip()
            if value and value not in {"미확인", "미연결"}:
                return default_source
            return "미확인"

        def inferred_confidence(key: str, default_confidence: str = "참고용") -> str:
            explicit = str(source.get(f"{key}_confidence", "") or "").strip()
            if explicit:
                return explicit
            value = str(source.get(key, "") or "").strip()
            if value and value not in {"미확인", "미연결"}:
                return default_confidence
            return "미확인"

        rows = [
            {
                "항목명": "사용자 유형",
                "현재 값": source.get("user_type", "미확인"),
                "값 출처": inferred_source("user_type", "사용자 입력"),
                "확인 수준": inferred_confidence("user_type"),
                "비고": source.get("note", "") if source.get("user_type") == "미확인" else "",
            },
            {
                "항목명": "건강보험 상태",
                "현재 값": source.get("health_insurance_status", "미확인"),
                "값 출처": inferred_source("health_insurance_status", "사용자 입력"),
                "확인 수준": inferred_confidence("health_insurance_status"),
                "비고": source.get("note", "") if source.get("health_insurance_status") == "미확인" else "",
            },
            {
                "항목명": "과세 상태",
                "현재 값": source.get("vat_status", "미확인"),
                "값 출처": inferred_source("vat_status", "사용자 입력"),
                "확인 수준": inferred_confidence("vat_status"),
                "비고": source.get("note", "") if source.get("vat_status") == "미확인" else "",
            },
            {
                "항목명": "사업자등록 유무",
                "현재 값": source.get("business_registration_status", "미확인"),
                "값 출처": inferred_source("business_registration_status", "사용자 입력"),
                "확인 수준": inferred_confidence("business_registration_status"),
                "비고": "",
            },
            {
                "항목명": "사업용 계좌 사용 여부",
                "현재 값": source.get("business_account_usage_status", "미확인"),
                "값 출처": inferred_source("business_account_usage_status", "계좌 연결 정보"),
                "확인 수준": inferred_confidence("business_account_usage_status"),
                "비고": source.get("note", ""),
            },
            {
                "항목명": "사업용 카드 사용 여부",
                "현재 값": source.get("business_card_usage_status", "미확인"),
                "값 출처": inferred_source("business_card_usage_status", "미확인"),
                "확인 수준": inferred_confidence("business_card_usage_status", "미확인"),
                "비고": "현재 구조에서 별도 확인값이 없습니다.",
            },
            {
                "항목명": "기준값 출처 요약",
                "현재 값": source.get("onboarding_basis", "미확인"),
                "값 출처": "시스템 정리",
                "확인 수준": "참고용",
                "비고": source.get("note", "") or "온보딩 입력값과 현재 연결 상태를 함께 봐 주세요.",
            },
        ]
        _write_table_sheet(
            ws,
            ["항목명", "현재 값", "값 출처", "확인 수준", "비고"],
            rows,
        )
        return wb

    return _workbook_bytes(build)


def _build_withholding_summary_workbook(snapshot: PackageSnapshot) -> bytes:
    def build() -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "원천징수·기납부세액 요약"
        source = (snapshot.withholding_summary_rows or [{}])[0]
        rows = [
            {
                "자료 종류": source.get("document_scope", "미확인"),
                "원천징수 기준 기간": source.get("withholding_period_basis", "미확인"),
                "기납부세액 기준 기간": source.get("paid_tax_period_basis", "미확인"),
                "원천징수 자료 있음": source.get("has_withholding_data", "아니오"),
                "총지급액 합계": source.get("gross_pay_total_krw", ""),
                "원천징수세액 합계": source.get("withholding_tax_total_krw", ""),
                "기납부세액 자료 있음": source.get("has_paid_tax_data", "아니오"),
                "기납부세액 합계": source.get("paid_tax_total_krw", ""),
                "지급처 참조": source.get("payer_reference", "미확인"),
                "다른 소득 있음": source.get("other_income_flag", "미확인"),
                "기준 자료": source.get("source_basis", "미확인"),
                "재확인 필요": source.get("needs_review", "아니오"),
                "비고": source.get("note", ""),
            }
        ]
        _write_table_sheet(
            ws,
            [
                "자료 종류",
                "원천징수 기준 기간",
                "기납부세액 기준 기간",
                "원천징수 자료 있음",
                "총지급액 합계",
                "원천징수세액 합계",
                "기납부세액 자료 있음",
                "기납부세액 합계",
                "지급처 참조",
                "다른 소득 있음",
                "기준 자료",
                "재확인 필요",
                "비고",
            ],
            rows,
        )
        return wb

    return _workbook_bytes(build)


def _build_vat_summary_workbook(snapshot: PackageSnapshot) -> bytes:
    def build() -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "부가세 자료 요약"
        source = (snapshot.vat_summary_rows or [{}])[0]
        rows = [
            {
                "과세 상태": source.get("vat_status", "미확인"),
                "최근 부가세 신고 여부": source.get("recent_vat_filing_status", "미확인"),
                "세금계산서 매출 합계": source.get("tax_invoice_sales_total_krw", ""),
                "세금계산서 매입 합계": source.get("tax_invoice_purchase_total_krw", ""),
                "카드 매입 합계": source.get("card_purchase_total_krw", ""),
                "현금영수증 매입 합계": source.get("cash_receipt_purchase_total_krw", ""),
                "기준 자료": source.get("source_basis", "미확인"),
                "재확인 필요": source.get("needs_review", "예"),
                "비고": source.get("note", ""),
            }
        ]
        _write_table_sheet(
            ws,
            ["과세 상태", "최근 부가세 신고 여부", "세금계산서 매출 합계", "세금계산서 매입 합계", "카드 매입 합계", "현금영수증 매입 합계", "기준 자료", "재확인 필요", "비고"],
            rows,
        )
        return wb

    return _workbook_bytes(build)


def _build_nhis_pension_summary_workbook(snapshot: PackageSnapshot) -> bytes:
    def build() -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "건보·연금 요약"
        source = (snapshot.nhis_pension_summary_rows or [{}])[0]
        rows = [
            {
                "건강보험 상태": source.get("health_insurance_status", "미확인"),
                "기준 기간": source.get("period_basis", "미확인"),
                "건강보험료 합계": source.get("nhis_total_krw", ""),
                "건강보험 자료 있음": source.get("has_nhis_data", "아니오"),
                "국민연금 납부 자료 있음": source.get("has_pension_data", "아니오"),
                "국민연금 합계": source.get("pension_total_krw", ""),
                "기준 자료": source.get("source_basis", "미확인"),
                "재확인 필요": source.get("needs_review", "예"),
                "비고": source.get("note", ""),
            }
        ]
        _write_table_sheet(
            ws,
            ["건강보험 상태", "기준 기간", "건강보험료 합계", "건강보험 자료 있음", "국민연금 납부 자료 있음", "국민연금 합계", "기준 자료", "재확인 필요", "비고"],
            rows,
        )
        return wb

    return _workbook_bytes(build)


def _build_reference_material_workbook(snapshot: PackageSnapshot, profile: TaxPackageProfile | None = None) -> bytes:
    profile = profile or get_tax_package_profile(DEFAULT_TAX_PACKAGE_PROFILE_CODE)

    def build() -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "참고자료 요약"
        rows = []
        if snapshot.reference_material_rows:
            sorted_rows = sorted(snapshot.reference_material_rows, key=lambda row: _reference_material_sort_key(row, profile))
            for row in sorted_rows:
                rows.append(
                    {
                        "참고자료 번호": row.get("reference_material_id", ""),
                        "제목": row.get("title", ""),
                        "자료 유형": row.get("reference_type", ""),
                        "기준 기간": row.get("reported_period", ""),
                        "기재 금액": row.get("reported_amount_krw", ""),
                        "연결된 공식자료 유형": row.get("linked_official_doc_type", ""),
                        "연결 상태": row.get("link_status", "참고용"),
                        "비교 기준": row.get("comparison_basis", "비교 기준 없음"),
                        "비교 대상": row.get("comparison_target", "연결 가능한 공식자료 요약값 없음"),
                        "차이 금액": row.get("difference_krw", ""),
                        "차이 설명": row.get("difference_description", ""),
                        "재확인 필요": row.get("needs_review", "예"),
                        "비고": row.get("note", ""),
                    }
                )
        else:
            rows.append(
                {
                    "참고자료 번호": "",
                    "제목": "현재 포함된 참고자료 없음",
                    "자료 유형": "",
                    "기준 기간": "",
                    "기재 금액": "",
                    "연결된 공식자료 유형": "",
                    "연결 상태": "참고용",
                    "비교 기준": "비교 기준 없음",
                    "비교 대상": "연결 가능한 공식자료 요약값 없음",
                    "차이 금액": "",
                    "차이 설명": "대상 월 기준으로 포함된 참고자료가 없습니다.",
                    "재확인 필요": "아니오",
                    "비고": "참고자료 원본은 기본 패키지에 포함하지 않습니다.",
                }
            )
        _write_table_sheet(
            ws,
            ["참고자료 번호", "제목", "자료 유형", "기준 기간", "기재 금액", "연결된 공식자료 유형", "연결 상태", "비교 기준", "비교 대상", "차이 금액", "차이 설명", "재확인 필요", "비고"],
            rows,
        )
        return wb

    return _workbook_bytes(build)


def _append_official_data_sheets(wb: Workbook, snapshot: PackageSnapshot, profile: TaxPackageProfile | None = None) -> None:
    profile = profile or get_tax_package_profile(DEFAULT_TAX_PACKAGE_PROFILE_CODE)
    ws = wb.create_sheet("공식자료목록")

    official_rows = []
    if snapshot.official_documents:
        for row in sorted(snapshot.official_documents, key=lambda item: _official_document_sort_key(item, profile)):
            official_rows.append(
                {
                    "자료번호": row.get("자료번호", ""),
                    "기관명": row.get("기관명", ""),
                    "문서종류": row.get("문서종류", ""),
                    "기준일": row.get("기준일", ""),
                    "원본파일명": row.get("원본파일명", ""),
                    "원본첨부여부": row.get("원본첨부여부", "아니오"),
                    "읽기상태": row.get("읽기상태", ""),
                    "검증상태": row.get("검증상태", ""),
                    "구조확인": row.get("구조확인", ""),
                    "신뢰등급": row.get("신뢰등급", ""),
                    "핵심값요약": row.get("핵심값요약", ""),
                    "교차검증 상태": row.get("교차검증상태", "비교 불가"),
                    "교차검증 사유": row.get("교차검증사유", "비교 가능한 공식자료 범위가 아닙니다."),
                    "교차검증 재확인 필요": row.get("교차검증재확인필요여부", "아니오"),
                    "목록반영여부": row.get("목록반영여부", "예"),
                    "재확인필요여부": row.get("재확인필요여부", "예"),
                    "메모": row.get("메모", ""),
                }
            )
    else:
        official_rows.append(
            {
                "자료번호": "",
                "기관명": "",
                "문서종류": "현재 포함된 공식자료 없음",
                "기준일": "",
                "원본파일명": "",
                "원본첨부여부": "아니오",
                "읽기상태": "자료 없음",
                "검증상태": "검증 미실시",
                "구조확인": "구조 미확인",
                "신뢰등급": "반영 보류",
                "핵심값요약": "",
                "교차검증 상태": "비교 불가",
                "교차검증 사유": "대상 월 기준으로 포함할 공식자료가 없습니다.",
                "교차검증 재확인 필요": "아니오",
                "목록반영여부": "아니오",
                "재확인필요여부": "아니오",
                "메모": "대상 월 기준으로 포함할 공식자료가 없습니다.",
            }
        )

    _write_table_sheet(
        ws,
        ["자료번호", "기관명", "문서종류", "기준일", "원본파일명", "원본첨부여부", "읽기상태", "검증상태", "구조확인", "신뢰등급", "핵심값요약", "교차검증 상태", "교차검증 사유", "교차검증 재확인 필요", "목록반영여부", "재확인필요여부", "메모"],
        official_rows,
    )

    ws2 = wb.create_sheet("공식자료상태요약")
    summary_rows = []
    if snapshot.official_documents:
        buckets: dict[str, dict[str, int]] = {}
        for row in snapshot.official_documents:
            key = row.get("문서종류", "문서 판별 전")
            bucket = buckets.setdefault(
                key,
                {
                    "개수": 0,
                    "읽기 가능 건수": 0,
                    "검토 필요 건수": 0,
                    "미지원 건수": 0,
                    "읽기 실패 건수": 0,
                    "교차검증 일치 건수": 0,
                    "교차검증 부분 일치 건수": 0,
                    "교차검증 재확인 필요 건수": 0,
                    "교차검증 불일치 건수": 0,
                    "교차검증 비교 불가 건수": 0,
                },
            )
            bucket["개수"] += 1
            status = row.get("읽기상태")
            if status == "반영 가능":
                bucket["읽기 가능 건수"] += 1
            elif status == "검토 필요":
                bucket["검토 필요 건수"] += 1
            elif status == "미지원 형식":
                bucket["미지원 건수"] += 1
            else:
                bucket["읽기 실패 건수"] += 1

            cross_validation_status = row.get("교차검증상태")
            if cross_validation_status == "일치":
                bucket["교차검증 일치 건수"] += 1
            elif cross_validation_status == "부분 일치":
                bucket["교차검증 부분 일치 건수"] += 1
            elif cross_validation_status == "재확인 필요":
                bucket["교차검증 재확인 필요 건수"] += 1
            elif cross_validation_status == "불일치":
                bucket["교차검증 불일치 건수"] += 1
            else:
                bucket["교차검증 비교 불가 건수"] += 1
        for key, bucket in sorted(buckets.items()):
            summary_rows.append({"문서종류": key, **bucket})
        summary_rows.sort(
            key=lambda row: (
                -int(row.get("교차검증 불일치 건수", 0) or 0),
                -int(row.get("교차검증 재확인 필요 건수", 0) or 0),
                -int(row.get("검토 필요 건수", 0) or 0),
                str(row.get("문서종류", "")),
            )
        )
    else:
        summary_rows.append(
            {
                "문서종류": "현재 포함된 공식자료 없음",
                "개수": 0,
                "읽기 가능 건수": 0,
                "검토 필요 건수": 0,
                "미지원 건수": 0,
                "읽기 실패 건수": 0,
                "교차검증 일치 건수": 0,
                "교차검증 부분 일치 건수": 0,
                "교차검증 재확인 필요 건수": 0,
                "교차검증 불일치 건수": 0,
                "교차검증 비교 불가 건수": 0,
            }
        )
    _write_table_sheet(
        ws2,
        ["문서종류", "개수", "읽기 가능 건수", "검토 필요 건수", "미지원 건수", "읽기 실패 건수", "교차검증 일치 건수", "교차검증 부분 일치 건수", "교차검증 재확인 필요 건수", "교차검증 불일치 건수", "교차검증 비교 불가 건수"],
        summary_rows,
    )

    ws3 = wb.create_sheet("공식자료핵심값")
    flattened_rows = []
    for row in snapshot.official_documents:
        summary_items = row.get("_summary_items") or []
        for item in summary_items:
            label = (item.get("label") or "").strip()
            value = (item.get("value") or "").strip()
            if not label or not value:
                continue
            flattened_rows.append(
                {
                    "자료번호": row.get("자료번호", ""),
                    "문서종류": row.get("문서종류", ""),
                    "기관명": row.get("기관명", ""),
                    "핵심항목": label,
                    "값": value,
                }
            )
    if not flattened_rows:
        flattened_rows.append(
            {
                "자료번호": "",
                "문서종류": "현재 추출된 공식자료 핵심값 없음",
                "기관명": "",
                "핵심항목": "",
                "값": "",
            }
        )
    _write_table_sheet(ws3, ["자료번호", "문서종류", "기관명", "핵심항목", "값"], flattened_rows)


def _build_attachment_index_rows(snapshot: PackageSnapshot) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for evidence in snapshot.evidences:
        zip_path = evidence.get("_zip_path") or ""
        rows.append(
            {
                "attachment_index_key": f"evidence-{evidence.get('증빙번호') or evidence.get('연결거래번호') or len(rows) + 1}",
                "display_file_name": evidence.get("파일명", ""),
                "document_type": evidence.get("증빙종류", "증빙자료"),
                "related_transaction_id": evidence.get("연결거래번호", ""),
                "period_basis": str(evidence.get("거래일시", ""))[:10],
                "contains_sensitive_info": "낮음",
                "package_status": "포함",
                "relative_path": zip_path,
                "file_open_link": ("열기", zip_path) if zip_path else "",
                "note": "거래와 연결된 대표 첨부입니다." if zip_path else "",
            }
        )

    for document in snapshot.official_documents:
        rows.append(
            {
                "attachment_index_key": document.get("_attachment_index_key", f"official-{len(rows) + 1}"),
                "display_file_name": document.get("원본파일명", ""),
                "document_type": f"{document.get('문서종류', '공식자료')} 원본",
                "related_transaction_id": "",
                "period_basis": document.get("_period_basis", document.get("기준일", "")),
                "contains_sensitive_info": "높음",
                "package_status": "기본 제외",
                "relative_path": "",
                "file_open_link": "",
                "note": "원본 공식자료는 기본 패키지에 포함하지 않고 목록/상태/핵심값만 제공합니다.",
            }
        )

    for material in snapshot.reference_material_rows:
        rows.append(
            {
                "attachment_index_key": material.get("_attachment_index_key", f"reference-{len(rows) + 1}"),
                "display_file_name": material.get("_original_filename", ""),
                "document_type": f"{material.get('reference_type', '참고자료')} 원본",
                "related_transaction_id": "",
                "period_basis": material.get("_period_basis", material.get("reported_period", "")),
                "contains_sensitive_info": "중간",
                "package_status": "기본 제외",
                "relative_path": "",
                "file_open_link": "",
                "note": "참고자료 원본은 기본 패키지에 포함하지 않고 요약값만 제공합니다.",
            }
        )

    return rows


def _build_attachment_index_workbook(snapshot: PackageSnapshot) -> bytes:
    def build() -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "첨부인덱스"
        rows = []
        for row in sorted(_build_attachment_index_rows(snapshot), key=_attachment_index_sort_key):
            rows.append(
                {
                    "첨부 인덱스 키": row.get("attachment_index_key", ""),
                    "표시 파일명": row.get("display_file_name", ""),
                    "자료 유형": row.get("document_type", ""),
                    "관련 거래번호": row.get("related_transaction_id", ""),
                    "기준 기간": row.get("period_basis", ""),
                    "민감정보 가능성": row.get("contains_sensitive_info", ""),
                    "패키지 포함 상태": row.get("package_status", ""),
                    "상대경로": row.get("relative_path", ""),
                    "파일 열기": row.get("file_open_link", ""),
                    "비고": row.get("note", ""),
                }
            )
        _write_table_sheet(
            ws,
            ["첨부 인덱스 키", "표시 파일명", "자료 유형", "관련 거래번호", "기준 기간", "민감정보 가능성", "패키지 포함 상태", "상대경로", "파일 열기", "비고"],
            rows,
        )
        return wb

    return _workbook_bytes(build)


def _render_package_guide(snapshot: PackageSnapshot, profile: TaxPackageProfile | None = None) -> str:
    profile = profile or get_tax_package_profile(DEFAULT_TAX_PACKAGE_PROFILE_CODE)
    stats = snapshot.stats
    included_file_lines = {
        "summary": "- 00_패키지요약.xlsx : 전체 요약, 패키지안내, 공식자료 목록/상태/핵심값, 교차검증 v1 요약",
        "business_status": "- 01_사업_상태_요약.xlsx : 사용자 유형/건보/과세 상태와 각 값의 출처/확인 수준 요약",
        "transactions": "- 03_거래원장.xlsx : 거래 목록, 원본 메타, 분류/증빙 연결",
        "evidence": "- 04_증빙상태표.xlsx : 첨부된 증빙 목록과 거래별 대표 첨부 연결",
        "withholding": "- 05_원천징수_기납부세액_요약.xlsx : 원천징수 자료 구분, 기준 기간, 지급처 참조, 주요 합계와 재확인 포인트 요약",
        "review": "- 06_세무사_확인필요목록.xlsx : 우선확인순서 기준 재확인 목록",
        "attachments": "- 07_첨부인덱스.xlsx : 패키지 첨부 전체 인덱스와 상대경로 링크",
        "vat": "- 08_부가세_자료_요약.xlsx : 과세 상태, 부가세 신고 여부, 매입/매출 관련 요약",
        "nhis_pension": "- 09_건보_연금_요약.xlsx : 건강보험 상태, 건보/연금 자료 존재 여부와 합계 요약",
        "reference": "- 10_참고자료_요약.xlsx : 공식자료 우선, 거래 합계 보조 기준으로 참고자료 비교/차이 설명 요약",
    }
    lines = [
        "[쓸수있어(SafeToSpend) 세무사 패키지 v2 4차]",
        f"- 패키지 버전: {PACKAGE_VERSION}",
        f"- 대상 기간: {stats.period_start_kst} ~ {stats.period_end_kst}",
        f"- 생성 시각(KST): {stats.generated_at_kst}",
        f"- 사용자명: {snapshot.display_name}",
        f"- 패키지 유형: {profile.display_name}",
        f"- 검토 관점: {profile.event_summary_note}",
        "",
        "[포함 파일]",
        *[included_file_lines[key] for key in profile.workbook_badge_order if key in profile.included_workbooks],
        "- attachments/evidence/ : 현재 연결된 대표 증빙 파일",
        "",
        "[권장 읽는 순서]",
        *profile.review_flow_lines,
        "",
        "[현재 포함되는 자료 범위]",
        "- 수동입력 거래",
        "- 수동업로드(CSV) 거래",
        "- 자동연동 거래",
        "- 거래에 연결된 대표 증빙 1개",
        "- 기준일 또는 업로드 시각 기준으로 대상 월에 포함된 공식자료 목록/상태/핵심 추출값",
        "- 교차검증 v1 기준의 공식자료 요약 수치와 재확인 포인트",
        "- 기준 기간에 업로드된 참고자료의 제목/유형/보조 설명 요약",
        "- 누락/검토 필요 상태",
        "",
        "[기본 제외되는 자료 범위]",
        "- 공식자료 원본 파일",
        "- NHIS 원본, 건강 상세 원문, 가족관계/공제 원본처럼 민감정보 포함 가능성이 있는 원본",
        "- 참고자료 원본",
        "- 추가설명 폴더",
        "",
        "[신뢰 구분 기준]",
        "- 반영됨: 구조화된 거래가 분류 완료되고 필요한 증빙 상태가 정리된 항목",
        "- 참고용: 수동입력처럼 구조화 근거가 상대적으로 약한 항목",
        "- 재확인필요: 분류 미확정, 필수 증빙 누락, 혼합 판단 등 추가 확인이 필요한 항목",
        "- 공식자료의 검증상태/신뢰등급은 현재 main 저장 메타를 그대로 보수적으로 전달합니다.",
        "- 교차검증 상태는 비교 가능한 공식자료만 대상으로 한 교차검증 v1 결과이며 자동 확정값이 아닙니다.",
        "",
        "[현재 한계]",
        "- 거래당 대표 증빙 1개 기준으로 정리됩니다. 전체 첨부 접근은 07_첨부인덱스.xlsx를 기준으로 봐 주세요.",
        f"- 공식자료는 {stats.official_data_total}건이 목록에 반영될 수 있지만, 교차검증은 v1 기준으로 비교 가능한 공식자료만 집계합니다.",
        "- 참고자료는 공식자료를 대체하지 않고, 공식자료 우선·거래 합계 보조 기준의 비교 설명 자료로만 요약합니다.",
        "- 사업 상태 요약의 사용자 입력값은 세무사 확인 전까지 참고용으로 봐 주세요.",
        "- 공식자료의 검증상태가 '검증 미실시'이면 세무사 확인이 필요합니다.",
        "- ZIP 내부 링크는 압축을 푼 뒤 여는 방식이 가장 안정적입니다.",
        "- 엑셀 링크가 열리지 않으면 07_첨부인덱스.xlsx의 상대경로를 기준으로 파일 위치를 먼저 확인해 주세요.",
    ]
    return "\n".join(lines) + "\n"


def _build_workbook_bytes_by_key(snapshot: PackageSnapshot, profile: TaxPackageProfile) -> dict[str, bytes]:
    return {
        "summary": _build_summary_workbook(snapshot, profile),
        "business_status": _build_business_status_workbook(snapshot),
        "transactions": _build_transactions_workbook(snapshot),
        "evidence": _build_evidence_workbook(snapshot),
        "withholding": _build_withholding_summary_workbook(snapshot),
        "review": _build_review_workbook(snapshot, profile),
        "attachments": _build_attachment_index_workbook(snapshot),
        "vat": _build_vat_summary_workbook(snapshot),
        "nhis_pension": _build_nhis_pension_summary_workbook(snapshot),
        "reference": _build_reference_material_workbook(snapshot, profile),
    }


def build_tax_package_zip_from_snapshot(
    snapshot: PackageSnapshot,
    profile_code: str | None = None,
) -> tuple[io.BytesIO, str]:
    profile = get_tax_package_profile(profile_code)
    root_name, download_name = _build_profiled_package_names(snapshot, profile)
    workbook_bytes = _build_workbook_bytes_by_key(snapshot, profile)
    out = io.BytesIO()
    with zipfile.ZipFile(out, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        root = root_name
        for workbook_key in ALL_WORKBOOK_KEYS:
            if workbook_key not in profile.included_workbooks:
                continue
            zf.writestr(f"{root}/{_workbook_filename(workbook_key)}", workbook_bytes[workbook_key])
        zf.writestr(f"{root}/attachments/", b"")
        zf.writestr(f"{root}/{EVIDENCE_ATTACHMENTS_DIR}/", b"")

        for evidence in snapshot.evidences:
            zip_path = evidence.get("_zip_path") or ""
            abs_path = evidence.get("_abs_path")
            if not zip_path or not abs_path:
                continue
            try:
                with Path(abs_path).open("rb") as f:
                    zf.writestr(f"{root}/{zip_path}", f.read())
            except Exception:
                continue

    out.seek(0)
    return out, download_name


def build_tax_package_zip(user_pk: int, month_key: str, profile_code: str | None = None) -> tuple[io.BytesIO, str]:
    snapshot = _collect_package_snapshot(user_pk=user_pk, month_key=month_key)
    return build_tax_package_zip_from_snapshot(snapshot, profile_code=profile_code)
